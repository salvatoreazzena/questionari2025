from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DEFAULT_INPUT = Path("questionari_sardi_sottocampioni.xlsx")
DEFAULT_OUTPUT = Path("analisi_spese_sardi.xlsx")
DEFAULT_SHEETS = ["campione_2", "campione_2a", "campione_2b", "campione_2c"]

NULL_TOKENS = {"", "ND", "NR", "NA", "N/D", "N.R.", "N.A.", "NON DISPONIBILE", "NULL"}
UNDEFINED_LABEL = "NON DEFINITO"
DEST_SARDEGNA = "SARDI IN VACANZA IN SARDEGNA"
DEST_ITALIA = "SARDI IN VACANZA IN ALTRE REGIONI ITALIANE"
DEST_ESTERO = "SARDI IN VACANZA ALL'ESTERO"

PROVINCIA_WEIGHT_MAP = {
    "CA": 0.83,
    "SS": 0.5,
    "OT": 1.53,
    "OR": 8.53,
    "NU": 1.02,
    "CI": 12.08,
    "SU": 12.08,
    "VS": 30.32,
    "OG": 5.48,
}

SPEND_COLUMNS = [
    "spese_trasporto_interno",
    "spese_alloggio",
    "spese_alimentazione",
    "spese_ristorazione",
    "spese_souvenir",
    "spese_altre",
]

REQUIRED_COLUMNS = [
    "provincia_provenienza",
    "comune_provenienza",
    "si_dove",
    "località_sardegna",
    "durata_soggiorno",
    "numero_componenti",
    "motivazione_principale",
    "web",
    *SPEND_COLUMNS,
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Costruisce l'analisi STP/STPG pesata delle spese dei sardi.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="File Excel di input.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="File Excel di output.")
    parser.add_argument(
        "--sheets",
        nargs="*",
        default=DEFAULT_SHEETS,
        help="Foglio o lista di fogli da elaborare. Default: campione_2 campione_2a campione_2b campione_2c.",
    )
    return parser.parse_args()


def normalize_text_series(series: pd.Series) -> pd.Series:
    s = series.astype("string").fillna("").str.strip()
    mask_null = s.str.upper().isin({token.upper() for token in NULL_TOKENS})
    return s.mask(mask_null, "")


def safe_divide(numerator: float, denominator: float) -> float:
    if denominator <= 0:
        return 0.0
    return round(float(numerator) / float(denominator), 4)


def uses_web_for_booking(value: object) -> bool:
    if pd.isna(value):
        return False
    text = str(value).strip().upper()
    if text in {"", "NO", "N", "0", "FALSE"}:
        return False
    return True


def classify_duration(value: object) -> str:
    if pd.isna(value):
        return ""
    try:
        duration = float(value)
    except (TypeError, ValueError):
        return ""
    if duration <= 0:
        return ""
    if duration <= 3:
        return "1-3"
    if duration <= 7:
        return "4-7"
    if duration <= 14:
        return "8-14"
    if duration <= 30:
        return "15-30"
    return ""


def build_destination_category(where_series: pd.Series) -> pd.Series:
    where = normalize_text_series(where_series).str.upper()
    out = pd.Series(UNDEFINED_LABEL, index=where.index, dtype="string")
    out = out.mask(where.eq("IN SARDEGNA"), DEST_SARDEGNA)
    out = out.mask(where.eq("IN ITALIA"), DEST_ITALIA)
    out = out.mask(where.eq("ALL'ESTERO"), DEST_ESTERO)
    return out


def normalize_localita_sardegna_prevalente(series: pd.Series) -> pd.Series:
    s = normalize_text_series(series)
    s = s.str.replace(r"\s*\([^)]*\)\s*", " ", regex=True)
    s = s.str.split(r"[;/]", n=1, regex=True).str[0].str.strip()
    s = s.str.split(",", n=1).str[0].str.strip()
    s = s.str.replace(r"(?i)\bS\.\s*TEODORO\b", "San Teodoro", regex=True)
    s = s.str.replace(r"(?i)\bS\.\s*TERESA\b", "Santa Teresa di Gallura", regex=True)
    s = s.str.replace(r"\s+", " ", regex=True)
    return s.astype("string").str.upper().fillna("")


def normalize_comune_residenza(series: pd.Series) -> pd.Series:
    s = normalize_text_series(series)
    s = s.str.split(",", n=1).str[0].str.strip()
    s = s.str.replace(r"\s+", " ", regex=True)
    return s.astype("string").str.upper().fillna("")


def compute_weighted_metrics(df: pd.DataFrame) -> dict[str, object]:
    component_sum = float(df["componenti_pesati"].sum())
    person_days_sum = float(df["person_days_pesati"].sum())
    category_sums = {col: float(df[col].sum()) for col in SPEND_COLUMNS}
    total_spend = float(sum(category_sums.values()))
    return {
        "stp_totale": safe_divide(total_spend, component_sum),
        "stpg_totale": safe_divide(total_spend, person_days_sum),
        "stp_categorie": {col: safe_divide(value, component_sum) for col, value in category_sums.items()},
        "stpg_categorie": {col: safe_divide(value, person_days_sum) for col, value in category_sums.items()},
    }


def build_metric_row(base: dict[str, object], metrics: dict[str, object], *, daily: bool) -> dict[str, object]:
    row = dict(base)
    row["totale"] = metrics["stpg_totale"] if daily else metrics["stp_totale"]
    category_metrics = metrics["stpg_categorie"] if daily else metrics["stp_categorie"]
    for col in SPEND_COLUMNS:
        row[col] = category_metrics[col]
    return row


def build_metric_tables(groups: list[tuple[dict[str, object], pd.DataFrame]]) -> list[dict[str, object]]:
    stp_rows: list[dict[str, object]] = []
    stpg_rows: list[dict[str, object]] = []

    for base, group_df in groups:
        metrics = compute_weighted_metrics(group_df)
        stp_rows.append(build_metric_row(base, metrics, daily=False))
        stpg_rows.append(build_metric_row(base, metrics, daily=True))

    shared_headers = ["totale", *SPEND_COLUMNS]
    leading_headers = [h for h in stp_rows[0].keys() if h not in shared_headers] if stp_rows else []
    headers = [*leading_headers, *shared_headers]
    return [
        {"title": "Tabella STP pesata", "headers": headers, "rows": stp_rows},
        {"title": "Tabella STPG pesata", "headers": headers, "rows": stpg_rows},
    ]


def build_total_only_tables(groups: list[tuple[dict[str, object], pd.DataFrame]]) -> list[dict[str, object]]:
    stp_rows: list[dict[str, object]] = []
    stpg_rows: list[dict[str, object]] = []

    for base, group_df in groups:
        metrics = compute_weighted_metrics(group_df)
        stp_row = dict(base)
        stpg_row = dict(base)
        stp_row["totale"] = metrics["stp_totale"]
        stpg_row["totale"] = metrics["stpg_totale"]
        stp_rows.append(stp_row)
        stpg_rows.append(stpg_row)

    headers = [*stp_rows[0].keys()] if stp_rows else ["totale"]
    return [
        {"title": "Tabella STP pesata", "headers": headers, "rows": stp_rows},
        {"title": "Tabella STPG pesata", "headers": headers, "rows": stpg_rows},
    ]


def summarize_dimension(df: pd.DataFrame, dimension_column: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    values = [v for v in df[dimension_column].dropna().unique().tolist() if str(v).strip() != ""]
    for value in sorted(values):
        subset = df[df[dimension_column] == value].copy()
        rows.append({"valore": value, "metrics": compute_weighted_metrics(subset)})
    return pd.DataFrame(rows)


def build_ranking_table(
    ranked: pd.DataFrame,
    *,
    ranking_label: str,
    dimension: str,
    daily: bool,
) -> dict[str, object]:
    rows: list[dict[str, object]] = []
    for _, r in ranked.iterrows():
        base = {
            "classifica": ranking_label,
            "dimensione": dimension,
            "valore": r["valore"],
        }
        rows.append(build_metric_row(base, r["metrics"], daily=daily))

    return {
        "title": "Tabella STPG pesata" if daily else "Tabella STP pesata",
        "headers": ["classifica", "dimensione", "valore", "totale", *SPEND_COLUMNS],
        "rows": rows,
    }


def rank_top_dimension(df: pd.DataFrame, dimension_column: str, metric_key: str) -> pd.DataFrame:
    summarized = summarize_dimension(df, dimension_column)
    if summarized.empty:
        return summarized
    return summarized.sort_values(
        by="metrics",
        key=lambda s: s.map(lambda m: m[metric_key]),
        ascending=False,
    ).head(5)


def build_sections_campione_2(df: pd.DataFrame) -> list[dict[str, object]]:
    sections: list[dict[str, object]] = []

    destination_groups = [
        ({"segmento": "DESTINAZIONE", "gruppo": DEST_SARDEGNA}, df[df["categoria_destinazione"] == DEST_SARDEGNA].copy()),
        ({"segmento": "DESTINAZIONE", "gruppo": DEST_ITALIA}, df[df["categoria_destinazione"] == DEST_ITALIA].copy()),
        ({"segmento": "DESTINAZIONE", "gruppo": DEST_ESTERO}, df[df["categoria_destinazione"] == DEST_ESTERO].copy()),
        ({"segmento": "DESTINAZIONE", "gruppo": "TOTALE"}, df.copy()),
    ]
    sections.append(
        {
            "title": "1. STP e STPG totale e per categoria per sardi in vacanza in Sardegna, in Italia, all'estero e totale",
            "tables": build_metric_tables(destination_groups),
        }
    )

    web_groups = []
    for label, subset in [
        ("USA_WEB", df[df["usa_web"]].copy()),
        ("TOTALE", df.copy()),
    ]:
        web_groups.extend(
            [
                ({"segmento": label, "gruppo": DEST_SARDEGNA}, subset[subset["categoria_destinazione"] == DEST_SARDEGNA].copy()),
                ({"segmento": label, "gruppo": DEST_ITALIA}, subset[subset["categoria_destinazione"] == DEST_ITALIA].copy()),
                ({"segmento": label, "gruppo": DEST_ESTERO}, subset[subset["categoria_destinazione"] == DEST_ESTERO].copy()),
                ({"segmento": label, "gruppo": "TOTALE"}, subset.copy()),
            ]
        )
    sections.append(
        {
            "title": "2. STP e STPG totale e per categoria per chi usa il web per prenotare in Sardegna, in Italia, all'estero e totale",
            "tables": build_metric_tables(web_groups),
        }
    )

    sardegna = df[df["categoria_destinazione"] == DEST_SARDEGNA].copy()
    top_dest_sardegna = (
        sardegna.groupby("destinazione_prevalente", as_index=False)
        .agg(questionari=("destinazione_prevalente", "size"))
        .query("destinazione_prevalente != ''")
        .sort_values(by=["questionari", "destinazione_prevalente"], ascending=[False, True], kind="mergesort")
        .head(5)["destinazione_prevalente"]
        .tolist()
    )
    sections.append(
        {
            "title": "3. STP e STPG totale e per categoria delle top 5 destinazioni prevalenti di chi fa vacanza in Sardegna",
            "tables": build_metric_tables(
                [({"destinazione_prevalente": dest}, sardegna[sardegna["destinazione_prevalente"] == dest].copy()) for dest in top_dest_sardegna]
            ),
        }
    )

    province_stp = rank_top_dimension(df, "provincia_provenienza_norm", "stp_totale")
    province_stpg = rank_top_dimension(df, "provincia_provenienza_norm", "stpg_totale")
    sections.append(
        {
            "title": "4. STP e STPG totale e per categoria delle 5 province con valori totali piu elevati",
            "tables": [
                build_ranking_table(province_stp, ranking_label="TOP_5_STP_TOTALE", dimension="provincia_residenza", daily=False),
                build_ranking_table(province_stpg, ranking_label="TOP_5_STPG_TOTALE", dimension="provincia_residenza", daily=True),
            ],
        }
    )

    comuni_stp = rank_top_dimension(df, "comune_provenienza_norm", "stp_totale")
    comuni_stpg = rank_top_dimension(df, "comune_provenienza_norm", "stpg_totale")
    sections.append(
        {
            "title": "5. STP e STPG totale e per categoria dei 5 comuni di residenza con valori totali piu elevati",
            "tables": [
                build_ranking_table(comuni_stp, ranking_label="TOP_5_STP_TOTALE", dimension="comune_residenza", daily=False),
                build_ranking_table(comuni_stpg, ranking_label="TOP_5_STPG_TOTALE", dimension="comune_residenza", daily=True),
            ],
        }
    )

    motivations = sorted(v for v in df["motivazione_principale_norm"].dropna().unique().tolist() if v != "")
    motivation_groups = [
        ({"motivazione_principale": motivation}, df[df["motivazione_principale_norm"] == motivation].copy())
        for motivation in motivations
    ]
    sections.append(
        {
            "title": "6. STP e STPG totale e per categoria in funzione di tutte le motivazioni principali",
            "tables": build_metric_tables(motivation_groups),
        }
    )
    return sections


def build_sections_campione_2a(df: pd.DataFrame) -> list[dict[str, object]]:
    motivations = sorted(v for v in df["motivazione_principale_norm"].dropna().unique().tolist() if v != "")
    motivation_groups = [
        ({"motivazione_principale": motivation}, df[df["motivazione_principale_norm"] == motivation].copy())
        for motivation in motivations
    ]
    return [
        {
            "title": "1. STP e STPG totale e per categoria in funzione di tutte le motivazioni principali",
            "tables": build_metric_tables(motivation_groups),
        }
    ]


def build_duration_total_groups(df: pd.DataFrame, segment_label: str) -> list[tuple[dict[str, object], pd.DataFrame]]:
    classes = [v for v in ["1-3", "4-7", "8-14", "15-30"] if v in set(df["classe_durata_soggiorno"].dropna())]
    groups: list[tuple[dict[str, object], pd.DataFrame]] = []
    for duration_class in classes:
        subset = df[df["classe_durata_soggiorno"] == duration_class].copy()
        groups.append(({"segmento": segment_label, "classe_durata_soggiorno": duration_class}, subset))
    return groups


def build_sections_campione_2b(df: pd.DataFrame) -> list[dict[str, object]]:
    sardegna = df[df["categoria_destinazione"] == DEST_SARDEGNA].copy()
    italia = df[df["categoria_destinazione"] == DEST_ITALIA].copy()
    estero = df[df["categoria_destinazione"] == DEST_ESTERO].copy()

    total_groups = (
        build_duration_total_groups(df, "TOTALE")
        + build_duration_total_groups(sardegna, "SARDEGNA")
        + build_duration_total_groups(italia, "ITALIA")
        + build_duration_total_groups(estero, "ESTERO")
    )

    sardegna_category_groups: list[tuple[dict[str, object], pd.DataFrame]] = []
    classes = [v for v in ["1-3", "4-7", "8-14", "15-30"] if v in set(sardegna["classe_durata_soggiorno"].dropna())]
    for duration_class in classes:
        subset = sardegna[sardegna["classe_durata_soggiorno"] == duration_class].copy()
        sardegna_category_groups.append(({"segmento": "SARDEGNA", "classe_durata_soggiorno": duration_class}, subset))

    return [
        {
            "title": "1. STP e STPG totale in funzione delle classi di durata soggiorno per tutti, in Sardegna, in Italia e all'estero",
            "tables": build_total_only_tables(total_groups),
        },
        {
            "title": "2. STP e STPG per categoria in funzione delle classi di durata soggiorno per chi va in Sardegna",
            "tables": build_metric_tables(sardegna_category_groups),
        },
    ]


def build_sections_campione_2c(df: pd.DataFrame) -> list[dict[str, object]]:
    sections: list[dict[str, object]] = []
    motivations = sorted(v for v in df["motivazione_principale_norm"].dropna().unique().tolist() if v != "")

    motivation_groups: list[tuple[dict[str, object], pd.DataFrame]] = []
    for segment_label, subset in [
        ("TOTALE", df),
        ("SARDEGNA", df[df["categoria_destinazione"] == DEST_SARDEGNA].copy()),
        ("ITALIA", df[df["categoria_destinazione"] == DEST_ITALIA].copy()),
        ("ESTERO", df[df["categoria_destinazione"] == DEST_ESTERO].copy()),
    ]:
        for motivation in motivations:
            filtered = subset[subset["motivazione_principale_norm"] == motivation].copy()
            if filtered.empty:
                continue
            motivation_groups.append(
                (
                    {"segmento": segment_label, "motivazione_principale": motivation},
                    filtered,
                )
            )

    sections.append(
        {
            "title": "1. STP e STPG totale e per categoria per tutti, in Sardegna, in Italia e all'estero in funzione delle 5 motivazioni",
            "tables": build_metric_tables(motivation_groups),
        }
    )

    giudizio_groups: list[tuple[dict[str, object], pd.DataFrame]] = []
    judgments = sorted(v for v in df["giudizio_norm"].dropna().unique().tolist() if v != "")
    for motivation in motivations:
        for judgment in judgments:
            subset = df[
                (df["motivazione_principale_norm"] == motivation)
                & (df["giudizio_norm"] == judgment)
            ].copy()
            if subset.empty:
                continue
            giudizio_groups.append(
                (
                    {"motivazione_principale": motivation, "giudizio": judgment},
                    subset,
                )
            )

    sections.append(
        {
            "title": "2. STP e STPG totale in funzione della motivazione principale e del giudizio complessivo",
            "tables": build_total_only_tables(giudizio_groups),
        }
    )
    return sections


SECTION_BUILDERS = {
    "campione_2": build_sections_campione_2,
    "campione_2a": build_sections_campione_2a,
    "campione_2b": build_sections_campione_2b,
    "campione_2c": build_sections_campione_2c,
}


def write_sheet(ws, sections: list[dict[str, object]]) -> None:
    bold_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold_font = Font(bold=True)
    title_font = Font(bold=True, size=12)

    row_idx = 1
    ws.cell(row=row_idx, column=1, value=f"Analisi spese pesate - {ws.title}")
    row_idx += 2

    for section in sections:
        ws.cell(row=row_idx, column=1, value=section["title"]).font = title_font
        row_idx += 1

        for table in section["tables"]:
            ws.cell(row=row_idx, column=1, value=table["title"]).font = bold_font
            row_idx += 1

            headers = table["headers"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=header)
                cell.font = bold_font
                cell.fill = bold_fill
            row_idx += 1

            for data_row in table["rows"]:
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=data_row.get(header))
                row_idx += 1

            row_idx += 1

    for col_idx, width in {
        1: 30,
        2: 28,
        3: 28,
        4: 24,
        5: 18,
        6: 18,
        7: 18,
        8: 18,
        9: 18,
        10: 18,
    }.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width


def prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti nel foglio: {missing}")

    out = df.copy()
    out["provincia_provenienza_norm"] = normalize_text_series(out["provincia_provenienza"]).str.upper()
    out["comune_provenienza_norm"] = normalize_comune_residenza(out["comune_provenienza"])
    out["categoria_destinazione"] = build_destination_category(out["si_dove"])
    out["destinazione_prevalente"] = normalize_localita_sardegna_prevalente(out["località_sardegna"])
    out["motivazione_principale_norm"] = normalize_text_series(out["motivazione_principale"]).str.upper()
    out["giudizio_norm"] = normalize_text_series(out["giudizio"]).str.upper() if "giudizio" in out.columns else ""
    out["usa_web"] = out["web"].map(uses_web_for_booking)
    out["peso_provincia"] = out["provincia_provenienza_norm"].map(PROVINCIA_WEIGHT_MAP).fillna(1.0)

    out["numero_componenti_num"] = pd.to_numeric(out["numero_componenti"], errors="coerce").fillna(0.0)
    out["durata_soggiorno_num"] = pd.to_numeric(out["durata_soggiorno"], errors="coerce").fillna(0.0)
    out["classe_durata_soggiorno"] = out["durata_soggiorno_num"].map(classify_duration)
    out["componenti_pesati"] = out["numero_componenti_num"] * out["peso_provincia"]
    out["person_days_pesati"] = out["numero_componenti_num"] * out["durata_soggiorno_num"] * out["peso_provincia"]

    for col in SPEND_COLUMNS:
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0.0)

    out = out[out["categoria_destinazione"].isin([DEST_SARDEGNA, DEST_ITALIA, DEST_ESTERO])].copy()
    return out


def build_workbook(input_path: Path, output_path: Path, sheets: list[str]) -> None:
    workbook_data = pd.read_excel(input_path, sheet_name=sheets)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name in sheets:
        if sheet_name not in workbook_data:
            raise ValueError(f"Foglio non trovato nell'input: {sheet_name}")
        if sheet_name not in SECTION_BUILDERS:
            raise ValueError(f"Nessuna analisi configurata per il foglio: {sheet_name}")
        prepared = prepare_dataframe(workbook_data[sheet_name])
        sections = SECTION_BUILDERS[sheet_name](prepared)
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, sections)

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"File non trovato: {args.input}")

    requested_sheets = list(dict.fromkeys(args.sheets))
    build_workbook(args.input, args.output, requested_sheets)
    print(f"Input letto: {args.input}")
    print(f"Fogli elaborati: {', '.join(requested_sheets)}")
    print(f"Output generato: {args.output}")


if __name__ == "__main__":
    main()
