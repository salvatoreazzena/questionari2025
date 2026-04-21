from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_INPUT = Path("questionari_fonte_veri_outlier_sostituiti.xlsx")
DEFAULT_OUTPUT = Path("questionari_fonte_veri_outlier_sostituiti_imputati.xlsx")

ID_COL = "ID"
DURATION_COL = "durata_soggiorno"
PACCHETTO_COL = "pacchetto"
STATO_COL = "stato_provenienza"
MOTIV_COL = "motivazione_principale"
COMPONENTI_COL = "numero_componenti"
SPESA_PACCHETTO_COL = "spesa_pacchetto"

SPESA_NO_PACCHETTO_COLUMNS = [
    "spese_trasporto_viaggio",
    "spese_trasporto_interno",
    "spese_alloggio",
    "spese_alimentazione",
    "spese_ristorazione",
    "spese_souvenir",
    "spese_altre",
]

PACCHETTO_IMPUTABILI = {"A", "B"}
NULL_TOKENS = {"", "ND", "NR", "NA", "N/D", "N.R.", "N.A."}
PASTEL_YELLOW_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")


@dataclass(frozen=True)
class ImputationStats:
    segment: str
    spend_col: str
    nd_initial_rows: int
    nd_initial_weight: int | float
    imputed_rows: int
    imputed_weight: int | float
    nd_remaining_rows: int
    nd_remaining_weight: int | float


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_strict_empty(value: object) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def is_null_like(value: object) -> bool:
    txt = normalize_text(value).upper()
    return txt in {t.upper() for t in NULL_TOKENS}


def parse_number(value: object) -> float | None:
    if pd.isna(value):
        return None

    if isinstance(value, (int, float, np.integer, np.floating)):
        v = float(value)
        if math.isfinite(v):
            return v
        return None

    txt = normalize_text(value)
    if txt == "":
        return None
    if txt.upper() in {t.upper() for t in NULL_TOKENS}:
        return None

    txt = txt.replace("€", "").replace(" ", "")

    if re.fullmatch(r"[-+]?\d{1,3}(\.\d{3})*(,\d+)?", txt):
        txt = txt.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"[-+]?\d{1,3}(,\d{3})*(\.\d+)?", txt):
        txt = txt.replace(",", "")
    else:
        txt = txt.replace(",", ".")

    try:
        return float(txt)
    except ValueError:
        return None


def build_tourist_weight_series(components_series: pd.Series) -> pd.Series:
    weights = pd.to_numeric(components_series, errors="coerce")
    weights = weights.where(weights > 0)
    return weights.fillna(0.0)


def format_count_value(value: float) -> int | float:
    if not math.isfinite(value):
        return 0
    rounded_int = int(round(value))
    if abs(value - rounded_int) < 1e-9:
        return rounded_int
    return round(value, 2)


def weighted_count(mask: pd.Series, weights: pd.Series) -> int | float:
    mask_bool = mask.fillna(False).astype(bool)
    total = float(weights[mask_bool].sum())
    return format_count_value(total)


def build_macro_provenienza(series: pd.Series) -> pd.Series:
    s = series.astype("string").fillna("").str.strip().str.upper()
    out = pd.Series(pd.NA, index=s.index, dtype="string")
    out = out.mask(s == "ITALIA", "ITALIANI")
    out = out.mask((s != "") & (s != "ITALIA"), "STRANIERI")
    return out


def build_header_map(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if value is None:
            continue
        text = str(value).strip()
        if text == "":
            continue
        headers[text] = col_idx
    return headers


def worksheet_to_dataframe(ws: Worksheet, header_map: dict[str, int]) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    columns = list(header_map.keys())

    for row_idx in range(2, ws.max_row + 1):
        row_data: dict[str, object] = {"_excel_row": row_idx}
        for col_name in columns:
            col_idx = header_map[col_name]
            row_data[col_name] = ws.cell(row=row_idx, column=col_idx).value
        records.append(row_data)

    if not records:
        return pd.DataFrame(columns=["_excel_row", *columns])

    return pd.DataFrame(records)


def validate_columns(header_map: dict[str, int]) -> None:
    required = {
        ID_COL,
        DURATION_COL,
        PACCHETTO_COL,
        STATO_COL,
        MOTIV_COL,
        COMPONENTI_COL,
        SPESA_PACCHETTO_COL,
        *SPESA_NO_PACCHETTO_COLUMNS,
    }
    missing = [c for c in required if c not in header_map]
    if missing:
        raise ValueError(f"Colonne mancanti nel file sorgente: {sorted(missing)}")


def prepare_base_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["_durata_num"] = pd.to_numeric(out[DURATION_COL], errors="coerce")
    out["_componenti_num"] = pd.to_numeric(out[COMPONENTI_COL], errors="coerce")
    out["_turisti_weight"] = build_tourist_weight_series(out["_componenti_num"])
    out["macro_provenienza"] = build_macro_provenienza(out[STATO_COL])
    out["motivazione_grp"] = out[MOTIV_COL].astype("string").fillna("").str.strip().str.upper()
    return out


def select_donor_id(
    donor_df: pd.DataFrame,
    *,
    value_col: str,
    target_value: float,
    target_duration: float | None,
    target_components: float | None,
) -> str:
    tmp = donor_df.copy()
    tmp["_abs_to_median"] = (tmp[value_col] - target_value).abs()

    if target_duration is None:
        tmp["_abs_duration"] = np.full(len(tmp), np.inf, dtype="float64")
    else:
        tmp["_abs_duration"] = (tmp["_durata_num"] - target_duration).abs().astype("float64")
        tmp["_abs_duration"] = tmp["_abs_duration"].where(tmp["_durata_num"].notna(), np.inf)

    if target_components is None:
        tmp["_abs_componenti"] = np.full(len(tmp), np.inf, dtype="float64")
    else:
        tmp["_abs_componenti"] = (tmp["_componenti_num"] - target_components).abs().astype("float64")
        tmp["_abs_componenti"] = tmp["_abs_componenti"].where(tmp["_componenti_num"].notna(), np.inf)

    tmp["_id_text"] = tmp[ID_COL].astype("string")
    tmp = tmp.sort_values(
        by=["_abs_to_median", "_abs_duration", "_abs_componenti", "_id_text"],
        kind="mergesort",
    )

    return str(tmp.iloc[0][ID_COL])


def infer_value_from_similar(
    df: pd.DataFrame,
    *,
    row_idx: int,
    value_col: str,
) -> tuple[float | None, str | None, str | None]:
    row = df.loc[row_idx]

    levels = [
        (
            "stesso_macro_motivazione_componenti_durata",
            ["macro_provenienza", "motivazione_grp", "_componenti_num", "_durata_num"],
        ),
        ("stesso_macro_motivazione", ["macro_provenienza", "motivazione_grp"]),
        ("stesso_macro", ["macro_provenienza"]),
        ("stessa_motivazione", ["motivazione_grp"]),
        ("globale", []),
    ]

    base_pool = df[(df.index != row_idx) & df[value_col].notna()].copy()
    if base_pool.empty:
        return None, None, None

    for level_name, cols in levels:
        pool = base_pool
        skip_level = False
        for c in cols:
            target = row[c]
            if pd.isna(target):
                skip_level = True
                break
            pool = pool[pool[c] == target]
        if skip_level or pool.empty:
            continue

        median_value = float(pool[value_col].median())
        donor_id = select_donor_id(
            pool,
            value_col=value_col,
            target_value=median_value,
            target_duration=row["_durata_num"] if pd.notna(row["_durata_num"]) else None,
            target_components=row["_componenti_num"] if pd.notna(row["_componenti_num"]) else None,
        )
        return median_value, donor_id, level_name

    return None, None, None


def impute_column_for_subset(
    *,
    df: pd.DataFrame,
    ws: Worksheet,
    header_map: dict[str, int],
    subset_mask: pd.Series,
    spend_col: str,
    segment_label: str,
) -> ImputationStats:
    subset = df.loc[subset_mask].copy()
    if subset.empty:
        return ImputationStats(
            segment=segment_label,
            spend_col=spend_col,
            nd_initial_rows=0,
            nd_initial_weight=0,
            imputed_rows=0,
            imputed_weight=0,
            nd_remaining_rows=0,
            nd_remaining_weight=0,
        )

    numeric_col = f"_{spend_col}_num"
    subset[numeric_col] = subset[spend_col].map(parse_number)

    nd_mask_initial = subset[spend_col].map(is_null_like)
    nd_indices = subset.index[nd_mask_initial].tolist()
    imputed_mask = pd.Series(False, index=subset.index)

    for idx in nd_indices:
        inferred, _, _ = infer_value_from_similar(
            subset,
            row_idx=idx,
            value_col=numeric_col,
        )
        if inferred is None:
            continue

        inferred_rounded = round(float(inferred), 2)
        subset.at[idx, spend_col] = inferred_rounded
        subset.at[idx, numeric_col] = inferred_rounded
        df.at[idx, spend_col] = inferred_rounded

        excel_row = int(df.at[idx, "_excel_row"])
        col_idx = header_map[spend_col]
        cell = ws.cell(row=excel_row, column=col_idx)
        cell.value = inferred_rounded
        cell.fill = PASTEL_YELLOW_FILL

        imputed_mask.at[idx] = True

    nd_mask_remaining = subset[spend_col].map(is_null_like)
    weights = df.loc[subset.index, "_turisti_weight"]

    return ImputationStats(
        segment=segment_label,
        spend_col=spend_col,
        nd_initial_rows=int(nd_mask_initial.sum()),
        nd_initial_weight=weighted_count(nd_mask_initial, weights),
        imputed_rows=int(imputed_mask.sum()),
        imputed_weight=weighted_count(imputed_mask, weights),
        nd_remaining_rows=int(nd_mask_remaining.sum()),
        nd_remaining_weight=weighted_count(nd_mask_remaining, weights),
    )


def run(input_file: Path, output_file: Path, sheet_name: str | None) -> None:
    if not input_file.exists():
        raise FileNotFoundError(f"File input non trovato: {input_file}")

    wb = load_workbook(input_file)
    selected_sheet = sheet_name or wb.sheetnames[0]
    if selected_sheet not in wb.sheetnames:
        raise ValueError(f"Foglio '{selected_sheet}' non presente in {input_file}")

    ws = wb[selected_sheet]
    header_map = build_header_map(ws)
    validate_columns(header_map)

    df = worksheet_to_dataframe(ws, header_map)
    if df.empty:
        wb.save(output_file)
        print("Nessuna riga dati da elaborare.")
        print(f"Output generato: {output_file}")
        return

    df = prepare_base_dataframe(df)

    pacchetto_norm = df[PACCHETTO_COL].astype("string").fillna("").str.strip().str.upper()
    mask_non_pacchetto = (df["_durata_num"] >= 1) & df[PACCHETTO_COL].map(is_strict_empty)
    mask_pacchetto_a = (df["_durata_num"] >= 1) & (pacchetto_norm == "A")
    mask_pacchetto_b = (df["_durata_num"] >= 1) & (pacchetto_norm == "B")
    mask_pacchetto_c = (df["_durata_num"] >= 1) & (pacchetto_norm == "C")

    all_stats: list[ImputationStats] = []

    for spend_col in SPESA_NO_PACCHETTO_COLUMNS:
        stats = impute_column_for_subset(
            df=df,
            ws=ws,
            header_map=header_map,
            subset_mask=mask_non_pacchetto,
            spend_col=spend_col,
            segment_label="pernottanti_senza_pacchetto",
        )
        all_stats.append(stats)

    for package_type, mask in (("A", mask_pacchetto_a), ("B", mask_pacchetto_b)):
        stats = impute_column_for_subset(
            df=df,
            ws=ws,
            header_map=header_map,
            subset_mask=mask,
            spend_col=SPESA_PACCHETTO_COL,
            segment_label=f"pernottanti_con_pacchetto_{package_type}",
        )
        all_stats.append(stats)

    c_subset = df.loc[mask_pacchetto_c]
    c_nd_mask = c_subset[SPESA_PACCHETTO_COL].map(is_null_like) if not c_subset.empty else pd.Series(dtype=bool)
    c_weights = df.loc[c_subset.index, "_turisti_weight"] if not c_subset.empty else pd.Series(dtype=float)

    wb.save(output_file)

    print(f"Input letto: {input_file}")
    print(f"Foglio elaborato: {selected_sheet}")
    print("Regole imputazione applicate:")
    print("- non pacchetto: donor pool solo non pacchetto")
    print("- pacchetto A: donor pool solo A")
    print("- pacchetto B: donor pool solo B")
    print("- pacchetto C: non imputato")
    print("- celle imputate: sfondo giallo pastello")
    print("Dettaglio imputazione ND:")
    for s in all_stats:
        print(
            f"- {s.segment} | {s.spend_col}: "
            f"nd_iniziali_righe={s.nd_initial_rows}, "
            f"nd_iniziali_pesati={s.nd_initial_weight}, "
            f"imputati_righe={s.imputed_rows}, "
            f"imputati_pesati={s.imputed_weight}, "
            f"nd_residui_righe={s.nd_remaining_rows}, "
            f"nd_residui_pesati={s.nd_remaining_weight}"
        )

    print(
        "- pernottanti_con_pacchetto_C | spesa_pacchetto: "
        f"nd_righe_non_imputati={int(c_nd_mask.sum())}, "
        f"nd_pesati_non_imputati={weighted_count(c_nd_mask, c_weights) if not c_subset.empty else 0}"
    )
    print(f"Output generato: {output_file}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Imputazione ND post-outlier su workbook completo: "
            "non-pacchetto usa donor non-pacchetto; pacchetto A/B usa donor dello stesso tipo; "
            "pacchetto C escluso. Celle imputate evidenziate in giallo pastello."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="File Excel sorgente")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="File Excel di output")
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Nome foglio da elaborare (default: primo foglio)",
    )
    args = parser.parse_args()

    run(args.input, args.output, args.sheet)


if __name__ == "__main__":
    main()