from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_INPUT = Path("questionari_fonte.xlsx")
DEFAULT_OUTPUT = Path("questionari_fonte_veri_outlier_sostituiti.xlsx")
DEFAULT_THRESHOLDS_FILE_NO_PACCHETTO = Path("questionari_outlier_pernottanti.xlsx")
DEFAULT_THRESHOLDS_FILE_PACCHETTO = Path("questionari_outlier_pernottanti_con_pacchetto.xlsx")

ID_COL = "ID"
DURATION_COL = "durata_soggiorno"
PACCHETTO_COL = "pacchetto"
STATO_COL = "stato_provenienza"
MOTIV_COL = "motivazione_principale"
COMPONENTI_COL = "numero_componenti"
SPESA_PACCHETTO_COL = "spesa_pacchetto"

SPESA_NO_PACCHETTO_COLUMNS = [
    "spese_trasporto_viaggio",
    "spese_trasporto_interno",
    "spese_alloggio",
    "spese_alimentazione",
    "spese_ristorazione",
    "spese_souvenir",
    "spese_altre",
]

PACCHETTO_ALLOWED_VALUES = {"A", "B", "C"}
NULL_TOKENS = {"", "ND", "NR", "NA", "N/D", "N.R.", "N.A."}
THRESHOLDS_REQUIRED_COLUMNS = [
    "colonna_spesa",
    "macro_provenienza",
    "motivazione_grp",
    "p5_gruppo",
    "p95_gruppo",
]


@dataclass(frozen=True)
class ThresholdInfo:
    p5: float | None
    p95: float | None
    n: int


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_strict_empty(value: object) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def parse_number(value: object) -> float | None:
    if pd.isna(value):
        return None

    if isinstance(value, (int, float, np.integer, np.floating)):
        v = float(value)
        if math.isfinite(v):
            return v
        return None

    txt = normalize_text(value)
    if txt == "":
        return None
    if txt.upper() in {t.upper() for t in NULL_TOKENS}:
        return None

    txt = txt.replace("€", "").replace(" ", "")

    if re.fullmatch(r"[-+]?\d{1,3}(\.\d{3})*(,\d+)?", txt):
        txt = txt.replace(".", "").replace(",", ".")
    elif re.fullmatch(r"[-+]?\d{1,3}(,\d{3})*(\.\d+)?", txt):
        txt = txt.replace(",", "")
    else:
        txt = txt.replace(",", ".")

    try:
        return float(txt)
    except ValueError:
        return None


def build_macro_provenienza(series: pd.Series) -> pd.Series:
    s = series.astype("string").fillna("").str.strip().str.upper()
    out = pd.Series("", index=s.index, dtype="string")
    out = out.mask(s == "ITALIA", "ITALIANI")
    out = out.mask((s != "") & (s != "ITALIA"), "STRANIERI")
    return out


def build_tourist_weight_series(components_series: pd.Series) -> pd.Series:
    weights = pd.to_numeric(components_series, errors="coerce")
    weights = weights.where(weights > 0)
    return weights.fillna(0.0)


def normalize_group_key(value: object) -> str:
    return normalize_text(value).upper()


def weighted_quantile(values: pd.Series, weights: pd.Series, q: float) -> float | None:
    values_num = pd.to_numeric(values, errors="coerce")
    weights_num = pd.to_numeric(weights, errors="coerce")

    valid_mask = values_num.notna() & weights_num.notna() & (weights_num > 0)
    if not bool(valid_mask.any()):
        return None

    vals = values_num[valid_mask].to_numpy(dtype="float64", copy=False)
    w = weights_num[valid_mask].to_numpy(dtype="float64", copy=False)

    order = np.argsort(vals, kind="mergesort")
    vals = vals[order]
    w = w[order]

    cum_weights = np.cumsum(w)
    total_weight = float(cum_weights[-1])
    if not math.isfinite(total_weight) or total_weight <= 0:
        return None

    q_clamped = min(max(float(q), 0.0), 1.0)
    threshold_weight = q_clamped * total_weight
    idx = int(np.searchsorted(cum_weights, threshold_weight, side="left"))
    idx = min(max(idx, 0), len(vals) - 1)
    return float(vals[idx])


def compute_thresholds(
    df: pd.DataFrame,
    *,
    value_col: str,
    group_cols: list[str],
    weights_col: str,
) -> dict[tuple[str, str], ThresholdInfo]:
    thresholds: dict[tuple[str, str], ThresholdInfo] = {}
    grouped = df.groupby(group_cols, dropna=False)

    for key, g in grouped:
        value_num = pd.to_numeric(g[value_col], errors="coerce")
        valid_mask = value_num.notna()
        n = int(valid_mask.sum())

        if n == 0:
            thresholds[key] = ThresholdInfo(p5=None, p95=None, n=0)
            continue

        p5 = weighted_quantile(g[value_col], g[weights_col], 0.05)
        p95 = weighted_quantile(g[value_col], g[weights_col], 0.95)

        if p5 is None or p95 is None:
            vals = value_num[valid_mask]
            p5 = float(vals.quantile(0.05))
            p95 = float(vals.quantile(0.95))

        thresholds[key] = ThresholdInfo(p5=p5, p95=p95, n=n)

    return thresholds


def compute_daily_per_capita_series(
    amount_series: pd.Series,
    components_series: pd.Series,
    duration_series: pd.Series,
) -> pd.Series:
    amount_num = pd.to_numeric(amount_series, errors="coerce")
    components_num = pd.to_numeric(components_series, errors="coerce")
    duration_num = pd.to_numeric(duration_series, errors="coerce")

    denominator = components_num * duration_num
    per_capita = amount_num / denominator
    return per_capita.where(denominator > 0)


def get_thresholds_sheet_name(thresholds_file: Path, preferred_sheet: str | None) -> str:
    xls = pd.ExcelFile(thresholds_file)
    if preferred_sheet:
        if preferred_sheet not in xls.sheet_names:
            raise ValueError(f"Foglio '{preferred_sheet}' non presente in {thresholds_file}")
        return preferred_sheet

    if len(xls.sheet_names) < 2:
        raise ValueError(
            f"Il file {thresholds_file} non contiene una seconda pagina con la tabella soglie"
        )
    return xls.sheet_names[1]


def find_thresholds_header_row(raw_sheet_df: pd.DataFrame) -> int:
    for row_idx in range(len(raw_sheet_df)):
        row = raw_sheet_df.iloc[row_idx]
        c0 = normalize_text(row.iloc[0]).lower() if len(row) > 0 else ""
        c1 = normalize_text(row.iloc[1]).lower() if len(row) > 1 else ""
        c2 = normalize_text(row.iloc[2]).lower() if len(row) > 2 else ""
        if c0 == "colonna_spesa" and c1 == "macro_provenienza" and c2 == "motivazione_grp":
            return row_idx

    raise ValueError(
        "Tabella soglie non trovata: intestazione 'colonna_spesa, macro_provenienza, motivazione_grp' assente"
    )


def load_thresholds_from_file(
    thresholds_file: Path,
    *,
    thresholds_sheet: str | None,
) -> tuple[str, dict[str, dict[tuple[str, str], ThresholdInfo]]]:
    if not thresholds_file.exists():
        raise FileNotFoundError(f"File soglie non trovato: {thresholds_file}")

    sheet_name = get_thresholds_sheet_name(thresholds_file, thresholds_sheet)
    raw_sheet_df = pd.read_excel(thresholds_file, sheet_name=sheet_name, header=None)
    header_row = find_thresholds_header_row(raw_sheet_df)
    table_df = pd.read_excel(thresholds_file, sheet_name=sheet_name, header=header_row)
    table_df = table_df.rename(columns=lambda c: normalize_text(c))

    missing_cols = [c for c in THRESHOLDS_REQUIRED_COLUMNS if c not in table_df.columns]
    if missing_cols:
        raise ValueError(
            "Colonne mancanti nella tabella soglie del file "
            f"{thresholds_file}, foglio {sheet_name}: {missing_cols}"
        )

    thresholds_by_col: dict[str, dict[tuple[str, str], ThresholdInfo]] = {}
    for _, row in table_df.iterrows():
        spend_col = normalize_text(row["colonna_spesa"])
        if spend_col == "":
            continue

        macro = normalize_group_key(row["macro_provenienza"])
        motiv = normalize_group_key(row["motivazione_grp"])
        if macro == "" or motiv == "":
            continue

        p5 = pd.to_numeric(pd.Series([row["p5_gruppo"]]), errors="coerce").iat[0]
        p95 = pd.to_numeric(pd.Series([row["p95_gruppo"]]), errors="coerce").iat[0]
        if pd.isna(p5) or pd.isna(p95):
            continue

        thresholds_by_col.setdefault(spend_col, {})[(macro, motiv)] = ThresholdInfo(
            p5=float(p5),
            p95=float(p95),
            n=0,
        )

    return sheet_name, thresholds_by_col


def build_header_map(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if value is None:
            continue
        text = str(value).strip()
        if text == "":
            continue
        headers[text] = col_idx
    return headers


def worksheet_to_dataframe(ws: Worksheet, header_map: dict[str, int]) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    columns = list(header_map.keys())

    for row_idx in range(2, ws.max_row + 1):
        row_data: dict[str, object] = {"_excel_row": row_idx}
        for col_name in columns:
            col_idx = header_map[col_name]
            row_data[col_name] = ws.cell(row=row_idx, column=col_idx).value
        records.append(row_data)

    if not records:
        return pd.DataFrame(columns=["_excel_row", *columns])

    return pd.DataFrame(records)


def color_to_hex(color: object) -> str | None:
    if color is None:
        return None

    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and rgb.strip() != "":
        raw = rgb.strip().upper()
        return raw[-6:] if len(raw) >= 6 else None

    indexed = getattr(color, "indexed", None)
    if indexed is not None:
        try:
            idx = int(indexed)
            if 0 <= idx < len(COLOR_INDEX):
                raw_index = str(COLOR_INDEX[idx]).strip().upper()
                if raw_index:
                    return raw_index[-6:] if len(raw_index) >= 6 else None
        except (TypeError, ValueError):
            return None

    return None


def is_red_hex(hex6: str) -> bool:
    if len(hex6) != 6:
        return False

    try:
        r = int(hex6[0:2], 16)
        g = int(hex6[2:4], 16)
        b = int(hex6[4:6], 16)
    except ValueError:
        return False

    known_reds = {"FF0000", "C00000", "9C0006", "E81123"}
    if hex6 in known_reds:
        return True

    return r >= 180 and g <= 120 and b <= 120


def is_true_outlier_red_cell(ws: Worksheet, row_idx: int, col_idx: int) -> bool:
    cell = ws.cell(row=row_idx, column=col_idx)
    fill = cell.fill
    if fill is None:
        return False

    if fill.fill_type != "solid":
        return False

    candidate_hex = color_to_hex(fill.fgColor) or color_to_hex(fill.start_color)
    if candidate_hex is None:
        return False

    return is_red_hex(candidate_hex)


def validate_columns(header_map: dict[str, int]) -> None:
    required = {
        ID_COL,
        DURATION_COL,
        PACCHETTO_COL,
        STATO_COL,
        MOTIV_COL,
        COMPONENTI_COL,
        SPESA_PACCHETTO_COL,
        *SPESA_NO_PACCHETTO_COLUMNS,
    }
    missing = [c for c in required if c not in header_map]
    if missing:
        raise ValueError(f"Colonne mancanti nel file sorgente: {sorted(missing)}")


def prepare_base_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["_durata_num"] = pd.to_numeric(out[DURATION_COL], errors="coerce")
    out["_componenti_num"] = pd.to_numeric(out[COMPONENTI_COL], errors="coerce")
    out["_turisti_weight"] = build_tourist_weight_series(out["_componenti_num"])
    out["macro_provenienza"] = build_macro_provenienza(out[STATO_COL]).astype("string")
    out["motivazione_grp"] = out[MOTIV_COL].astype("string").fillna("").str.strip().str.upper()
    out["motivazione_grp"] = out["motivazione_grp"].astype("string")
    return out


def replace_red_outliers_for_columns(
    *,
    ws: Worksheet,
    subset_df: pd.DataFrame,
    spend_columns: list[str],
    thresholds_by_col: dict[str, dict[tuple[str, str], ThresholdInfo]],
    header_map: dict[str, int],
) -> dict[str, dict[str, int]]:
    stats: dict[str, dict[str, int]] = {
        col: {
            "red_cells": 0,
            "replaced_p5": 0,
            "replaced_p95": 0,
            "red_inside_range": 0,
            "red_skipped_missing_threshold": 0,
            "red_skipped_non_numeric": 0,
        }
        for col in spend_columns
    }

    for spend_col in spend_columns:
        metric_col = f"_{spend_col}_metric"
        thresholds = thresholds_by_col.get(spend_col, {})
        col_idx = header_map[spend_col]

        for idx, row in subset_df.iterrows():
            excel_row = int(row["_excel_row"])
            if not is_true_outlier_red_cell(ws, excel_row, col_idx):
                continue

            stats[spend_col]["red_cells"] += 1

            current_value = parse_number(row[spend_col])
            duration = row["_durata_num"]
            components = row["_componenti_num"]
            denominator = None

            if (
                current_value is None
                or pd.isna(duration)
                or pd.isna(components)
                or float(duration) <= 0
                or float(components) <= 0
            ):
                stats[spend_col]["red_skipped_non_numeric"] += 1
                continue

            denominator = float(duration) * float(components)
            if denominator <= 0:
                stats[spend_col]["red_skipped_non_numeric"] += 1
                continue

            metric_val = row[metric_col]
            if pd.isna(metric_val):
                metric_val = current_value / denominator

            key = (
                normalize_group_key(row["macro_provenienza"]),
                normalize_group_key(row["motivazione_grp"]),
            )
            info = thresholds.get(key, ThresholdInfo(p5=None, p95=None, n=0))

            if info.p5 is None or info.p95 is None:
                stats[spend_col]["red_skipped_missing_threshold"] += 1
                continue

            target_metric: float | None = None
            target_kind: str | None = None

            if float(metric_val) <= float(info.p5):
                target_metric = float(info.p5)
                target_kind = "p5"
            elif float(metric_val) >= float(info.p95):
                target_metric = float(info.p95)
                target_kind = "p95"
            else:
                dist_p5 = abs(float(metric_val) - float(info.p5))
                dist_p95 = abs(float(metric_val) - float(info.p95))
                if dist_p5 <= dist_p95:
                    target_metric = float(info.p5)
                    target_kind = "p5"
                else:
                    target_metric = float(info.p95)
                    target_kind = "p95"
                stats[spend_col]["red_inside_range"] += 1

            replacement_value = round(target_metric * denominator, 2)
            ws.cell(row=excel_row, column=col_idx).value = replacement_value

            if target_kind == "p5":
                stats[spend_col]["replaced_p5"] += 1
            else:
                stats[spend_col]["replaced_p95"] += 1

    return stats


def run(
    input_file: Path,
    output_file: Path,
    sheet_name: str | None,
    thresholds_file_no_pacchetto: Path,
    thresholds_sheet_no_pacchetto: str | None,
    thresholds_file_pacchetto: Path,
    thresholds_sheet_pacchetto: str | None,
) -> None:
    if not input_file.exists():
        raise FileNotFoundError(f"File input non trovato: {input_file}")

    wb = load_workbook(input_file)
    selected_sheet = sheet_name or wb.sheetnames[0]
    if selected_sheet not in wb.sheetnames:
        raise ValueError(f"Foglio '{selected_sheet}' non presente in {input_file}")

    ws = wb[selected_sheet]
    header_map = build_header_map(ws)
    validate_columns(header_map)

    df = worksheet_to_dataframe(ws, header_map)
    if df.empty:
        wb.save(output_file)
        print("Nessuna riga dati da elaborare.")
        print(f"Output generato: {output_file}")
        return

    df = prepare_base_dataframe(df)
    (
        thresholds_sheet_no_pacchetto_name,
        thresholds_no_pacchetto_from_file,
    ) = load_thresholds_from_file(
        thresholds_file_no_pacchetto,
        thresholds_sheet=thresholds_sheet_no_pacchetto,
    )
    (
        thresholds_sheet_pacchetto_name,
        thresholds_pacchetto_from_file,
    ) = load_thresholds_from_file(
        thresholds_file_pacchetto,
        thresholds_sheet=thresholds_sheet_pacchetto,
    )

    no_pacchetto_mask = (df["_durata_num"] >= 1) & df[PACCHETTO_COL].map(is_strict_empty)
    pacchetto_norm = df[PACCHETTO_COL].astype("string").fillna("").str.strip().str.upper()
    con_pacchetto_mask = (df["_durata_num"] >= 1) & pacchetto_norm.isin(PACCHETTO_ALLOWED_VALUES)

    subset_spese = df.loc[no_pacchetto_mask].copy()
    subset_pacchetto = df.loc[con_pacchetto_mask].copy()

    thresholds_spese: dict[str, dict[tuple[str, str], ThresholdInfo]] = {}
    for spend_col in SPESA_NO_PACCHETTO_COLUMNS:
        metric_col = f"_{spend_col}_metric"
        subset_spese[metric_col] = compute_daily_per_capita_series(
            subset_spese[spend_col].map(parse_number),
            subset_spese["_componenti_num"],
            subset_spese["_durata_num"],
        )
        thresholds_spese[spend_col] = thresholds_no_pacchetto_from_file.get(spend_col, {})

    thresholds_pacchetto: dict[str, dict[tuple[str, str], ThresholdInfo]] = {}
    metric_col_pacchetto = f"_{SPESA_PACCHETTO_COL}_metric"
    subset_pacchetto[metric_col_pacchetto] = compute_daily_per_capita_series(
        subset_pacchetto[SPESA_PACCHETTO_COL].map(parse_number),
        subset_pacchetto["_componenti_num"],
        subset_pacchetto["_durata_num"],
    )
    thresholds_pacchetto[SPESA_PACCHETTO_COL] = thresholds_pacchetto_from_file.get(
        SPESA_PACCHETTO_COL,
        {},
    )

    stats_spese = replace_red_outliers_for_columns(
        ws=ws,
        subset_df=subset_spese,
        spend_columns=SPESA_NO_PACCHETTO_COLUMNS,
        thresholds_by_col=thresholds_spese,
        header_map=header_map,
    )

    stats_pacchetto = replace_red_outliers_for_columns(
        ws=ws,
        subset_df=subset_pacchetto,
        spend_columns=[SPESA_PACCHETTO_COL],
        thresholds_by_col=thresholds_pacchetto,
        header_map=header_map,
    )

    wb.save(output_file)

    all_stats = {**stats_spese, **stats_pacchetto}

    print(f"Input letto: {input_file}")
    print(f"Foglio elaborato: {selected_sheet}")
    print(
        "Soglie no-pacchetto lette da: "
        f"{thresholds_file_no_pacchetto} (foglio: {thresholds_sheet_no_pacchetto_name})"
    )
    print(
        "Soglie pacchetto lette da: "
        f"{thresholds_file_pacchetto} (foglio: {thresholds_sheet_pacchetto_name})"
    )
    print(f"Output generato: {output_file}")
    print("Dettaglio sostituzioni celle rosse:")

    for spend_col in [*SPESA_NO_PACCHETTO_COLUMNS, SPESA_PACCHETTO_COL]:
        s = all_stats[spend_col]
        print(
            f"- {spend_col}: "
            f"rosse={s['red_cells']}, "
            f"sostituite_p5={s['replaced_p5']}, "
            f"sostituite_p95={s['replaced_p95']}, "
            f"rosse_interne_range={s['red_inside_range']}, "
            f"saltate_no_soglia={s['red_skipped_missing_threshold']}, "
            f"saltate_non_numeriche={s['red_skipped_non_numeric']}"
        )


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Nel file questionari fonte sostituisce solo i veri outlier (celle rosse) "
            "con la soglia corrispondente p5/p95, mantenendo invariati i valori verdi e tutti gli altri."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="File Excel sorgente")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="File Excel di output")
    parser.add_argument(
        "--thresholds-file-no-pacchetto",
        type=Path,
        default=DEFAULT_THRESHOLDS_FILE_NO_PACCHETTO,
        help="File Excel soglie per spese senza pacchetto (seconda pagina di default)",
    )
    parser.add_argument(
        "--thresholds-sheet-no-pacchetto",
        type=str,
        default=None,
        help="Nome foglio soglie no-pacchetto (default: seconda pagina del file soglie)",
    )
    parser.add_argument(
        "--thresholds-file-pacchetto",
        type=Path,
        default=DEFAULT_THRESHOLDS_FILE_PACCHETTO,
        help="File Excel soglie per spesa_pacchetto (seconda pagina di default)",
    )
    parser.add_argument(
        "--thresholds-sheet-pacchetto",
        type=str,
        default=None,
        help="Nome foglio soglie pacchetto (default: seconda pagina del file soglie)",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default=None,
        help="Nome foglio da elaborare (default: primo foglio)",
    )
    args = parser.parse_args()

    run(
        args.input,
        args.output,
        args.sheet,
        args.thresholds_file_no_pacchetto,
        args.thresholds_sheet_no_pacchetto,
        args.thresholds_file_pacchetto,
        args.thresholds_sheet_pacchetto,
    )


if __name__ == "__main__":
    main()
