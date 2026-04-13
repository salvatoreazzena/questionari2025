from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill

NULL_VALUE = "NULL"
NULL_TOKENS = {"", "ND", "NR", "NA", "N/D", "N.R.", "N.A."}
PACKAGE_CODES = {"A", "B", "C"}

SPEND_COLUMNS = [
    "spese_trasporto_viaggio",
    "spese_trasporto_interno",
    "spese_alloggio",
    "spese_alimentazione",
    "spese_ristorazione",
    "spese_souvenir",
    "spese_altre",
]

MANDATORY_COLUMNS = [
    "stato_provenienza",
    "motivazione_principale",
    "tipologia_sistemazione",
    "fascia_età",
    "durata_soggiorno",
    "numero_componenti",
    "pacchetto",
    "spesa_pacchetto",
    *SPEND_COLUMNS,
]

MATCH_KEYS = [
    "stato_provenienza",
    "motivazione_principale",
    "tipologia_sistemazione",
    "fascia_età",
    "classe_numero_componenti",
]

STRONG_MATCH_KEYS = ["stato_provenienza", "motivazione_principale"]

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")


@dataclass(frozen=True)
class Threshold:
    n_valid: int
    p5: float
    p95: float


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return NULL_VALUE
    txt = str(value).strip()
    if txt.upper() in NULL_TOKENS:
        return NULL_VALUE
    return txt.upper()


def is_blank_text(value: Any) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def parse_numeric_value(value: Any) -> tuple[float | None, bool, bool]:
    """Return numeric value, missing_token_flag, parse_error_flag."""
    if pd.isna(value):
        return None, True, False

    txt = str(value).strip()
    if txt.upper() in NULL_TOKENS:
        return None, True, False

    cleaned = txt.replace("\u20ac", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")

    try:
        return float(cleaned), False, False
    except ValueError:
        return None, False, True


def parse_numeric_series(series: pd.Series) -> tuple[pd.Series, pd.Series, pd.Series]:
    parsed: list[float | None] = []
    missing_flags: list[bool] = []
    parse_error_flags: list[bool] = []

    for value in series:
        num, is_missing, is_parse_error = parse_numeric_value(value)
        parsed.append(num)
        missing_flags.append(is_missing)
        parse_error_flags.append(is_parse_error)

    return (
        pd.Series(parsed, index=series.index, dtype="float64"),
        pd.Series(missing_flags, index=series.index, dtype="bool"),
        pd.Series(parse_error_flags, index=series.index, dtype="bool"),
    )


def build_numero_componenti_class(series_num: pd.Series) -> pd.Series:
    out = pd.Series(index=series_num.index, dtype="string")

    out = out.mask(series_num.isna(), "UNKNOWN")
    out = out.mask(series_num <= 1.5, "1")
    out = out.mask((series_num > 1.5) & (series_num <= 2.5), "2")
    out = out.mask((series_num > 2.5) & (series_num <= 4.5), "3-4")
    out = out.mask(series_num > 4.5, "5+")
    return out


def ensure_columns(df: pd.DataFrame, required: list[str]) -> None:
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti nel file sorgente: {missing}")


def _build_threshold_map(
    df: pd.DataFrame,
    value_col: str,
    group_cols: list[str],
) -> dict[Any, Threshold]:
    if not group_cols:
        values = df[value_col].dropna()
        values = values[values > 0]
        if values.empty:
            return {}
        return {
            "__GLOBAL__": Threshold(
                n_valid=int(values.shape[0]),
                p5=float(values.quantile(0.05, interpolation="linear")),
                p95=float(values.quantile(0.95, interpolation="linear")),
            )
        }

    out: dict[Any, Threshold] = {}
    grouped = df.groupby(group_cols, dropna=False)
    for key, part in grouped:
        values = part[value_col].dropna()
        values = values[values > 0]
        if values.empty:
            continue
        out[key] = Threshold(
            n_valid=int(values.shape[0]),
            p5=float(values.quantile(0.05, interpolation="linear")),
            p95=float(values.quantile(0.95, interpolation="linear")),
        )
    return out


def choose_threshold(
    row: pd.Series,
    *,
    threshold_full: dict[Any, Threshold],
    threshold_prov: dict[Any, Threshold],
    threshold_motiv: dict[Any, Threshold],
    threshold_global: dict[Any, Threshold],
    min_n: int,
) -> tuple[Threshold | None, str]:
    key_full = (row["__norm_stato_provenienza"], row["__norm_motivazione_principale"])
    th = threshold_full.get(key_full)
    if th is not None and th.n_valid >= min_n:
        return th, "FULL_STATO_X_MOTIVAZIONE"

    key_prov = row["__norm_stato_provenienza"]
    th = threshold_prov.get(key_prov)
    if th is not None and th.n_valid >= min_n:
        return th, "SOLO_STATO"

    key_motiv = row["__norm_motivazione_principale"]
    th = threshold_motiv.get(key_motiv)
    if th is not None and th.n_valid >= min_n:
        return th, "SOLO_MOTIVAZIONE"

    th = threshold_global.get("__GLOBAL__")
    if th is not None:
        return th, "TOTALE_CAMPIONE"

    return None, "NESSUNA_SOGLIA"


def prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.insert(0, "id_riga", range(1, len(out) + 1))

    out["__norm_stato_provenienza"] = out["stato_provenienza"].map(normalize_text)
    out["__norm_motivazione_principale"] = out["motivazione_principale"].map(normalize_text)
    out["__norm_tipologia_sistemazione"] = out["tipologia_sistemazione"].map(normalize_text)
    out["__norm_fascia_eta"] = out["fascia_età"].map(normalize_text)

    numero_componenti_num, _, _ = parse_numeric_series(out["numero_componenti"])
    out["__numero_componenti_num"] = numero_componenti_num
    out["classe_numero_componenti"] = build_numero_componenti_class(numero_componenti_num)

    durata_soggiorno_num, _, _ = parse_numeric_series(out["durata_soggiorno"])
    out["__durata_soggiorno_num"] = durata_soggiorno_num

    denominatore = numero_componenti_num * durata_soggiorno_num
    out["__denominatore_spesa"] = denominatore

    out["__is_pernottante"] = durata_soggiorno_num.ge(1).fillna(False)
    out["__pacchetto_vuoto"] = out["pacchetto"].map(is_blank_text)
    out["__is_pernottante_senza_pacchetto"] = out["__is_pernottante"] & out["__pacchetto_vuoto"]

    pacchetto_norm = out["pacchetto"].map(normalize_text)
    out["__pacchetto_norm"] = pacchetto_norm
    spesa_pacchetto_num, _, _ = parse_numeric_series(out["spesa_pacchetto"])
    out["__spesa_pacchetto_num"] = spesa_pacchetto_num

    package_active = pacchetto_norm.isin(PACKAGE_CODES) & spesa_pacchetto_num.gt(0)
    out["__package_active"] = package_active

    for col in SPEND_COLUMNS:
        parsed, missing_mask, parse_error_mask = parse_numeric_series(out[col])
        out[f"__num_{col}"] = parsed
        out[f"__missing_{col}"] = missing_mask
        out[f"__parse_error_{col}"] = parse_error_mask

        pcg = parsed / denominatore
        pcg = pcg.mask(denominatore <= 0)
        out[f"__pcg_{col}"] = pcg

        zero_structural = package_active & parsed.eq(0)
        out[f"__zero_structural_{col}"] = zero_structural
        out[f"__eligible_outlier_{col}"] = (
            out["__is_pernottante_senza_pacchetto"] & pcg.notna() & (~zero_structural)
        )
        out[f"__eligible_threshold_{col}"] = out[f"__eligible_outlier_{col}"] & parsed.gt(0)

    return out


def build_thresholds(df: pd.DataFrame, spend_col: str) -> tuple[dict[Any, Threshold], dict[Any, Threshold], dict[Any, Threshold], dict[Any, Threshold], pd.DataFrame]:
    eligible = df[df[f"__eligible_threshold_{spend_col}"]].copy()
    value_col = f"__pcg_{spend_col}"

    threshold_full = _build_threshold_map(eligible, value_col, ["__norm_stato_provenienza", "__norm_motivazione_principale"])
    threshold_prov = _build_threshold_map(eligible, value_col, ["__norm_stato_provenienza"])
    threshold_motiv = _build_threshold_map(eligible, value_col, ["__norm_motivazione_principale"])
    threshold_global = _build_threshold_map(eligible, value_col, [])

    rows: list[dict[str, Any]] = []
    for (stato, motivazione), th in threshold_full.items():
        rows.append(
            {
                "spesa": spend_col,
                "livello": "FULL_STATO_X_MOTIVAZIONE",
                "metrica": "SPESA_PRO_CAPITE_GIORNALIERA",
                "stato_provenienza": stato,
                "motivazione_principale": motivazione,
                "n_valid": th.n_valid,
                "p5": th.p5,
                "p95": th.p95,
            }
        )
    for stato, th in threshold_prov.items():
        rows.append(
            {
                "spesa": spend_col,
                "livello": "SOLO_STATO",
                "metrica": "SPESA_PRO_CAPITE_GIORNALIERA",
                "stato_provenienza": stato,
                "motivazione_principale": "",
                "n_valid": th.n_valid,
                "p5": th.p5,
                "p95": th.p95,
            }
        )
    for motivazione, th in threshold_motiv.items():
        rows.append(
            {
                "spesa": spend_col,
                "livello": "SOLO_MOTIVAZIONE",
                "metrica": "SPESA_PRO_CAPITE_GIORNALIERA",
                "stato_provenienza": "",
                "motivazione_principale": motivazione,
                "n_valid": th.n_valid,
                "p5": th.p5,
                "p95": th.p95,
            }
        )
    global_th = threshold_global.get("__GLOBAL__")
    if global_th is not None:
        rows.append(
            {
                "spesa": spend_col,
                "livello": "TOTALE_CAMPIONE",
                "metrica": "SPESA_PRO_CAPITE_GIORNALIERA",
                "stato_provenienza": "",
                "motivazione_principale": "",
                "n_valid": global_th.n_valid,
                "p5": global_th.p5,
                "p95": global_th.p95,
            }
        )

    return threshold_full, threshold_prov, threshold_motiv, threshold_global, pd.DataFrame(rows)


def detect_outliers(df: pd.DataFrame, min_n: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    out = df.copy()
    threshold_tables: list[pd.DataFrame] = []

    for col in SPEND_COLUMNS:
        th_full, th_prov, th_motiv, th_global, th_table = build_thresholds(out, col)
        threshold_tables.append(th_table)

        pcg_values: list[float | None] = []
        p5_values: list[float | None] = []
        p95_values: list[float | None] = []
        lvl_values: list[str] = []
        outlier_types: list[str] = []

        for _, row in out.iterrows():
            value = row[f"__pcg_{col}"]
            zero_structural = bool(row[f"__zero_structural_{col}"])
            in_scope = bool(row["__is_pernottante_senza_pacchetto"])

            if not in_scope:
                pcg_values.append(float(value) if pd.notna(value) else None)
                p5_values.append(None)
                p95_values.append(None)
                lvl_values.append("FUORI_PERIMETRO")
                outlier_types.append("NON_VALUTABILE_FUORI_PERIMETRO")
                continue

            if pd.isna(value):
                pcg_values.append(None)
                p5_values.append(None)
                p95_values.append(None)
                lvl_values.append("NON_VALUTABILE")
                outlier_types.append("MISSING")
                continue

            if zero_structural:
                pcg_values.append(None)
                p5_values.append(None)
                p95_values.append(None)
                lvl_values.append("ECCEZIONE_PACCHETTO")
                outlier_types.append("ZERO_STRUTTURALE_PACCHETTO")
                continue

            threshold, level = choose_threshold(
                row,
                threshold_full=th_full,
                threshold_prov=th_prov,
                threshold_motiv=th_motiv,
                threshold_global=th_global,
                min_n=min_n,
            )

            if threshold is None:
                pcg_values.append(float(value))
                p5_values.append(None)
                p95_values.append(None)
                lvl_values.append(level)
                outlier_types.append("NON_VALUTABILE")
                continue

            pcg_values.append(float(value))
            p5_values.append(threshold.p5)
            p95_values.append(threshold.p95)
            lvl_values.append(level)

            if float(value) < threshold.p5:
                outlier_types.append("OUTLIER_BASSO")
            elif float(value) > threshold.p95:
                outlier_types.append("OUTLIER_ALTO")
            else:
                outlier_types.append("OK")

        out[f"valore_pcg_{col}"] = pcg_values
        out[f"p5_{col}"] = p5_values
        out[f"p95_{col}"] = p95_values
        out[f"livello_soglia_{col}"] = lvl_values
        out[f"tipo_outlier_{col}"] = outlier_types
        out[f"flag_outlier_{col}"] = out[f"tipo_outlier_{col}"].isin({"OUTLIER_BASSO", "OUTLIER_ALTO"})

    flag_cols = [f"flag_outlier_{col}" for col in SPEND_COLUMNS]
    out["outlier_count"] = out[flag_cols].sum(axis=1)
    out["ha_outlier"] = out["outlier_count"] > 0

    def _join_outlier_cols(row: pd.Series) -> str:
        cols = [col for col in SPEND_COLUMNS if bool(row[f"flag_outlier_{col}"])]
        return "; ".join(cols)

    out["outlier_cols"] = out.apply(_join_outlier_cols, axis=1)

    threshold_df = pd.concat(threshold_tables, ignore_index=True) if threshold_tables else pd.DataFrame()
    return out, threshold_df


def build_initial_stats(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    global_rows: list[dict[str, Any]] = []
    group_rows: list[dict[str, Any]] = []

    total_rows = len(df)

    for col in SPEND_COLUMNS:
        pcg_col = f"__pcg_{col}"
        missing_col = f"__missing_{col}"
        parse_error_col = f"__parse_error_{col}"
        zero_col = f"__zero_structural_{col}"
        eligible_col = f"__eligible_outlier_{col}"
        threshold_eligible_col = f"__eligible_threshold_{col}"
        outlier_col = f"flag_outlier_{col}"

        eligible_series = df.loc[df[threshold_eligible_col], pcg_col]

        global_rows.append(
            {
                "spesa": col,
                "n_totale": total_rows,
                "n_mancanti": int(df[missing_col].sum()),
                "n_parse_error": int(df[parse_error_col].sum()),
                "n_in_scope": int(df["__is_pernottante_senza_pacchetto"].sum()),
                "n_zero_strutturali_pacchetto": int(df[zero_col].sum()),
                "n_eligible_outlier": int(df[eligible_col].sum()),
                "n_eligible_soglie": int(df[threshold_eligible_col].sum()),
                "n_outlier": int(df[outlier_col].sum()) if outlier_col in df.columns else 0,
                "tasso_mancanti_pct": round((df[missing_col].sum() / total_rows) * 100.0, 4),
                "min": float(eligible_series.min()) if not eligible_series.empty else None,
                "p5": float(eligible_series.quantile(0.05, interpolation="linear")) if not eligible_series.empty else None,
                "mediana": float(eligible_series.median()) if not eligible_series.empty else None,
                "p95": float(eligible_series.quantile(0.95, interpolation="linear")) if not eligible_series.empty else None,
                "max": float(eligible_series.max()) if not eligible_series.empty else None,
            }
        )

        grouped = df[df["__is_pernottante_senza_pacchetto"]].groupby(
            ["__norm_stato_provenienza", "__norm_motivazione_principale"],
            dropna=False,
        )
        for (stato, motivazione), part in grouped:
            part_eligible = part.loc[part[threshold_eligible_col], pcg_col]
            group_rows.append(
                {
                    "spesa": col,
                    "metrica": "SPESA_PRO_CAPITE_GIORNALIERA",
                    "stato_provenienza": stato,
                    "motivazione_principale": motivazione,
                    "n_totale_gruppo": int(len(part)),
                    "n_eligible_outlier": int(part[eligible_col].sum()),
                    "n_eligible_soglie": int(part[threshold_eligible_col].sum()),
                    "n_zero_strutturali_pacchetto": int(part[zero_col].sum()),
                    "min": float(part_eligible.min()) if not part_eligible.empty else None,
                    "p5": float(part_eligible.quantile(0.05, interpolation="linear")) if not part_eligible.empty else None,
                    "mediana": float(part_eligible.median()) if not part_eligible.empty else None,
                    "p95": float(part_eligible.quantile(0.95, interpolation="linear")) if not part_eligible.empty else None,
                    "max": float(part_eligible.max()) if not part_eligible.empty else None,
                }
            )

    return pd.DataFrame(global_rows), pd.DataFrame(group_rows)


def find_donor_value(df: pd.DataFrame, row_index: int, spend_col: str) -> tuple[float | None, int, str, float | None]:
    target = df.loc[row_index]

    in_scope_mask = df["__is_pernottante_senza_pacchetto"]
    strong_mask = (
        (df["__norm_stato_provenienza"] == target["__norm_stato_provenienza"])
        & (df["__norm_motivazione_principale"] == target["__norm_motivazione_principale"])
        & in_scope_mask
    )

    donor_mask_base = (
        strong_mask
        & df[f"__pcg_{spend_col}"].notna()
        & (~df[f"__zero_structural_{spend_col}"])
    )

    donor_mask_base.iloc[row_index] = False
    if not donor_mask_base.any():
        return None, 0, "NO_DONOR_STRONG", None

    candidates = df[donor_mask_base].copy()

    def _match_score(candidate_row: pd.Series) -> int:
        score = 0
        if candidate_row["__norm_stato_provenienza"] == target["__norm_stato_provenienza"]:
            score += 1
        if candidate_row["__norm_motivazione_principale"] == target["__norm_motivazione_principale"]:
            score += 1
        if candidate_row["__norm_tipologia_sistemazione"] == target["__norm_tipologia_sistemazione"]:
            score += 1
        if candidate_row["__norm_fascia_eta"] == target["__norm_fascia_eta"]:
            score += 1
        if candidate_row["classe_numero_componenti"] == target["classe_numero_componenti"]:
            score += 1
        return score

    candidates["__match_score"] = candidates.apply(_match_score, axis=1)

    stage1 = candidates[candidates["__match_score"] >= 3]
    if not stage1.empty:
        best_score = int(stage1["__match_score"].max())
        best = stage1[stage1["__match_score"] == best_score]
        donor_pcg = float(best[f"__pcg_{spend_col}"].median())
        denominator = target["__denominatore_spesa"]
        if pd.notna(denominator) and float(denominator) > 0:
            value = float(donor_pcg * float(denominator))
            return value, int(len(best)), f"STRONG_3_OF_5_BEST_{best_score}_FROM_PCG", donor_pcg
        value = float(best[f"__num_{spend_col}"].median())
        return value, int(len(best)), f"STRONG_3_OF_5_BEST_{best_score}_FROM_TOTAL_NO_DEN", donor_pcg

    stage2 = candidates[candidates["__match_score"] >= 2]
    if not stage2.empty:
        donor_pcg = float(stage2[f"__pcg_{spend_col}"].median())
        denominator = target["__denominatore_spesa"]
        if pd.notna(denominator) and float(denominator) > 0:
            value = float(donor_pcg * float(denominator))
            return value, int(len(stage2)), "STRONG_ONLY_FALLBACK_FROM_PCG", donor_pcg
        value = float(stage2[f"__num_{spend_col}"].median())
        return value, int(len(stage2)), "STRONG_ONLY_FALLBACK_FROM_TOTAL_NO_DEN", donor_pcg

    return None, 0, "NO_DONOR_AFTER_FALLBACK", None


def apply_imputations(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, list[int]]]:
    out = df.copy()
    log_rows: list[dict[str, Any]] = []
    imputed_row_indices: dict[str, list[int]] = {col: [] for col in SPEND_COLUMNS}

    for col in SPEND_COLUMNS:
        missing_mask = out[f"__missing_{col}"]
        for idx in out.index[missing_mask]:
            if not bool(out.at[idx, "__is_pernottante_senza_pacchetto"]):
                continue
            if bool(out.at[idx, "__package_active"]):
                continue
            imputed_value, donor_n, rule, donor_pcg = find_donor_value(out, idx, col)
            if imputed_value is None:
                continue

            original_value = out.at[idx, col]
            out.at[idx, col] = imputed_value
            out.at[idx, f"__num_{col}"] = imputed_value
            denominator = out.at[idx, "__denominatore_spesa"]
            if pd.notna(denominator) and float(denominator) > 0:
                out.at[idx, f"__pcg_{col}"] = float(imputed_value) / float(denominator)
            else:
                out.at[idx, f"__pcg_{col}"] = donor_pcg
            out.at[idx, f"__missing_{col}"] = False
            out.at[idx, f"__eligible_outlier_{col}"] = (
                bool(out.at[idx, "__is_pernottante_senza_pacchetto"])
                and pd.notna(out.at[idx, f"__pcg_{col}"])
            )
            out.at[idx, f"__eligible_threshold_{col}"] = (
                bool(out.at[idx, f"__eligible_outlier_{col}"]) and float(imputed_value) > 0
            )
            imputed_row_indices[col].append(int(idx))

            log_rows.append(
                {
                    "id_riga": int(out.at[idx, "id_riga"]),
                    "colonna_spesa": col,
                    "valore_pre": original_value,
                    "valore_post": imputed_value,
                    "numero_donor": donor_n,
                    "regola_match_usata": rule,
                    "donor_valore_pcg": donor_pcg,
                    "denominatore_target": denominator,
                    "stato_provenienza": out.at[idx, "__norm_stato_provenienza"],
                    "motivazione_principale": out.at[idx, "__norm_motivazione_principale"],
                    "flag_pacchetto_attivo": bool(out.at[idx, "__package_active"]),
                    "flag_in_scope_pernottante_senza_pacchetto": bool(
                        out.at[idx, "__is_pernottante_senza_pacchetto"]
                    ),
                }
            )

    log_df = pd.DataFrame(log_rows)
    return out, log_df, imputed_row_indices


def build_outliers_export_df(df: pd.DataFrame) -> pd.DataFrame:
    detail_cols: list[str] = []
    for col in SPEND_COLUMNS:
        detail_cols.extend(
            [
                f"tipo_outlier_{col}",
                f"valore_pcg_{col}",
                f"p5_{col}",
                f"p95_{col}",
                f"livello_soglia_{col}",
                f"__zero_structural_{col}",
            ]
        )

    generated_cols = {
        "id_riga",
        "classe_numero_componenti",
        "outlier_count",
        "ha_outlier",
        "outlier_cols",
    }
    for col in SPEND_COLUMNS:
        generated_cols.add(f"p5_{col}")
        generated_cols.add(f"p95_{col}")
        generated_cols.add(f"livello_soglia_{col}")
        generated_cols.add(f"tipo_outlier_{col}")
        generated_cols.add(f"valore_pcg_{col}")
        generated_cols.add(f"flag_outlier_{col}")

    original_like_cols = [
        c for c in df.columns if (not c.startswith("__")) and (c not in generated_cols)
    ]

    base_cols = ["id_riga", *original_like_cols]
    export_cols = [
        *base_cols,
        "outlier_count",
        "outlier_cols",
        *detail_cols,
    ]

    out = df[df["ha_outlier"]].copy()
    out = out[export_cols]

    rename_map = {f"__zero_structural_{col}": f"zero_strutturale_pacchetto_{col}" for col in SPEND_COLUMNS}
    return out.rename(columns=rename_map)


def build_rules_sheet(min_n: int) -> pd.DataFrame:
    rows = [
        {
            "regola": "Soglie outlier",
            "valore": "Percentili p5 e p95 (interpolazione lineare) su spesa pro capite giornaliera",
        },
        {
            "regola": "Perimetro analisi outlier",
            "valore": "Solo pernottanti senza pacchetto: durata_soggiorno >= 1 e pacchetto vuoto",
        },
        {
            "regola": "Esclusione zeri da soglie",
            "valore": "Soglie p5/p95 calcolate solo su record con spesa totale categoria > 0",
        },
        {
            "regola": "Metrica spesa per soglie",
            "valore": "spesa_totale_categoria / (numero_componenti * durata_soggiorno)",
        },
        {
            "regola": "Gruppo principale",
            "valore": "stato_provenienza x motivazione_principale",
        },
        {
            "regola": "Fallback soglie",
            "valore": "solo provenienza -> solo motivazione -> totale campione",
        },
        {"regola": "Minimo n valido", "valore": str(min_n)},
        {
            "regola": "Eccezione pacchetto",
            "valore": "pacchetto A/B/C + spesa_pacchetto valorizzata + spesa categoria = 0",
        },
        {
            "regola": "Effetto eccezione pacchetto",
            "valore": "record fuori perimetro analisi outlier; zero strutturali non imputati",
        },
        {
            "regola": "Imputazione mancanti",
            "valore": "solo su pernottanti senza pacchetto, donor equivalenti con vincolo forte stato+motivazione",
        },
        {
            "regola": "Matching donor",
            "valore": "almeno 3/5 chiavi; fallback strong only",
        },
        {
            "regola": "Chiavi matching",
            "valore": "stato_provenienza, motivazione_principale, tipologia_sistemazione, fascia_eta, classe_numero_componenti",
        },
    ]
    return pd.DataFrame(rows)


def write_outliers_workbook(
    output_path: Path,
    outliers_df: pd.DataFrame,
    stats_global_df: pd.DataFrame,
    stats_group_df: pd.DataFrame,
    thresholds_df: pd.DataFrame,
    rules_df: pd.DataFrame,
) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        outliers_df.to_excel(writer, sheet_name="OUTLIERS", index=False)
        stats_global_df.to_excel(writer, sheet_name="STATS", index=False)
        stats_group_df.to_excel(writer, sheet_name="STATS_GRUPPI", index=False)
        thresholds_df.to_excel(writer, sheet_name="SOGLIE", index=False)
        rules_df.to_excel(writer, sheet_name="REGOLE", index=False)

        wb = writer.book
        ws = wb["OUTLIERS"]

        if not outliers_df.empty:
            col_pos = {name: i + 1 for i, name in enumerate(outliers_df.columns)}
            for r in range(len(outliers_df)):
                excel_row = r + 2
                for spend_col in SPEND_COLUMNS:
                    type_col = f"tipo_outlier_{spend_col}"
                    if type_col not in outliers_df.columns:
                        continue
                    kind = outliers_df.iloc[r][type_col]
                    if kind in {"OUTLIER_BASSO", "OUTLIER_ALTO"}:
                        spend_col_idx = col_pos.get(spend_col)
                        if spend_col_idx is not None:
                            ws.cell(row=excel_row, column=spend_col_idx).fill = YELLOW_FILL


def write_imputation_workbook(
    output_path: Path,
    original_df: pd.DataFrame,
    imputed_df: pd.DataFrame,
    log_df: pd.DataFrame,
    imputed_indices: dict[str, list[int]],
    stats_global_df: pd.DataFrame,
) -> None:
    out_main = imputed_df[original_df.columns].copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        out_main.to_excel(writer, sheet_name="DATI_IMPUTATI", index=False)
        log_df.to_excel(writer, sheet_name="LOG_IMPUTAZIONI", index=False)
        stats_global_df.to_excel(writer, sheet_name="STATS_POST_IMPUT", index=False)

        wb = writer.book
        ws = wb["DATI_IMPUTATI"]
        col_pos = {name: i + 1 for i, name in enumerate(out_main.columns)}

        for spend_col, row_indices in imputed_indices.items():
            col_idx = col_pos.get(spend_col)
            if col_idx is None:
                continue
            for row_idx in row_indices:
                excel_row = int(row_idx) + 2
                ws.cell(row=excel_row, column=col_idx).fill = GREEN_FILL


def run_pipeline(input_path: Path, output_dir: Path, min_n: int) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"File input non trovato: {input_path}")

    output_dir.mkdir(parents=True, exist_ok=True)

    xls = pd.ExcelFile(input_path)
    first_sheet = xls.sheet_names[0]
    original_df = pd.read_excel(input_path, sheet_name=first_sheet)

    ensure_columns(original_df, MANDATORY_COLUMNS)

    prepared = prepare_dataframe(original_df)

    outlier_scored_df, thresholds_df = detect_outliers(prepared, min_n=min_n)
    stats_global_df, stats_group_df = build_initial_stats(outlier_scored_df)

    outliers_export_df = build_outliers_export_df(outlier_scored_df)
    rules_df = build_rules_sheet(min_n=min_n)

    outliers_path = output_dir / "outliers_da_verificare.xlsx"
    write_outliers_workbook(
        outliers_path,
        outliers_export_df,
        stats_global_df,
        stats_group_df,
        thresholds_df,
        rules_df,
    )

    imputed_df, imputation_log_df, imputed_indices = apply_imputations(prepared)
    post_scored_df, _ = detect_outliers(imputed_df, min_n=min_n)
    stats_post_global_df, _ = build_initial_stats(post_scored_df)

    imputazioni_path = output_dir / "questionari_imputazioni.xlsx"
    write_imputation_workbook(
        imputazioni_path,
        original_df,
        imputed_df,
        imputation_log_df,
        imputed_indices,
        stats_post_global_df,
    )

    print(f"Input: {input_path}")
    print(f"Righe input: {len(original_df)}")
    print(f"Outlier workbook: {outliers_path}")
    print(f"Imputazioni workbook: {imputazioni_path}")
    print(f"Righe con almeno un outlier: {int(outlier_scored_df['ha_outlier'].sum())}")
    print(f"Numero totale imputazioni: {len(imputation_log_df)}")


def parse_args() -> argparse.Namespace:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description=(
            "Pipeline outlier/imputazione separata: statistiche iniziali, export outlier da verificare, "
            "imputazione ND/NR con donor equivalenti e output con evidenziazione celle."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=base_dir.parent / "questionari_fonte.xlsx",
        help="Path del file xlsx sorgente (default: ../questionari_fonte.xlsx)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=base_dir / "output",
        help="Cartella output (default: outlier_pipeline/output)",
    )
    parser.add_argument(
        "--min-n",
        type=int,
        default=30,
        help="Numero minimo di osservazioni valide per usare il gruppo principale nelle soglie.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run_pipeline(input_path=args.input, output_dir=args.output_dir, min_n=args.min_n)


if __name__ == "__main__":
    main()
