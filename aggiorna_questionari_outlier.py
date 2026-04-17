from __future__ import annotations

import argparse
from collections import Counter
from copy import copy
import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_SOURCE = Path("questionari_fonte.xlsx")
DEFAULT_OUTLIER_PACCHETTI = Path("outlier_pacchetti.xlsx")
DEFAULT_OUTLIER_PERNOTTANTI = Path("outlier_pernottanti.xlsx")
DEFAULT_OUTPUT = Path("questionari_update.xlsx")
ID_COL = "ID"


def normalize_id(value: object) -> str:
    if pd.isna(value):
        return ""

    if isinstance(value, (int, np.integer)):
        return str(int(value))

    if isinstance(value, (float, np.floating)):
        value_f = float(value)
        if not math.isfinite(value_f):
            return ""
        if value_f.is_integer():
            return str(int(value_f))
        return str(value_f)

    text = str(value).strip()
    if text == "":
        return ""

    # Gestisce ID numerici serializzati come testo (es. "123.0").
    try:
        parsed = float(text)
    except ValueError:
        return text

    if not math.isfinite(parsed):
        return ""
    if parsed.is_integer():
        return str(int(parsed))
    return text


def get_worksheet(workbook_path: Path, preferred_sheet: str | None = None) -> tuple[Worksheet, str]:
    wb = load_workbook(workbook_path)
    sheet_name = preferred_sheet or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Foglio '{sheet_name}' non presente in {workbook_path}")
    return wb[sheet_name], sheet_name


def build_header_map(ws: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col_idx).value
        if value is None:
            continue
        text = str(value).strip()
        if text == "":
            continue
        headers[text] = col_idx
    return headers


def build_id_rows_index(ws: Worksheet, id_col_idx: int) -> tuple[dict[str, list[int]], int]:
    id_to_rows: dict[str, list[int]] = {}
    for row_idx in range(2, ws.max_row + 1):
        key = normalize_id(ws.cell(row=row_idx, column=id_col_idx).value)
        if key == "":
            continue
        id_to_rows.setdefault(key, []).append(row_idx)

    duplicated_rows = int(sum(len(rows) for rows in id_to_rows.values() if len(rows) > 1))
    return id_to_rows, duplicated_rows


def build_outlier_last_row_index(
    ws: Worksheet,
    id_col_idx: int,
) -> tuple[dict[str, int], int, int, list[str]]:
    id_to_last_row: dict[str, int] = {}
    raw_ids: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        key = normalize_id(ws.cell(row=row_idx, column=id_col_idx).value)
        if key == "":
            continue
        raw_ids.append(key)
        id_to_last_row[key] = row_idx

    counts = Counter(raw_ids)
    duplicated_rows = int(sum(count for count in counts.values() if count > 1))
    return id_to_last_row, len(id_to_last_row), duplicated_rows, raw_ids


def copy_cell_value_and_style(source_cell, donor_cell) -> None:
    source_cell.value = donor_cell.value
    source_cell.font = copy(donor_cell.font)
    source_cell.fill = copy(donor_cell.fill)
    source_cell.border = copy(donor_cell.border)
    source_cell.alignment = copy(donor_cell.alignment)
    source_cell.protection = copy(donor_cell.protection)
    source_cell.number_format = donor_cell.number_format

    source_cell.comment = copy(donor_cell.comment) if donor_cell.comment is not None else None

    if donor_cell.hyperlink is not None:
        source_cell._hyperlink = copy(donor_cell.hyperlink)
    else:
        source_cell._hyperlink = None


def get_outlier_columns(header: dict[str, int]) -> list[str]:
    return [col_name for col_name in header if "outlier" in col_name.strip().lower()]


def append_missing_outlier_columns(
    source_ws: Worksheet,
    source_header: dict[str, int],
    outlier_sources: list[dict[str, object]],
) -> list[str]:
    added_columns: list[str] = []

    for source in outlier_sources:
        ws = source["ws"]
        header = source["header"]
        for col_name in get_outlier_columns(header):
            if col_name in source_header:
                continue

            donor_col_idx = int(header[col_name])
            new_col_idx = source_ws.max_column + 1

            source_header[col_name] = new_col_idx
            source_header_cell = source_ws.cell(row=1, column=new_col_idx)
            donor_header_cell = ws.cell(row=1, column=donor_col_idx)
            copy_cell_value_and_style(source_header_cell, donor_header_cell)

            if source_header_cell.value is None or str(source_header_cell.value).strip() == "":
                source_header_cell.value = col_name

            added_columns.append(col_name)

    return added_columns


def replace_rows_with_style(
    source_ws: Worksheet,
    source_header: dict[str, int],
    source_id_rows: dict[str, list[int]],
    outlier_sources: list[dict[str, object]],
) -> tuple[int, int, int, int, int, int, int]:
    combined_last_by_id: dict[str, tuple[Worksheet, int, dict[str, int]]] = {}
    all_raw_ids: list[str] = []

    unique_p = duplicated_p = unique_n = duplicated_n = 0
    for source in outlier_sources:
        ws = source["ws"]
        header = source["header"]
        id_col_idx = int(source["id_col_idx"])
        label = str(source["label"])

        id_to_last_row, unique_count, duplicated_rows, raw_ids = build_outlier_last_row_index(ws, id_col_idx)
        if label == "outlier_pacchetti":
            unique_p = unique_count
            duplicated_p = duplicated_rows
        elif label == "outlier_pernottanti":
            unique_n = unique_count
            duplicated_n = duplicated_rows

        all_raw_ids.extend(raw_ids)
        for id_key, donor_row in id_to_last_row.items():
            combined_last_by_id[id_key] = (ws, donor_row, header)

    cross_counts = Counter(all_raw_ids)
    duplicated_cross = int(sum(count for count in cross_counts.values() if count > 1))

    matched_rows = 0
    missing_ids = 0

    for id_key, donor_info in combined_last_by_id.items():
        ws_outlier, donor_row, donor_header = donor_info
        target_rows = source_id_rows.get(id_key)
        if not target_rows:
            missing_ids += 1
            continue

        common_cols = [col for col in donor_header if col in source_header]
        if not common_cols:
            continue

        for source_row in target_rows:
            for col_name in common_cols:
                source_col_idx = source_header[col_name]
                donor_col_idx = donor_header[col_name]
                source_cell = source_ws.cell(row=source_row, column=source_col_idx)
                donor_cell = ws_outlier.cell(row=donor_row, column=donor_col_idx)
                copy_cell_value_and_style(source_cell, donor_cell)
            matched_rows += 1

    return matched_rows, missing_ids, unique_p, duplicated_p, unique_n, duplicated_n, duplicated_cross

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Sostituisce nel file questionari_fonte le righe con stesso ID presenti nei file "
            "outlier_pacchetti e outlier_pernottanti, generando questionari_update."
        )
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="File Excel sorgente.")
    parser.add_argument(
        "--outlier-pacchetti",
        type=Path,
        default=DEFAULT_OUTLIER_PACCHETTI,
        help="File Excel outlier pacchetti.",
    )
    parser.add_argument(
        "--outlier-pernottanti",
        type=Path,
        default=DEFAULT_OUTLIER_PERNOTTANTI,
        help="File Excel outlier pernottanti.",
    )
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="File Excel di output.")
    parser.add_argument(
        "--sheet-source",
        type=str,
        default=None,
        help="Nome foglio da usare nel file sorgente (default: primo foglio).",
    )
    parser.add_argument(
        "--sheet-outlier-pacchetti",
        type=str,
        default=None,
        help="Nome foglio del file outlier pacchetti (default: primo foglio).",
    )
    parser.add_argument(
        "--sheet-outlier-pernottanti",
        type=str,
        default=None,
        help="Nome foglio del file outlier pernottanti (default: primo foglio).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    for path_arg in (args.source, args.outlier_pacchetti, args.outlier_pernottanti):
        if not path_arg.exists():
            raise FileNotFoundError(f"File non trovato: {path_arg}")

    source_wb = load_workbook(args.source)
    source_sheet = args.sheet_source or source_wb.sheetnames[0]
    if source_sheet not in source_wb.sheetnames:
        raise ValueError(f"Foglio '{source_sheet}' non presente in {args.source}")

    source_ws = source_wb[source_sheet]
    source_header = build_header_map(source_ws)
    if ID_COL not in source_header:
        raise ValueError(f"Colonna mancante nel file sorgente: {ID_COL}")
    source_id_rows, _ = build_id_rows_index(source_ws, source_header[ID_COL])

    outlier_p_ws, _ = get_worksheet(args.outlier_pacchetti, args.sheet_outlier_pacchetti)
    outlier_n_ws, _ = get_worksheet(args.outlier_pernottanti, args.sheet_outlier_pernottanti)

    outlier_p_header = build_header_map(outlier_p_ws)
    outlier_n_header = build_header_map(outlier_n_ws)
    if ID_COL not in outlier_p_header:
        raise ValueError(f"Colonna mancante nel file outlier_pacchetti: {ID_COL}")
    if ID_COL not in outlier_n_header:
        raise ValueError(f"Colonna mancante nel file outlier_pernottanti: {ID_COL}")

    outlier_sources = [
        {
            "ws": outlier_p_ws,
            "header": outlier_p_header,
            "id_col_idx": outlier_p_header[ID_COL],
            "label": "outlier_pacchetti",
        },
        {
            "ws": outlier_n_ws,
            "header": outlier_n_header,
            "id_col_idx": outlier_n_header[ID_COL],
            "label": "outlier_pernottanti",
        },
    ]

    added_outlier_columns = append_missing_outlier_columns(source_ws, source_header, outlier_sources)

    matched_rows, missing_ids, unique_p, duplicated_p, unique_n, duplicated_n, duplicated_cross = replace_rows_with_style(
        source_ws,
        source_header,
        source_id_rows,
        outlier_sources,
    )

    source_wb.save(args.output)

    print(f"File scritto: {args.output}")
    print(f"Foglio aggiornato: {source_sheet}")
    print(f"Righe sorgente totali: {source_ws.max_row - 1}")
    print(f"Righe sostituite per ID: {matched_rows}")
    print(f"ID outlier non trovati nel sorgente: {missing_ids}")
    if added_outlier_columns:
        print("Colonne outlier aggiunte: " + ", ".join(added_outlier_columns))
    else:
        print("Colonne outlier aggiunte: nessuna")
    print(
        "Dettaglio outlier letti: "
        f"pacchetti={unique_p} (duplicati interni rilevati={duplicated_p}), "
        f"pernottanti={unique_n} (duplicati interni rilevati={duplicated_n}), "
        f"duplicati tra i due file={duplicated_cross}"
    )


if __name__ == "__main__":
    main()