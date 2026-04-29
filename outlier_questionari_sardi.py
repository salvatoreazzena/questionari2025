from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_INPUT = Path('questionari_sardi.xlsx')
DEFAULT_OUTPUT = Path('questionari_sardi_outlier_sostituiti.xlsx')

ID_COL = 'ID'
VACANZA_COL = 'vacanza_2025'
DOVE_COL = 'si_dove'
DURATION_COL = 'durata_soggiorno'
COMPONENTI_COL = 'numero_componenti'
MOTIV_COL = 'motivazione_principale'
PACCHETTO_COL = 'pacchetto'
SPESA_PACCHETTO_COL = 'spesa_pacchetto'

SPESA_COLUMNS_NO_PACCHETTO = [
    'spese_trasporto_viaggio',
    'spese_trasporto_interno',
    'spese_alloggio',
    'spese_alimentazione',
    'spese_ristorazione',
    'spese_souvenir',
    'spese_altre',
]
PACCHETTO_ALLOWED_VALUES = {'A', 'B', 'C'}

NULL_TOKENS = {'', 'ND', 'NR', 'NA', 'N/D', 'N.R.', 'N.A.'}
ANALYSIS_SHEET = 'outlier_sostituiti'
STATS_SHEET = 'statistiche'
RED_FILL = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')


@dataclass(frozen=True)
class ThresholdInfo:
    p5: float | None
    p95: float | None


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ''
    return str(value).strip()


def is_null_like(value: object) -> bool:
    return normalize_text(value).upper() in {t.upper() for t in NULL_TOKENS}


def parse_number(value: object) -> float | None:
    if pd.isna(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        v = float(value)
        return v if math.isfinite(v) else None

    txt = normalize_text(value)
    if txt == '' or txt.upper() in {t.upper() for t in NULL_TOKENS}:
        return None

    txt = txt.replace('€', '').replace(' ', '')
    if re.fullmatch(r'[-+]?\d{1,3}(\.\d{3})*(,\d+)?', txt):
        txt = txt.replace('.', '').replace(',', '.')
    elif re.fullmatch(r'[-+]?\d{1,3}(,\d{3})*(\.\d+)?', txt):
        txt = txt.replace(',', '')
    else:
        txt = txt.replace(',', '.')

    try:
        return float(txt)
    except ValueError:
        return None


def compute_daily_per_capita_series(amount_series: pd.Series, components_series: pd.Series, duration_series: pd.Series) -> pd.Series:
    amount_num = pd.to_numeric(amount_series, errors='coerce')
    components_num = pd.to_numeric(components_series, errors='coerce')
    duration_num = pd.to_numeric(duration_series, errors='coerce')
    denominator = components_num * duration_num
    return (amount_num / denominator).where(denominator > 0)


def build_tourist_weight_series(components_series: pd.Series) -> pd.Series:
    weights = pd.to_numeric(components_series, errors='coerce')
    weights = weights.where(weights > 0)
    return weights.fillna(0.0)


def weighted_count(mask: pd.Series, weights: pd.Series) -> int | float:
    total = float(weights[mask.fillna(False).astype(bool)].sum())
    if not math.isfinite(total):
        return 0
    as_int = int(round(total))
    return as_int if abs(total - as_int) < 1e-9 else round(total, 2)


def build_destinazione(series: pd.Series) -> pd.Series:
    s = series.astype('string').fillna('').str.strip().str.upper()
    out = pd.Series(pd.NA, index=s.index, dtype='string')
    out = out.mask(s == 'IN SARDEGNA', 'SARDEGNA')
    out = out.mask(s == 'IN ITALIA', 'ITALIA')
    out = out.mask(s == "ALL'ESTERO", 'ESTERO')
    return out


def is_vacanza_si(value: object) -> bool:
    return normalize_text(value).lower() in {'sì', 'si', 'sė'}


def weighted_quantile(values: pd.Series, weights: pd.Series, q: float) -> float | None:
    v = pd.to_numeric(values, errors='coerce')
    w = pd.to_numeric(weights, errors='coerce')
    m = v.notna() & w.notna() & (w > 0)
    if not bool(m.any()):
        return None
    vals = v[m].to_numpy(dtype='float64', copy=False)
    wei = w[m].to_numpy(dtype='float64', copy=False)
    order = np.argsort(vals, kind='mergesort')
    vals = vals[order]
    wei = wei[order]
    cum = np.cumsum(wei)
    tw = float(cum[-1])
    if not math.isfinite(tw) or tw <= 0:
        return None
    idx = int(np.searchsorted(cum, min(max(q, 0.0), 1.0) * tw, side='left'))
    idx = min(max(idx, 0), len(vals) - 1)
    return float(vals[idx])


def compute_thresholds(df: pd.DataFrame, value_col: str) -> dict[tuple[str, str], ThresholdInfo]:
    out: dict[tuple[str, str], ThresholdInfo] = {}
    for key, g in df.groupby(['destinazione_grp', 'motivazione_grp'], dropna=False):
        p5 = weighted_quantile(g[value_col], g['_turisti_weight'], 0.05)
        p95 = weighted_quantile(g[value_col], g['_turisti_weight'], 0.95)
        if p5 is None or p95 is None:
            vals = pd.to_numeric(g[value_col], errors='coerce').dropna()
            if vals.empty:
                out[key] = ThresholdInfo(None, None)
                continue
            p5 = float(vals.quantile(0.05))
            p95 = float(vals.quantile(0.95))
        out[key] = ThresholdInfo(p5, p95)
    return out


def select_donor_id(donor_df: pd.DataFrame, value_col: str, target_value: float, target_duration: float | None, target_components: float | None) -> str:
    tmp = donor_df.copy()
    tmp['_abs_to_median'] = (tmp[value_col] - target_value).abs()
    if target_duration is None:
        tmp['_abs_duration'] = np.inf
    else:
        tmp['_abs_duration'] = (tmp['_durata_num'] - target_duration).abs().where(tmp['_durata_num'].notna(), np.inf)
    if target_components is None:
        tmp['_abs_componenti'] = np.inf
    else:
        tmp['_abs_componenti'] = (tmp['_componenti_num'] - target_components).abs().where(tmp['_componenti_num'].notna(), np.inf)
    tmp['_id_text'] = tmp[ID_COL].astype('string')
    tmp = tmp.sort_values(by=['_abs_to_median', '_abs_duration', '_abs_componenti', '_id_text'], kind='mergesort')
    return str(tmp.iloc[0][ID_COL])


def infer_value_from_similar(df: pd.DataFrame, row_idx: int, value_col: str) -> tuple[float | None, str | None, str | None]:
    row = df.loc[row_idx]
    levels = [
        ('stessa_dest_motivazione_componenti_durata', ['destinazione_grp', 'motivazione_grp', '_componenti_num', '_durata_num']),
        ('stessa_dest_motivazione', ['destinazione_grp', 'motivazione_grp']),
        ('stessa_destinazione', ['destinazione_grp']),
        ('stessa_motivazione', ['motivazione_grp']),
        ('globale', []),
    ]
    base_pool = df[(df.index != row_idx) & df[value_col].notna()].copy()
    if base_pool.empty:
        return None, None, None

    for level_name, cols in levels:
        pool = base_pool
        skip = False
        for c in cols:
            if pd.isna(row[c]):
                skip = True
                break
            pool = pool[pool[c] == row[c]]
        if skip or pool.empty:
            continue

        median_value = float(pool[value_col].median())
        donor_id = select_donor_id(
            pool,
            value_col=value_col,
            target_value=median_value,
            target_duration=row['_durata_num'] if pd.notna(row['_durata_num']) else None,
            target_components=row['_componenti_num'] if pd.notna(row['_componenti_num']) else None,
        )
        return median_value, donor_id, level_name
    return None, None, None


def format_imputation_note(method: str | None, donor_id: str | None) -> str:
    labels = {
        'stessa_dest_motivazione_componenti_durata': 'MEDIANA[dest+motivazione+componenti+durata]',
        'stessa_dest_motivazione': 'MEDIANA[dest+motivazione]',
        'stessa_destinazione': 'MEDIANA[destinazione]',
        'stessa_motivazione': 'MEDIANA[motivazione]',
        'globale': 'MEDIANA[globale]',
    }
    base = labels.get(method or '', 'MEDIANA')
    return base if donor_id is None else f'{base}; donor_ID={donor_id}'


def validate_columns(df: pd.DataFrame) -> None:
    required = [ID_COL, VACANZA_COL, DOVE_COL, DURATION_COL, COMPONENTI_COL, MOTIV_COL, PACCHETTO_COL, SPESA_PACCHETTO_COL, *SPESA_COLUMNS_NO_PACCHETTO]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f'Colonne mancanti nel dataset: {missing}')


def process_subset(df_source: pd.DataFrame, eligible_mask: pd.Series, spend_cols: list[str], stats_rows: list[dict[str, object]]) -> tuple[pd.DataFrame, dict[str, pd.Series], dict[str, pd.Series], dict[str, pd.Series]]:
    df_work = df_source.loc[eligible_mask].copy()
    df_work['_componenti_num'] = pd.to_numeric(df_work[COMPONENTI_COL], errors='coerce')
    df_work['_durata_num'] = pd.to_numeric(df_work[DURATION_COL], errors='coerce')
    df_work['_turisti_weight'] = build_tourist_weight_series(df_work['_componenti_num'])
    df_work['destinazione_grp'] = build_destinazione(df_work[DOVE_COL])
    df_work['motivazione_grp'] = df_work[MOTIV_COL].astype('string').fillna('').str.strip().str.upper()

    out_masks: dict[str, pd.Series] = {}
    imp_masks: dict[str, pd.Series] = {}
    imp_out_masks: dict[str, pd.Series] = {}

    for spesa_col in spend_cols:
        imputation_col = f'imputazione_{spesa_col}'
        if imputation_col not in df_source.columns:
            df_source[imputation_col] = pd.NA

        numeric_col = f'_{spesa_col}_num'
        metric_col = f'_{spesa_col}_metric'
        df_work[numeric_col] = df_work[spesa_col].map(parse_number)
        metric_before = compute_daily_per_capita_series(df_work[numeric_col], df_work['_componenti_num'], df_work['_durata_num'])
        df_work[metric_col] = metric_before.copy()

        thresholds = compute_thresholds(df_work, metric_col)

        nd_mask = df_work[spesa_col].map(is_null_like)
        imputed_mask_local = pd.Series(False, index=df_work.index)
        imputed_outlier_local = pd.Series(False, index=df_work.index)

        for idx in df_work.index[nd_mask].tolist():
            inferred, donor_id, method = infer_value_from_similar(df_work, idx, numeric_col)
            if inferred is None:
                continue
            inferred_rounded = round(float(inferred), 2)
            df_work.at[idx, spesa_col] = inferred_rounded
            df_work.at[idx, numeric_col] = inferred_rounded
            df_source.at[idx, spesa_col] = inferred_rounded
            df_source.at[idx, imputation_col] = format_imputation_note(method, donor_id)
            imputed_mask_local.at[idx] = True

        df_work[metric_col] = compute_daily_per_capita_series(df_work[numeric_col], df_work['_componenti_num'], df_work['_durata_num'])

        outlier_local = pd.Series(False, index=df_work.index)
        replaced_rows = 0

        for idx, row in df_work.iterrows():
            val = row[metric_col]
            if pd.isna(val):
                continue
            info = thresholds.get((row['destinazione_grp'], row['motivazione_grp']), ThresholdInfo(None, None))
            if info.p5 is None or info.p95 is None:
                continue
            target_metric = None
            if val < info.p5:
                target_metric = info.p5
            elif val > info.p95:
                target_metric = info.p95
            if target_metric is None:
                continue

            denom = row['_componenti_num'] * row['_durata_num']
            if pd.isna(denom) or float(denom) <= 0:
                continue

            new_value = round(float(target_metric) * float(denom), 2)
            df_source.at[idx, spesa_col] = new_value
            outlier_local.at[idx] = True
            replaced_rows += 1
            if bool(imputed_mask_local.at[idx]):
                imputed_outlier_local.at[idx] = True

        out_masks[spesa_col] = outlier_local.reindex(df_source.index, fill_value=False)
        imp_masks[spesa_col] = imputed_mask_local.reindex(df_source.index, fill_value=False)
        imp_out_masks[spesa_col] = imputed_outlier_local.reindex(df_source.index, fill_value=False)

        stats_rows.append({
            'sezione': 'con_pacchetto' if spesa_col == SPESA_PACCHETTO_COL else 'senza_pacchetto',
            'colonna_spesa': spesa_col,
            'n_analizzati': int(df_work.shape[0]),
            'n_numerici_iniziali': weighted_count(metric_before.notna(), df_work['_turisti_weight']),
            'n_nd_iniziali': weighted_count(nd_mask, df_work['_turisti_weight']),
            'n_imputati': weighted_count(imputed_mask_local, df_work['_turisti_weight']),
            'n_outlier_sostituiti': weighted_count(outlier_local, df_work['_turisti_weight']),
            'n_imputati_outlier_sostituiti': weighted_count(imputed_outlier_local, df_work['_turisti_weight']),
            'n_outlier_sostituiti_righe': replaced_rows,
        })

    return df_source, out_masks, imp_masks, imp_out_masks


def write_output(output_file: Path, output_df: pd.DataFrame, outlier_masks: dict[str, pd.Series], imputed_masks: dict[str, pd.Series], imputed_outlier_masks: dict[str, pd.Series], stats_df: pd.DataFrame) -> None:
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        output_df.to_excel(writer, sheet_name=ANALYSIS_SHEET, index=False)
        stats_df.to_excel(writer, sheet_name=STATS_SHEET, index=False)
        ws = writer.book[ANALYSIS_SHEET]
        header_map = {str(c.value): c.column for c in ws[1]}

        for col, out_mask in outlier_masks.items():
            col_idx = header_map.get(col)
            if col_idx is None:
                continue
            col_letter = get_column_letter(col_idx)
            imp_mask = imputed_masks.get(col, pd.Series(False, index=output_df.index))
            imp_out_mask = imputed_outlier_masks.get(col, pd.Series(False, index=output_df.index))
            for i in range(len(output_df)):
                r = i + 2
                if bool(imp_out_mask.iat[i]):
                    ws[f'{col_letter}{r}'].fill = ORANGE_FILL
                elif bool(imp_mask.iat[i]):
                    ws[f'{col_letter}{r}'].fill = YELLOW_FILL
                elif bool(out_mask.iat[i]):
                    ws[f'{col_letter}{r}'].fill = RED_FILL


def run(input_file: Path, output_file: Path) -> None:
    xls = pd.ExcelFile(input_file)
    first_sheet = xls.sheet_names[0]
    df = pd.read_excel(input_file, sheet_name=first_sheet)
    validate_columns(df)

    output_df = df.copy()

    vacanza_si = output_df[VACANZA_COL].map(is_vacanza_si)
    valid_dest = build_destinazione(output_df[DOVE_COL]).notna()
    durata_ok = pd.to_numeric(output_df[DURATION_COL], errors='coerce') >= 1
    pac_norm = output_df[PACCHETTO_COL].astype('string').fillna('').str.strip().str.upper()

    eligible_no_pacchetto = vacanza_si & valid_dest & durata_ok & pac_norm.eq('')
    eligible_con_pacchetto = vacanza_si & valid_dest & durata_ok & pac_norm.isin(PACCHETTO_ALLOWED_VALUES)

    stats_rows: list[dict[str, object]] = []

    output_df, out1, imp1, impout1 = process_subset(output_df, eligible_no_pacchetto, SPESA_COLUMNS_NO_PACCHETTO, stats_rows)
    output_df, out2, imp2, impout2 = process_subset(output_df, eligible_con_pacchetto, [SPESA_PACCHETTO_COL], stats_rows)

    outlier_masks = {**out1, **out2}
    imputed_masks = {**imp1, **imp2}
    imputed_outlier_masks = {**impout1, **impout2}

    stats_rows.append({'sezione': 'filtro', 'colonna_spesa': '-', 'n_analizzati': int((vacanza_si & valid_dest & durata_ok).sum())})
    stats_rows.append({'sezione': 'filtro_esclusi_no_vacanza', 'colonna_spesa': '-', 'n_analizzati': int((~vacanza_si).sum())})

    stats_df = pd.DataFrame(stats_rows)
    write_output(output_file, output_df, outlier_masks, imputed_masks, imputed_outlier_masks, stats_df)

    print(f'Input letto: {input_file}')
    print(f'Foglio sorgente: {first_sheet}')
    print(f'Output generato: {output_file}')


def main() -> None:
    parser = argparse.ArgumentParser(description='Outlier + imputazione per questionari_sardi (solo chi ha fatto vacanza).')
    parser.add_argument('--input', type=Path, default=DEFAULT_INPUT)
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    run(args.input, args.output)


if __name__ == '__main__':
    main()
