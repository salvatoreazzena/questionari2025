from __future__ import annotations

import argparse
import math
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_INPUT = Path("questionari_fonte.xlsx")
DEFAULT_OUTPUT = Path("questionari_outlier_pernottanti.xlsx")
DEFAULT_OUTPUT_NO_IMPUTATION = Path("questionari_outlier_pernottanti_no_imputazioni.xlsx")

ID_COL = "ID"
DURATION_COL = "durata_soggiorno"
PACCHETTO_COL = "pacchetto"
STATO_COL = "stato_provenienza"
MOTIV_COL = "motivazione_principale"
COMPONENTI_COL = "numero_componenti"

SPESA_COLUMNS = [
    "spese_trasporto_viaggio",
    "spese_trasporto_interno",
    "spese_alloggio",
    "spese_alimentazione",
    "spese_ristorazione",
    "spese_souvenir",
    "spese_altre",
]

NULL_TOKENS = {"", "ND", "NR", "NA", "N/D", "N.R.", "N.A."}

ANALYSIS_SHEET = "pernottanti_senza_pacchetto"
STATS_SHEET = "statistiche"
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")


@dataclass(frozen=True)
class ThresholdInfo:
    p5: float | None
    p95: float | None
    n: int
    n_turisti: int | float


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_strict_empty(value: object) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def is_null_like(value: object) -> bool:
    txt = normalize_text(value).upper()
    return txt in {t.upper() for t in NULL_TOKENS}


def parse_number(value: object) -> float | None:
    if pd.isna(value):
        return None

    if isinstance(value, (int, float, np.integer, np.floating)):
        v = float(value)
        if math.isfinite(v):
            return v
        return None

    txt = normalize_text(value)
    if txt == "":
        return None
    if txt.upper() in {t.upper() for t in NULL_TOKENS}:
        return None

    txt = txt.replace("€", "").replace(" ", "")

    # 1.234,56 -> 1234.56
    if re.fullmatch(r"[-+]?\d{1,3}(\.\d{3})*(,\d+)?", txt):
        txt = txt.replace(".", "").replace(",", ".")
    # 1,234.56 -> 1234.56
    elif re.fullmatch(r"[-+]?\d{1,3}(,\d{3})*(\.\d+)?", txt):
        txt = txt.replace(",", "")
    else:
        txt = txt.replace(",", ".")

    try:
        return float(txt)
    except ValueError:
        return None


def compute_daily_per_capita_value(
    amount: float | None,
    components: float | None,
    duration: float | None,
) -> float | None:
    if amount is None or components is None or duration is None:
        return None
    if pd.isna(amount) or pd.isna(components) or pd.isna(duration):
        return None

    amount_f = float(amount)
    components_f = float(components)
    duration_f = float(duration)
    denominator = components_f * duration_f

    if not math.isfinite(amount_f) or not math.isfinite(denominator):
        return None
    if denominator <= 0:
        return None
    return amount_f / denominator


def compute_daily_per_capita_series(
    amount_series: pd.Series,
    components_series: pd.Series,
    duration_series: pd.Series,
) -> pd.Series:
    amount_num = pd.to_numeric(amount_series, errors="coerce")
    components_num = pd.to_numeric(components_series, errors="coerce")
    duration_num = pd.to_numeric(duration_series, errors="coerce")

    denominator = components_num * duration_num
    per_capita = amount_num / denominator
    return per_capita.where(denominator > 0)


def build_tourist_weight_series(components_series: pd.Series) -> pd.Series:
    weights = pd.to_numeric(components_series, errors="coerce")
    weights = weights.where(weights > 0)
    return weights.fillna(0.0)


def format_count_value(value: float) -> int | float:
    if not math.isfinite(value):
        return 0
    rounded_int = int(round(value))
    if abs(value - rounded_int) < 1e-9:
        return rounded_int
    return round(value, 2)


def weighted_count(mask: pd.Series, weights: pd.Series) -> int | float:
    mask_bool = mask.fillna(False).astype(bool)
    total = float(weights[mask_bool].sum())
    return format_count_value(total)


def build_macro_provenienza(series: pd.Series) -> pd.Series:
    s = series.astype("string").fillna("").str.strip().str.upper()
    out = pd.Series(pd.NA, index=s.index, dtype="string")
    out = out.mask(s == "ITALIA", "ITALIANI")
    out = out.mask((s != "") & (s != "ITALIA"), "STRANIERI")
    return out


def compute_thresholds(
    df: pd.DataFrame,
    value_col: str,
    group_cols: list[str],
    weights_col: str | None = None,
) -> dict[tuple[str, str], ThresholdInfo]:
    thresholds: dict[tuple[str, str], ThresholdInfo] = {}
    grouped = df.groupby(group_cols, dropna=False)
    for key, g in grouped:
        value_num = pd.to_numeric(g[value_col], errors="coerce")
        valid_mask = value_num.notna()
        vals = value_num[valid_mask]
        n = int(vals.shape[0])

        if weights_col and weights_col in g.columns:
            tourist_weights = build_tourist_weight_series(g[weights_col])
            n_turisti = weighted_count(valid_mask, tourist_weights)
        else:
            n_turisti = n

        if n == 0:
            thresholds[key] = ThresholdInfo(p5=None, p95=None, n=n, n_turisti=0)
            continue

        p5 = float(vals.quantile(0.05))
        p95 = float(vals.quantile(0.95))
        thresholds[key] = ThresholdInfo(p5=p5, p95=p95, n=n, n_turisti=n_turisti)
    return thresholds


def select_donor_id(
    donor_df: pd.DataFrame,
    *,
    value_col: str,
    target_value: float,
    target_duration: float | None,
    target_components: float | None,
) -> str:
    tmp = donor_df.copy()
    tmp["_abs_to_median"] = (tmp[value_col] - target_value).abs()

    if target_duration is None:
        tmp["_abs_duration"] = np.full(len(tmp), np.inf, dtype="float64")
    else:
        tmp["_abs_duration"] = (tmp["_durata_num"] - target_duration).abs().astype("float64")
        tmp["_abs_duration"] = tmp["_abs_duration"].where(tmp["_durata_num"].notna(), np.inf)

    if target_components is None:
        tmp["_abs_componenti"] = np.full(len(tmp), np.inf, dtype="float64")
    else:
        tmp["_abs_componenti"] = (tmp["_componenti_num"] - target_components).abs().astype("float64")
        tmp["_abs_componenti"] = tmp["_abs_componenti"].where(tmp["_componenti_num"].notna(), np.inf)

    tmp["_id_text"] = tmp[ID_COL].astype("string")

    tmp = tmp.sort_values(
        by=["_abs_to_median", "_abs_duration", "_abs_componenti", "_id_text"],
        kind="mergesort",
    )

    return str(tmp.iloc[0][ID_COL])


def infer_value_from_similar(
    df: pd.DataFrame,
    *,
    row_idx: int,
    value_col: str,
) -> tuple[float | None, str | None, str | None]:
    row = df.loc[row_idx]

    levels = [
        (
            "stesso_macro_motivazione_componenti_durata",
            ["macro_provenienza", "motivazione_grp", "_componenti_num", "_durata_num"],
        ),
        ("stesso_macro_motivazione", ["macro_provenienza", "motivazione_grp"]),
        ("stesso_macro", ["macro_provenienza"]),
        ("stessa_motivazione", ["motivazione_grp"]),
        ("globale", []),
    ]

    base_pool = df[(df.index != row_idx) & df[value_col].notna()].copy()
    if base_pool.empty:
        return None, None, None

    for level_name, cols in levels:
        pool = base_pool
        skip_level = False
        for c in cols:
            target = row[c]
            if pd.isna(target):
                skip_level = True
                break
            pool = pool[pool[c] == target]
        if skip_level or pool.empty:
            continue

        median_value = float(pool[value_col].median())
        donor_id = select_donor_id(
            pool,
            value_col=value_col,
            target_value=median_value,
            target_duration=row["_durata_num"] if pd.notna(row["_durata_num"]) else None,
            target_components=row["_componenti_num"] if pd.notna(row["_componenti_num"]) else None,
        )
        return median_value, donor_id, level_name

    return None, None, None


def validate_columns(df: pd.DataFrame) -> None:
    required = [
        ID_COL,
        DURATION_COL,
        PACCHETTO_COL,
        STATO_COL,
        MOTIV_COL,
        COMPONENTI_COL,
        *SPESA_COLUMNS,
    ]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti nel dataset: {missing}")


def format_imputation_note(method: str | None, donor_id: str | None) -> str:
    method_labels = {
        "stesso_macro_motivazione_componenti_durata": "MEDIANA[stesso_macro+motivazione+componenti+durata]",
        "stesso_macro_motivazione": "MEDIANA[stesso_macro+motivazione]",
        "stesso_macro": "MEDIANA[stesso_macro]",
        "stessa_motivazione": "MEDIANA[stessa_motivazione]",
        "globale": "MEDIANA[globale]",
    }
    base = method_labels.get(method or "", "MEDIANA")
    if donor_id is None:
        return base
    return f"{base}; donor_ID={donor_id}"


def build_stats_table(
    *,
    analyzed_tourists: int | float,
    full_tourists: int | float,
    spend_stats: list[dict[str, object]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    meta = pd.DataFrame(
        [
            {"metrica": "turisti_input_totali", "valore": full_tourists},
            {"metrica": "turisti_analizzati_filtro", "valore": analyzed_tourists},
            {
                "metrica": "filtro_applicato",
                "valore": "durata_soggiorno >= 1 AND pacchetto vuoto",
            },
            {
                "metrica": "gruppo_omogeneo_outlier",
                "valore": "macro_provenienza (ITALIANI/STRANIERI) x motivazione_principale",
            },
            {
                "metrica": "unita_conteggi",
                "valore": "conteggi pesati per numero_componenti",
            },
            {
                "metrica": "metrica_outlier",
                "valore": "spesa_pro_capite_giornaliera = spesa / (numero_componenti * durata_soggiorno)",
            },
        ]
    )
    detail = pd.DataFrame(spend_stats)
    return meta, detail


def build_thresholds_stats_table(
    thresholds_by_spesa: dict[str, dict[tuple[str, str], ThresholdInfo]],
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for spesa_col in SPESA_COLUMNS:
        thresholds = thresholds_by_spesa.get(spesa_col, {})
        for key, info in thresholds.items():
            macro, motivazione = key
            rows.append(
                {
                    "colonna_spesa": spesa_col,
                    "macro_provenienza": macro,
                    "motivazione_grp": motivazione,
                    "n_turisti_gruppo": info.n_turisti,
                    "p5_gruppo": info.p5,
                    "p95_gruppo": info.p95,
                }
            )

    if not rows:
        return pd.DataFrame(
            columns=[
                "colonna_spesa",
                "macro_provenienza",
                "motivazione_grp",
                "n_turisti_gruppo",
                "p5_gruppo",
                "p95_gruppo",
            ]
        )

    out = pd.DataFrame(rows)
    return out.sort_values(
        by=["colonna_spesa", "macro_provenienza", "motivazione_grp"],
        kind="mergesort",
    ).reset_index(drop=True)


def write_output_excel(
    *,
    output_file: Path,
    output_df: pd.DataFrame,
    stats_detail_df: pd.DataFrame,
    stats_thresholds_df: pd.DataFrame,
    stats_meta_df: pd.DataFrame,
    analyzed_len: int,
    outlier_masks: dict[str, pd.Series],
    imputed_masks: dict[str, pd.Series],
    imputed_outlier_masks: dict[str, pd.Series],
    apply_imputation_colors: bool,
) -> None:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        output_df.to_excel(writer, sheet_name=ANALYSIS_SHEET, index=False)
        stats_detail_df.to_excel(writer, sheet_name=STATS_SHEET, index=False)
        stats_thresholds_df.to_excel(
            writer,
            sheet_name=STATS_SHEET,
            index=False,
            startrow=len(stats_detail_df) + 2,
        )
        stats_meta_df.to_excel(
            writer,
            sheet_name=STATS_SHEET,
            index=False,
            startrow=len(stats_detail_df) + len(stats_thresholds_df) + 4,
        )

        ws = writer.book[ANALYSIS_SHEET]
        header_map = {str(c.value): c.column for c in ws[1]}

        for spesa_col in SPESA_COLUMNS:
            col_idx = header_map.get(spesa_col)
            if col_idx is None:
                continue
            col_letter = get_column_letter(col_idx)

            outlier = outlier_masks[spesa_col]
            imputed = imputed_masks[spesa_col]
            imputed_outlier = imputed_outlier_masks[spesa_col]

            for row_idx in range(analyzed_len):
                excel_row = row_idx + 2
                if apply_imputation_colors:
                    if bool(imputed_outlier.iat[row_idx]):
                        ws[f"{col_letter}{excel_row}"].fill = ORANGE_FILL
                    elif bool(imputed.iat[row_idx]):
                        ws[f"{col_letter}{excel_row}"].fill = YELLOW_FILL
                    elif bool(outlier.iat[row_idx]):
                        ws[f"{col_letter}{excel_row}"].fill = RED_FILL
                elif bool(outlier.iat[row_idx]):
                    ws[f"{col_letter}{excel_row}"].fill = RED_FILL


def run(input_file: Path, output_file: Path, output_file_no_imputation: Path) -> None:
    if not input_file.exists():
        raise FileNotFoundError(f"File input non trovato: {input_file}")

    xls = pd.ExcelFile(input_file)
    first_sheet = xls.sheet_names[0]
    df = pd.read_excel(input_file, sheet_name=first_sheet)

    validate_columns(df)
    original_columns = df.columns.tolist()

    df = df.copy()
    df["_durata_num"] = pd.to_numeric(df[DURATION_COL], errors="coerce")
    df["_componenti_num"] = pd.to_numeric(df[COMPONENTI_COL], errors="coerce")
    df["_turisti_weight"] = build_tourist_weight_series(df["_componenti_num"])

    pacchetto_empty = df[PACCHETTO_COL].map(is_strict_empty)
    filter_mask = (df["_durata_num"] >= 1) & pacchetto_empty

    analyzed = df.loc[filter_mask].copy()
    analyzed.reset_index(drop=True, inplace=True)
    output_df = analyzed[original_columns].copy()
    output_df_no_imputation = analyzed[original_columns].copy()
    tourist_weights = analyzed["_turisti_weight"]

    analyzed["macro_provenienza"] = build_macro_provenienza(analyzed[STATO_COL])
    analyzed["motivazione_grp"] = analyzed[MOTIV_COL].astype("string").fillna("").str.strip().str.upper()

    outlier_masks: dict[str, pd.Series] = {}
    imputed_masks: dict[str, pd.Series] = {}
    imputed_outlier_masks: dict[str, pd.Series] = {}
    thresholds_by_spesa: dict[str, dict[tuple[str, str], ThresholdInfo]] = {}
    spend_stats: list[dict[str, object]] = []
    spend_stats_no_imputation: list[dict[str, object]] = []

    for spesa_col in SPESA_COLUMNS:
        analyzed[spesa_col] = analyzed[spesa_col].astype("object")
        output_df[spesa_col] = output_df[spesa_col].astype("object")
        output_df_no_imputation[spesa_col] = output_df_no_imputation[spesa_col].astype("object")
        imputation_col = f"imputazione_{spesa_col}"
        output_df[imputation_col] = pd.NA

        numeric_col = f"_{spesa_col}_num"
        analyzed[numeric_col] = analyzed[spesa_col].map(parse_number)
        outlier_metric_col = f"_{spesa_col}_procapite_giornaliera"
        analyzed[outlier_metric_col] = compute_daily_per_capita_series(
            analyzed[numeric_col],
            analyzed["_componenti_num"],
            analyzed["_durata_num"],
        )
        metric_before = analyzed[outlier_metric_col].copy()

        nd_mask = analyzed[spesa_col].map(is_null_like)

        thresholds = compute_thresholds(
            analyzed,
            value_col=outlier_metric_col,
            group_cols=["macro_provenienza", "motivazione_grp"],
            weights_col="_turisti_weight",
        )
        thresholds_by_spesa[spesa_col] = thresholds

        is_outlier = pd.Series(False, index=analyzed.index)
        for i, row in analyzed.iterrows():
            key = (row["macro_provenienza"], row["motivazione_grp"])
            info = thresholds.get(key, ThresholdInfo(p5=None, p95=None, n=0, n_turisti=0))

            val = row[outlier_metric_col]
            if val is None or pd.isna(val):
                continue
            if info.p5 is None or info.p95 is None:
                continue
            if val < info.p5 or val > info.p95:
                is_outlier.at[i] = True

        outlier_masks[spesa_col] = is_outlier

        imputed_mask = pd.Series(False, index=analyzed.index)
        imputed_outlier_mask = pd.Series(False, index=analyzed.index)
        nd_indices = analyzed.index[nd_mask].tolist()

        for idx in nd_indices:
            inferred, donor_id, method = infer_value_from_similar(
                analyzed,
                row_idx=idx,
                value_col=numeric_col,
            )
            if inferred is None:
                continue

            inferred_rounded = round(float(inferred), 2)
            analyzed.at[idx, spesa_col] = inferred_rounded
            output_df.at[idx, spesa_col] = inferred_rounded
            analyzed.at[idx, numeric_col] = inferred_rounded
            inferred_metric = compute_daily_per_capita_value(
                inferred_rounded,
                analyzed.at[idx, "_componenti_num"],
                analyzed.at[idx, "_durata_num"],
            )
            analyzed.at[idx, outlier_metric_col] = inferred_metric
            output_df.at[idx, imputation_col] = format_imputation_note(method, donor_id)
            imputed_mask.at[idx] = True

            key = (analyzed.at[idx, "macro_provenienza"], analyzed.at[idx, "motivazione_grp"])
            info = thresholds.get(key, ThresholdInfo(p5=None, p95=None, n=0, n_turisti=0))
            if info.p5 is not None and info.p95 is not None and inferred_metric is not None:
                if inferred_metric < info.p5 or inferred_metric > info.p95:
                    imputed_outlier_mask.at[idx] = True

        imputed_masks[spesa_col] = imputed_mask
        imputed_outlier_masks[spesa_col] = imputed_outlier_mask

        n_valori_numerici = weighted_count(metric_before.notna(), tourist_weights)
        n_imputed = weighted_count(imputed_mask, tourist_weights)
        n_imputed_outlier = weighted_count(imputed_outlier_mask, tourist_weights)
        n_nd = weighted_count(nd_mask, tourist_weights)
        n_nd_non_imputati = weighted_count(nd_mask & (~imputed_mask), tourist_weights)
        n_outliers = weighted_count(is_outlier, tourist_weights)
        spend_stats.append(
            {
                "colonna_spesa": spesa_col,
                "n_valori_numerici": n_valori_numerici,
                "n_nd_iniziali": n_nd,
                "n_imputati": n_imputed,
                "n_imputati_outlier": n_imputed_outlier,
                "n_nd_non_imputati": n_nd_non_imputati,
                "n_outlier": n_outliers,
            }
        )

        spend_stats_no_imputation.append(
            {
                "colonna_spesa": spesa_col,
                "n_valori_numerici": n_valori_numerici,
                "n_nd_iniziali": n_nd,
                "n_imputati": 0,
                "n_imputati_outlier": 0,
                "n_nd_non_imputati": n_nd,
                "n_outlier": n_outliers,
            }
        )

    if outlier_masks:
        any_outlier_mask = pd.concat(outlier_masks.values(), axis=1).any(axis=1)
    else:
        any_outlier_mask = pd.Series(False, index=analyzed.index)
    output_df_no_imputation["outlier_presente"] = np.where(any_outlier_mask, "SI", "NO")

    full_tourists = format_count_value(float(df["_turisti_weight"].sum()))
    analyzed_tourists = format_count_value(float(analyzed["_turisti_weight"].sum()))

    stats_meta_df, stats_detail_df = build_stats_table(
        analyzed_tourists=analyzed_tourists,
        full_tourists=full_tourists,
        spend_stats=spend_stats,
    )
    _, stats_detail_no_imputation_df = build_stats_table(
        analyzed_tourists=analyzed_tourists,
        full_tourists=full_tourists,
        spend_stats=spend_stats_no_imputation,
    )
    stats_thresholds_df = build_thresholds_stats_table(thresholds_by_spesa)

    if analyzed.empty:
        print("Nessun questionario soddisfa il filtro: durata_soggiorno >= 1 e pacchetto vuoto.")

    write_output_excel(
        output_file=output_file,
        output_df=output_df,
        stats_detail_df=stats_detail_df,
        stats_thresholds_df=stats_thresholds_df,
        stats_meta_df=stats_meta_df,
        analyzed_len=int(len(analyzed)),
        outlier_masks=outlier_masks,
        imputed_masks=imputed_masks,
        imputed_outlier_masks=imputed_outlier_masks,
        apply_imputation_colors=True,
    )

    write_output_excel(
        output_file=output_file_no_imputation,
        output_df=output_df_no_imputation,
        stats_detail_df=stats_detail_no_imputation_df,
        stats_thresholds_df=stats_thresholds_df,
        stats_meta_df=stats_meta_df,
        analyzed_len=int(len(analyzed)),
        outlier_masks=outlier_masks,
        imputed_masks=imputed_masks,
        imputed_outlier_masks=imputed_outlier_masks,
        apply_imputation_colors=False,
    )

    total_analyzed = format_count_value(float(analyzed["_turisti_weight"].sum()))
    total_outliers = {
        c: weighted_count(outlier_masks[c], tourist_weights)
        for c in SPESA_COLUMNS
    }
    total_imputed = {
        c: weighted_count(imputed_masks[c], tourist_weights)
        for c in SPESA_COLUMNS
    }
    total_imputed_outlier = {
        c: weighted_count(imputed_outlier_masks[c], tourist_weights)
        for c in SPESA_COLUMNS
    }
    total_not_imputed = {
        c: weighted_count(analyzed[c].map(is_null_like), tourist_weights)
        for c in SPESA_COLUMNS
    }

    print(f"Input letto: {input_file}")
    print(f"Foglio sorgente: {first_sheet}")
    print(f"Turisti analizzati (filtro applicato): {total_analyzed}")
    print("Outlier per colonna:")
    for c, n in total_outliers.items():
        print(f"- {c}: {n}")

    print("ND imputati per colonna:")
    for c, n in total_imputed.items():
        print(f"- {c}: {n}")

    print("ND imputati che risultano outlier per colonna:")
    for c, n in total_imputed_outlier.items():
        print(f"- {c}: {n}")

    print("ND rimasti non imputati per colonna:")
    for c, n in total_not_imputed.items():
        print(f"- {c}: {n}")

    print(f"Output generato: {output_file}")
    print(f"Output generato (senza imputazioni): {output_file_no_imputation}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Rilevazione outlier (p5-p95) su spesa pro capite giornaliera "
            "(spesa/(numero_componenti*durata_soggiorno)) e imputazione ND sulle spese "
            "per pernottanti senza pacchetto."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="File Excel sorgente")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="File Excel di output")
    parser.add_argument(
        "--output-no-imputation",
        type=Path,
        default=DEFAULT_OUTPUT_NO_IMPUTATION,
        help="File Excel di output parallelo senza imputazioni ND",
    )
    args = parser.parse_args()

    run(args.input, args.output, args.output_no_imputation)


if __name__ == "__main__":
    main()
