from __future__ import annotations

import argparse
import os
import re
import unicodedata
from functools import lru_cache
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tqdm.auto import tqdm
from build_analisi_preliminare import extract_prevalent_destination

DEFAULT_INPUT = Path("questionari_sottocampioni.xlsx")
DEFAULT_OUTPUT = Path("analisi_qualitative_campione_1.xlsx")
DEFAULT_SHEET = "campione_1"
DEFAULT_SHEET_1D = "campione_1d"

NULL_VALUE = "NULL"
UNDEFINED_LABEL = "NON DEFINITO"
NULL_TOKENS = {"", "ND", "NR", "NA", "N/D", "N.R.", "N.A.", "NULL", "NULLO", "VUOTO", "VUOTA"}

ID_COL = "ID"
AGE_COL = "fascia_età"
STATE_COL = "stato_provenienza"
REGION_COL = "regione_provenienza"
ARRIVAL_COL = "arrivo_sardegna"
LOCATIONS_COL = "località_visitate"
DURATION_COL = "durata_soggiorno"
COMPONENTS_COL = "numero_componenti"
PRIMARY_REASON_COL = "motivazione_principale"
SECONDARY_REASON_COL = "motivazione_secondaria"
LODGING_COL = "tipologia_sistemazione"
WEB_COL = "web"
WEB_DETAIL_COL = "web_specifica"
STRENGTHS_COL = "apprezzamenti"
WEAKNESSES_COL = "miglioramenti"
JUDGMENT_COL = "giudizio"
FUTURE_SPEND_COL = "previsione_spesa"

REQUIRED_COLUMNS = [
    ID_COL,
    AGE_COL,
    STATE_COL,
    REGION_COL,
    ARRIVAL_COL,
    LOCATIONS_COL,
    DURATION_COL,
    COMPONENTS_COL,
    PRIMARY_REASON_COL,
    SECONDARY_REASON_COL,
    LODGING_COL,
    WEB_COL,
    WEB_DETAIL_COL,
    STRENGTHS_COL,
    WEAKNESSES_COL,
    JUDGMENT_COL,
    FUTURE_SPEND_COL,
]

JUDGMENT_SCORE_MAP = {
    "PESSIMO": 1,
    "SUFFICIENTE": 2,
    "BUONO": 3,
    "OTTIMO": 4,
}

TEXT_SPLIT_RE = re.compile(r"[;,/|\n]+|\be\b|\bed\b|\by\b", flags=re.IGNORECASE)
NEGATION_PATTERNS = [
    r"\b(nulla|niente|nessuno|nessuna|no|non saprei|non so|na|nd|nr)\b",
    r"\b(nulla da migliorare|niente da migliorare|va bene cosi|va bene cosi'|tutto bene|nessun miglioramento)\b",
    r"^\s*(tutto|tutta|tutti|tutte)\s*$",
    r"^\s*(niente|nulla)\s*$",
    r"^\s*(tutto quanto|tutto q\.?to|tutto ok|tutto perfetto)\s*$",
]
WEB_NEGATION_PATTERNS = [
    r"\b(non ho prenotato|nessuna prenotazione|non prenotato|non usato il web)\b",
]

WEB_THEME_PATTERNS: list[tuple[str, str]] = [
    ("VOLO/AEREO", r"\b(vol\w*|aere\w*|flight\w*)\b"),
    ("NAVE/TRAGHETTO", r"\b(nav\w*|traghett\w*|ferr\w*)\b"),
    ("ALLOGGIO", r"\b(allogg\w*|hotel\w*|b&b|appartament\w*|casa vacanz\w*|campegg\w*|residence\w*|agriturism\w*|ostell\w*)\b"),
    ("AUTONOLEGGIO/TRASPORTI", r"\b(auto\b|nolegg\w*|scooter\w*|moto\w*|bus\w*|autobus\w*|tren\w*|transfer\w*|taxi\w*)\b"),
    ("ESCURSIONI/ATTIVITA", r"\b(escursion\w*|git\w*|tour\w*|barc\w*|boat\w*|attivit\w*|esperienz\w*|bigliett\w*|ticket\w*|muse\w*|parc\w*|visit\w*)\b"),
    ("RISTORANTI", r"\b(ristor\w*|pizzeri\w*|cen\w*|pranz\w*|food\w*|colazion\w*)\b"),
    ("ATTREZZATURA MARE", r"\b(ombrellon\w*|lettin\w*|sdrai\w*|attrezzatura per il mare)\b"),
]

THEME_PATTERNS: list[tuple[str, str]] = [
    ("MARE/SPIAGGE", r"\b(mar\w*|spiagg\w*|acqu\w*|cost\w*|litoral\w*|cal\w*)\b"),
    ("NATURA/PAESAGGIO", r"\b(natur\w*|paesagg\w*|panoram\w*|montagn\w*|parc\w*|verd\w*)\b"),
    ("CLIMA", r"\b(clim\w*|meteo\w*|tempo\b)\b"),
    ("OSPITALITA/ACCOGLIENZA", r"\b(ospitalit\w*|ospital\w*|accoglien\w*|gentilezz\w*|cordial\w*)\b"),
    ("CIBO/ENOGASTRONOMIA", r"\b(cib\w*|cucin\w*|ristor\w*|mang\w*|enogastr\w*|gastron\w*|food\w*)\b"),
    ("TRANQUILLITA/RELAX", r"\b(tranquill\w*|relax\w*|silenz\w*|pac\w*|calm\w*)\b"),
    ("DIVERTIMENTO/VITA NOTTURNA", r"\b(divert\w*|movid\w*|local\w*|notturn\w*|serat\w*|mojito\w*)\b"),
    ("CULTURA/BORGHI", r"\b(cultur\w*|muse\w*|stori\w*|borgh\w*|artist\w*|architett\w*)\b"),
]

WEAKNESS_PATTERNS: list[tuple[str, str]] = [
    ("VIABILITA/STRADE", r"\b(strad\w*|viabil\w*|cantier\w*|asfalt\w*)\b"),
    ("TRASPORTI/COLLEGAMENTI", r"\b(collegament\w*|trasport\w*|autobus\w*|bus\w*|tren\w*|vol\w* dirett\w*|traghett\w*|aere\w*|taxi\w*|mezzi pubblici|mezz\w*)\b"),
    ("PARCHEGGI", r"\b(parchegg\w*)\b"),
    ("PREZZI/COSTI", r"\b(prezz\w*|car\w*|cost\w*)\b"),
    ("SERVIZI DIGITALI/INFORMAZIONI", r"\b(online\w*|sit\w*|app\b|informazion\w*|servizi online|servizi digitali)\b"),
    ("PULIZIA/RIFIUTI", r"\b(pulizi\w*|rifiut\w*|cassonett\w*|sporc\w*)\b"),
    ("SOVRAFFOLLAMENTO", r"\b(affoll\w*|confusion\w*|caos\w*|tropp\w* turisti\w*)\b"),
    ("SPIAGGE/ACCESSIBILITA", r"\b(spiagg\w*|access\w*|accessibil\w*|prenotazion\w*|barrier\w* architettonic\w*)\b"),
]

SENTIMENT_THEME_DEFINITIONS: list[tuple[str, str]] = [
    (
        "Mare e paesaggio",
        "riferimenti espliciti alla bellezza del mare, alla qualità delle spiagge, ai paesaggi costieri e naturali",
    ),
    (
        "Accoglienza e ospitalità",
        "valutazioni relative alla cordialità degli operatori turistici, alla disponibilità della popolazione locale, dei residenti e dei sardi, al clima relazionale",
    ),
    (
        "Enogastronomia",
        "giudizi sulla qualità dei prodotti alimentari locali, sull’autenticità della cucina tradizionale, sulla varietà dell’offerta gastronomica",
    ),
    (
        "Clima",
        "riferimenti alle condizioni meteorologiche e climatiche",
    ),
    (
        "Autenticità dei luoghi",
        "percezione dell’autenticità culturale, della preservazione delle tradizioni, dell’identità territoriale",
    ),
    (
        "Rapporto qualità/prezzo",
        "valutazioni comparative tra il livello dei servizi fruiti e i costi sostenuti",
    ),
    (
        "Servizi turistici",
        "giudizi sull’organizzazione complessiva dei servizi, come informazione turistica, segnaletica, accessibilità, pulizia",
    ),
    (
        "Infrastrutture e viabilità",
        "valutazioni sulla qualità delle strade, sulla disponibilità di parcheggi, sull’adeguatezza delle infrastrutture",
    ),
    (
        "Costi elevati",
        "segnalazioni esplicite di prezzi ritenuti eccessivi o non proporzionati al servizio ricevuto",
    ),
    (
        "Carenze nei servizi",
        "indicazioni di disservizi, mancanza di servizi attesi, inefficienze organizzative",
    ),
]

SENTIMENT_THEME_LABELS = [label for label, _ in SENTIMENT_THEME_DEFINITIONS]
SENTIMENT_THEME_CANDIDATES = [
    f"{label} - {description}" for label, description in SENTIMENT_THEME_DEFINITIONS
]
SENTIMENT_THEME_LABEL_MAP = {
    candidate: label for (label, _), candidate in zip(SENTIMENT_THEME_DEFINITIONS, SENTIMENT_THEME_CANDIDATES)
}

EXPLICIT_THEME_PATTERNS: list[tuple[str, str]] = [
    ("Mare e paesaggio", r"\b(mare|spiagg\w*|litor\w*|costa|costier\w*|scoglier\w*|calett\w*|panoram\w*|paesagg\w*|natur\w*)\b"),
    ("Accoglienza e ospitalità", r"\b(accoglienz\w*|ospitalit\w*|ospital\w*|gentilezz\w*|cordial\w*|disponibil\w*|relazional\w*|resident\w*|abitant\w*|popolazion\w* locale|gente del posto|persone del posto|sard(?:o|a|i|e))\b"),
    ("Enogastronomia", r"\b(cibo|cucin\w*|mang\w*|ristoran\w*|gastron\w*|enogastr\w*|prodott\w* locali|piatt\w* tipic\w*)\b"),
    ("Clima", r"\b(clima|meteo|temperatur\w*|vento|soleggiat\w*|caldo|fresc\w*|umid\w*|pioggi\w*)\b"),
    ("Autenticità dei luoghi", r"\b(autentic\w*|tradizion\w*|identit\w*|cultural\w*|borg\w*|tipic\w*|genuin\w*)\b"),
    ("Rapporto qualità/prezzo", r"\b(qualit[aà]\s*/?\s*prezzo|rapporto qualit[aà]|buon prezzo|prezzo giusto|convenient\w*|caro ma)\b"),
    ("Servizi turistici", r"\b(segnaletic\w*|informazion\w* turistic\w*|ufficio turistico|accessibil\w*|pulizi\w*|serviz\w* turistic\w*|organizzazion\w*)\b"),
    ("Infrastrutture e viabilità", r"\b(strad\w*|viabil\w*|parchegg\w*|infrastruttur\w*|collegament\w*|traffico|asfalt\w*)\b"),
    ("Costi elevati", r"\b(prezz\w* alt\w*|cost\w* eccessiv\w*|troppo car\w*|carissim\w*|prezzi eccessiv\w*)\b"),
    ("Carenze nei servizi", r"\b(disserviz\w*|mancanz\w* di servizi|servizi assenti|inefficien\w*|servizi carent\w*|manca\w*)\b"),
]

DEFAULT_ZERO_SHOT_MODEL = os.environ.get("QUAL_SENTIMENT_MODEL", "MoritzLaurer/mDeBERTa-v3-base-mnli-xnli")
ZERO_SHOT_FALLBACK_MODELS = [
    DEFAULT_ZERO_SHOT_MODEL,
    "joeddav/xlm-roberta-large-xnli",
]
DEFAULT_ZERO_SHOT_THRESHOLD = float(os.environ.get("QUAL_SENTIMENT_THRESHOLD", "0.30"))
DEFAULT_ZERO_SHOT_MAX_LABELS = int(os.environ.get("QUAL_SENTIMENT_MAX_LABELS", "1"))
DEFAULT_ZERO_SHOT_BATCH_SIZE = int(os.environ.get("QUAL_SENTIMENT_BATCH_SIZE", "16"))
ZERO_SHOT_HYPOTHESIS_TEMPLATE = "Questo testo riguarda {}."
ZERO_SHOT_METHOD_LABEL = "regola_esplicita_più_ml_zero_shot_huggingface"
HF_CACHE_DIR = Path(os.environ.get("HF_HOME", Path(__file__).resolve().parent / ".hf_cache"))

STOPWORDS = {
    "il", "lo", "la", "i", "gli", "le", "un", "una", "uno", "di", "a", "da", "in", "su", "per", "con",
    "del", "della", "dello", "dei", "degli", "delle", "al", "alla", "allo", "ai", "agli", "alle", "ed",
    "e", "o", "the", "and", "to", "of", "is", "are", "non", "piu", "più", "molto", "molti", "molte",
    "nulla", "niente", "tutto", "tutti", "tutte", "same", "good", "nice", "very",
}

LEADING_FILLER_PATTERNS = [
    r"^(?:il|lo|la|i|gli|le|un|uno|una|l)\s+",
    r"^(?:l')",
    r"^(?:del|della|dello|dei|degli|delle|al|allo|alla|ai|agli|alle|dal|dallo|dalla|dai|dagli|dalle)\s+",
    r"^(?:di|da|da parte di|con|per|su|in)\s+",
    r"^(?:mi e piaciut\w*|ho apprezzat\w*|mi sono piaciut\w*|bella\b|bello\b|belli\b|belle\b|ottim\w*|buon\w*)\s+",
]

FRAGMENT_LEADING_PATTERNS = [
    r"^(?:mi e piaciut\w*|mi sono piaciut\w*|mi piace\w*|ho apprezzat\w*|ho gradit\w*|ho amat\w*|ador\w*|amo)\s+",
    r"^(?:mi sono trovat\w* bene|mi son trovat\w* bene|mi sono trovat\w* benissimo|mi son trovat\w* benissimo)\s+",
    r"^(?:quello che ho apprezzato di piu|quello che mi e piaciuto di piu|la cosa migliore|la cosa piu bella|la cosa che mi e piaciuta di piu)\s+",
    r"^(?:molto |davvero |veramente |particolarmente |soprattutto )+",
]

FRAGMENT_TRAILING_PATTERNS = [
    r"\s+(?:bellissim\w*|stupend\w*|splendid\w*|fantastic\w*|meraviglios\w*|ottim\w*|buon\w*|eccellent\w*|top)$",
]

GENERIC_FRAGMENT_TOKENS = {
    "bello", "bella", "belli", "belle", "bellissimo", "bellissima", "bellissimi", "bellissime",
    "buono", "buona", "buoni", "buone", "ottimo", "ottima", "ottimi", "ottime",
    "stupendo", "stupenda", "stupendi", "stupende", "fantastico", "fantastica", "fantastici", "fantastiche",
    "molto", "davvero", "veramente", "particolarmente", "soprattutto",
}

TOKEN_NORMALIZATION_MAP = {
    "spiagge": "spiaggia",
    "spiaggia": "spiaggia",
    "mari": "mare",
    "paesaggi": "paesaggio",
    "coste": "costa",
    "costieri": "costiero",
    "costiere": "costiero",
    "naturali": "naturale",
    "operatori": "operatore",
    "operatori turistici": "operatore turistico",
    "prodotti": "prodotto",
    "prodotti locali": "prodotto locale",
    "piatti": "piatto",
    "piatti tipici": "piatto tipico",
    "tradizioni": "tradizione",
    "luoghi": "luogo",
    "servizi": "servizio",
    "servizi turistici": "servizio turistico",
    "strade": "strada",
    "parcheggi": "parcheggio",
    "infrastrutture": "infrastruttura",
    "prezzi": "prezzo",
    "costi": "costo",
    "disservizi": "disservizio",
    "inefficienze": "inefficienza",
    "carenze": "carenza",
    "cordiali": "cordiale",
    "gentili": "gentile",
    "disponibili": "disponibile",
    "costose": "costoso",
    "costosi": "costoso",
    "costosa": "costoso",
    "cari": "caro",
    "care": "caro",
    "cara": "caro",
    "autentica": "autentico",
    "autentici": "autentico",
    "autentiche": "autentico",
    "turistiche": "turistico",
    "turistici": "turistico",
    "turistica": "turistico",
}

PHRASE_NORMALIZATION_MAP = {
    "prodotti locali": "prodotto locale",
    "piatti tipici": "piatto tipico",
    "servizi turistici": "servizio turistico",
    "operatori turistici": "operatore turistico",
}

INVARIABLE_TOKEN_ENDINGS = (
    "zione", "zioni", "mente", "ita", "ista", "iste", "isti", "ale", "ali", "ile", "ili",
)

NON_INFORMATIVE_EXACT_PATTERNS = [
    r"^\s*$",
    r"^\s*(null|nulla|nullo|nullo|vuoto|vuota|nessuna risposta|nessun commento)\s*$",
    r"^\s*(tutto|tutta|tutti|tutte)\s*$",
    r"^\s*(niente|nulla)\s*$",
    r"^\s*(tutto bene|tutto ok|tutto perfetto|tutto bellissimo|tutto bello)\s*$",
    r"^\s*(niente da migliorare|nulla da migliorare|nessun miglioramento)\s*$",
    r"^\s*(boh|non so|non saprei)\s*$",
]

TEXT_NORMALIZATION_REPLACEMENTS = [
    (r"\bb b\b", "b&b"),
    (r"\bbnb\b", "b&b"),
    (r"\bbed and breakfast\b", "b&b"),
    (r"\bcasa vacanza\b", "casa vacanze"),
    (r"\bcase vacanza\b", "casa vacanze"),
    (r"\bautonoleggio\b", "noleggio auto"),
    (r"\bauto noleggio\b", "noleggio auto"),
    (r"\bvoli diretti\b", "voli diretti"),
    (r"\bmezzi pubblici\b", "mezzi pubblici"),
    (r"\bspiaggie\b", "spiagge"),
    (r"\btranquilit[aà]\b", "tranquillita"),
]

DESTINATION_VARIANT_MAP = {
    "AGHERO": "ALGHERO",
    "ALGHERO.": "ALGHERO",
    "BAIA SARDINIA": "BAIA SARDINIA",
    "BAJA SARDINIA": "BAIA SARDINIA",
    "LOIRI PORTO S. PAOLO": "LOIRI PORTO SAN PAOLO",
    "LOIRI PORTO S PAOLO": "LOIRI PORTO SAN PAOLO",
    "LOIRI PORTO SAN PAOLO": "LOIRI PORTO SAN PAOLO",
    "BARI SARDO": "BARISARDO",
    "BARISARDO": "BARISARDO",
    "BUGERRU": "BUGGERRU",
    "CABONIA": "CARBONIA",
    "CAAGLIARI": "CAGLIARI",
    "CAGLARI": "CAGLIARI",
    "CAGLIAR": "CAGLIARI",
    "CAGLIRI": "CAGLIARI",
    "ORISTNO": "ORISTANO",
    "CARDEDDU": "CARDEDU",
    "CARLO FORTE": "CARLOFORTE",
    "CASTELASARDO": "CASTELSARDO",
    "CAOTERRA": "CAPOTERRA",
    "SAN GAVINO": "SAN GAVINO MONREALE",
    "SANT TEODORO": "SAN TEODORO",
    "SANT'ANNA ARRSI": "SANT'ANNA ARRESI",
    "SANT'ANTIIOCO": "SANT'ANTIOCO",
    "SANTA TERESA": "SANTA TERESA DI GALLURA",
    "S. TERESA DI GALLURA": "SANTA TERESA DI GALLURA",
    "S. TERESA": "SANTA TERESA DI GALLURA",
    "SANTA TERSA DI GALLURA": "SANTA TERESA DI GALLURA",
    "S. TEODORO": "SAN TEODORO",
    "S. ANTIOCO": "SANT'ANTIOCO",
    "S. MARIA LA PALMA": "SANTA MARIA LA PALMA",
    "S. PANTALEO": "SAN PANTALEO",
    "QUARTU SANT' ELENA": "QUARTU SANT'ELENA",
    "QUARTU SANT'ELANA": "QUARTU SANT'ELENA",
    "DORGALIA": "DORGALI",
    "DOMUS DE MARI": "DOMUS DE MARIA",
    "LA MADDALEMA": "LA MADDALENA",
    "ORGOSOLO": "ORGOSOLO",
    "OROGOSOLO": "ORGOSOLO",
    "PORTO SCUSO": "PORTOSCUSO",
    "PORTUSCUSO": "PORTOSCUSO",
    "PORTO TORRESE": "PORTO TORRES",
    "SANTA MARIA NAVVARESE": "SANTA MARIA NAVARRESE",
    "SANTA MARIA NOVARRESE": "SANTA MARIA NAVARRESE",
    "SILIQUIA": "SILIQUA",
    "TORTOLIE": "TORTOLI",
    "TORTOLII": "TORTOLI",
    "TORTOLÌ": "TORTOLI",
    "VALLANOVA TULO": "VILLANOVATULO",
    "VILLANOVATULO": "VILLANOVATULO",
    "MURAVAERA": "MURAVERA",
    "IGLIESIA": "IGLESIAS",
    "IGLIESIAS": "IGLESIAS",
    "TERTENITA": "TERTENIA",
    "VILLASIMUS": "VILLASIMIUS",
    "VIALLASIMIUS": "VILLASIMIUS",
    "CALGIARI": "CAGLIARI",
    "FLUMINI MAGGIORE": "FLUMINIMAGGIORE",
    "TRINITÀ D'AGULTU E VIGNOLA": "TRINITÀ D'AGULTU",
    "TRINITA D'AGULTU E VIGNOLA": "TRINITA D'AGULTU",
    "TRINITA D AGULTU E VIGNOLA": "TRINITA D'AGULTU",
}


def normalize_text_series(series: pd.Series) -> pd.Series:
    s = series.astype("string").fillna("").str.strip()
    upper = s.str.upper()
    null_upper = {token.upper() for token in NULL_TOKENS}
    return s.mask(upper.isin(null_upper), NULL_VALUE)


def normalize_upper_series(series: pd.Series) -> pd.Series:
    return normalize_text_series(series).str.upper()


def validate_columns(df: pd.DataFrame) -> None:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti nel dataset: {missing}")


def strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def sanitize_location_text(raw_text: str) -> str:
    text = raw_text.strip()
    text = re.sub(r"(?<=[A-Za-zÀ-ÿ])(?=\d+(?:\s*(?:;|,|:|$)))", " ", text)
    return text


def normalize_destination_name(destination_raw: str) -> str:
    destination_clean = re.sub(r"\s*\([^)]*\)\s*", " ", destination_raw)
    destination_clean = re.sub(r"^(?:\s*\d+\s*:\s*)+", "", destination_clean)
    destination_clean = re.sub(r",\s*\d+[A-Za-z]+\d*\s*$", "", destination_clean)
    destination_clean = re.sub(r"\s+\d+\s*$", "", destination_clean)
    destination_clean = re.sub(r"\.+$", "", destination_clean)
    destination_clean = re.sub(r"\s+", " ", destination_clean).strip(" ,;")
    if not destination_clean:
        return ""

    destination_upper = destination_clean.upper()
    destination_upper = DESTINATION_VARIANT_MAP.get(destination_upper, destination_upper)
    if destination_upper == destination_clean.upper():
        destination_ascii = strip_accents(destination_upper)
        destination_upper = DESTINATION_VARIANT_MAP.get(destination_ascii, destination_upper)
    return destination_upper


def extract_location_without_valid_days(segment: str) -> str:
    cleaned_segment = sanitize_location_text(segment.strip())
    cleaned_segment = re.sub(r",\s*\d+[A-Za-z]+(?:\s*\d+)?\s*$", "", cleaned_segment)
    cleaned_segment = re.sub(r":\s*\d+[A-Za-z]+(?:\s*\d+)?\s*$", "", cleaned_segment)
    return normalize_destination_name(cleaned_segment)


def extract_prevalent_destination_local(value: object) -> str:
    pairs = extract_destination_days(value)
    if pairs:
        totals: dict[str, int] = {}
        first_seen: dict[str, int] = {}
        for idx, (dest, days) in enumerate(pairs):
            if dest not in first_seen:
                first_seen[dest] = idx
            totals[dest] = totals.get(dest, 0) + days

        positive_totals = {dest: total for dest, total in totals.items() if total > 0}
        if positive_totals:
            return min(
                positive_totals.keys(),
                key=lambda k: (-positive_totals[k], first_seen[k], k),
            )
        return pairs[0][0]

    prevalent = extract_prevalent_destination(value)
    if prevalent == NULL_VALUE:
        if pd.isna(value):
            return NULL_VALUE
        raw_text = str(value).strip()
        if not raw_text or raw_text.upper() == NULL_VALUE:
            return NULL_VALUE
        raw_text = sanitize_location_text(raw_text)
        if raw_text.startswith("(") and raw_text.endswith(")"):
            raw_text = raw_text[1:-1].strip()
        if ";" not in raw_text:
            single_destination = normalize_destination_name(raw_text)
            if single_destination:
                return single_destination
        return NULL_VALUE

    return normalize_destination_name(prevalent)


def normalize_free_text(value: object) -> str:
    if pd.isna(value):
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    upper = raw.upper()
    if upper in {token.upper() for token in NULL_TOKENS}:
        return ""
    text = strip_accents(raw.lower())
    for pattern, repl in TEXT_NORMALIZATION_REPLACEMENTS:
        text = re.sub(pattern, repl, text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def prettify_value(value: object) -> str:
    if pd.isna(value):
        return UNDEFINED_LABEL
    text = str(value).strip()
    if not text:
        return UNDEFINED_LABEL
    if text.upper() == NULL_VALUE:
        return UNDEFINED_LABEL
    return text


def build_macro_provenienza(state_series: pd.Series) -> pd.Series:
    upper = normalize_upper_series(state_series)
    out = pd.Series(UNDEFINED_LABEL, index=upper.index, dtype="string")
    out = out.mask(upper.eq("ITALIA"), "ITALIANI")
    out = out.mask((upper != "") & (upper != NULL_VALUE) & (~upper.eq("ITALIA")), "STRANIERI")
    return out


def build_provenienza_label(df: pd.DataFrame) -> pd.Series:
    macro = df["macro_provenienza"]
    state = normalize_upper_series(df[STATE_COL])
    region = normalize_upper_series(df[REGION_COL])

    out = pd.Series(UNDEFINED_LABEL, index=df.index, dtype="string")
    out = out.mask(macro.eq("ITALIANI") & region.ne(NULL_VALUE), region)
    out = out.mask(macro.eq("STRANIERI") & state.ne(NULL_VALUE), state)
    return out


def extract_destination_days(value: object) -> list[tuple[str, int]]:
    if pd.isna(value):
        return []

    raw_text = sanitize_location_text(str(value).strip())
    if not raw_text or raw_text.upper() == NULL_VALUE:
        return []

    if raw_text.startswith("(") and raw_text.endswith(")"):
        raw_text = raw_text[1:-1].strip()

    pair_pattern = re.compile(
        r"\s*([^\d,;:][^,;:]*?)\s*(?:,+\s*|:\s*|\s*)?([0-9]+(?:[.,][0-9]+)?)(?![A-Za-z])\s*(?=;|,|:|$)"
    )
    matches = list(pair_pattern.finditer(raw_text))

    results: list[tuple[str, int]] = []
    for match in matches:
        destination_raw = match.group(1).strip()
        days_raw = match.group(2).strip()
        destination_clean = normalize_destination_name(destination_raw)
        if not destination_clean:
            continue
        if destination_clean[0].isdigit():
            continue
        try:
            days_float = float(days_raw.replace(",", "."))
        except ValueError:
            continue
        if not days_float.is_integer():
            continue
        days_int = int(days_float)
        if days_int < 0:
            continue
        results.append((destination_clean, days_int))

    if ";" in raw_text:
        parsed_destinations = {dest for dest, _ in results}
        for segment in raw_text.split(";"):
            segment = segment.strip()
            if not segment:
                continue
            strict_match = re.fullmatch(
                r"\s*([^\d,;:][^,;:]*?)\s*(?:,+\s*|:\s*|\s*)?([0-9]+(?:[.,][0-9]+)?)\s*",
                segment,
            )
            if strict_match:
                continue
            fallback_destination = extract_location_without_valid_days(segment)
            if fallback_destination and fallback_destination not in parsed_destinations:
                results.append((fallback_destination, 0))
                parsed_destinations.add(fallback_destination)

    if results:
        return results

    if ";" not in raw_text and not re.search(r"\d", raw_text):
        single_destination = normalize_destination_name(raw_text)
        if single_destination:
            return [(single_destination, 1)]

    return []


def build_destination_metrics(locations_series: pd.Series) -> pd.DataFrame:
    prevalent: list[str] = []
    unique_positive_counts: list[int] = []
    travel_type: list[str] = []

    for value in locations_series.tolist():
        pairs = extract_destination_days(value)
        if not pairs:
            prevalent.append(extract_prevalent_destination_local(value))
            unique_positive_counts.append(0)
            travel_type.append(UNDEFINED_LABEL)
            continue

        totals: dict[str, int] = {}
        first_seen: dict[str, int] = {}
        positive_destinations: list[str] = []
        for idx, (dest, days) in enumerate(pairs):
            if dest not in first_seen:
                first_seen[dest] = idx
            totals[dest] = totals.get(dest, 0) + days
            if days > 0:
                positive_destinations.append(dest)

        prev = extract_prevalent_destination_local(value)
        if prev == NULL_VALUE:
            prev = UNDEFINED_LABEL
        if prev not in totals:
            prev = min(totals.keys(), key=lambda k: (-totals[k], first_seen[k], k))
        positive_unique = len(dict.fromkeys(positive_destinations))
        total_unique = len(totals)

        prevalent.append(prev)
        if positive_unique <= 0:
            unique_positive_counts.append(total_unique)
            if total_unique <= 1:
                travel_type.append("STANZIALE")
            else:
                travel_type.append("ITINERANTE")
        else:
            unique_positive_counts.append(positive_unique)
            if positive_unique == 1:
                travel_type.append("STANZIALE")
            else:
                travel_type.append("ITINERANTE")

    return pd.DataFrame(
        {
            "destinazione_prevalente": pd.Series(prevalent, dtype="string"),
            "numero_destinazioni_positive": pd.Series(unique_positive_counts, dtype="int64"),
            "tipo_turismo": pd.Series(travel_type, dtype="string"),
        }
    )


def normalize_judgment(series: pd.Series) -> pd.Series:
    return normalize_upper_series(series)


def normalize_yes_no(series: pd.Series) -> pd.Series:
    upper = normalize_upper_series(series)
    out = pd.Series(UNDEFINED_LABEL, index=upper.index, dtype="string")
    out = out.mask(upper.isin({"SI", "SÌ"}), "SI")
    out = out.mask(upper.eq("NO"), "NO")
    return out


def split_apprezzamenti_1d(text: str) -> list[str]:
    parts = re.split(r"[;,/|]+", text)
    out: list[str] = []
    for part in parts:
        cleaned = re.sub(r"\s+", " ", part).strip(" .:-")
        if not cleaned:
            continue
        if cleaned.upper() in {t.upper() for t in NULL_TOKENS}:
            continue
        out.append(cleaned.title())
    return out


def classify_text_themes(text: str, patterns: list[tuple[str, str]], default_label: str) -> list[str]:
    if not text:
        return []
    themes = [label for label, pattern in patterns if re.search(pattern, text, flags=re.IGNORECASE)]
    if themes:
        return themes
    return [default_label]


def is_negative_or_empty_text(text: str, *, extra_patterns: list[str] | None = None) -> bool:
    if not text:
        return True
    patterns = list(NEGATION_PATTERNS)
    if extra_patterns:
        patterns.extend(extra_patterns)
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in patterns)


def split_text_fragments(text: str) -> list[str]:
    if not text:
        return []
    fragments = []
    for part in TEXT_SPLIT_RE.split(text):
        cleaned = normalize_fragment_text(part)
        if not cleaned or cleaned in STOPWORDS:
            continue
        if len(cleaned) <= 2:
            continue
        fragments.append(cleaned)
    return fragments


def deduplicate_preserve_order(values: list[str]) -> list[str]:
    return list(dict.fromkeys(values))


def parse_bool_env(name: str, default: bool = False) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


def should_show_progress() -> bool:
    return parse_bool_env("QUAL_SHOW_PROGRESS", default=True)


def classify_fragment_with_explicit_rule(fragment: str) -> str | None:
    matches = [
        label
        for label, pattern in EXPLICIT_THEME_PATTERNS
        if re.search(pattern, fragment, flags=re.IGNORECASE)
    ]
    matches = deduplicate_preserve_order(matches)
    if len(matches) == 1:
        return matches[0]
    return None


def is_non_informative_text(text: str) -> bool:
    if not text:
        return True
    return any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in NON_INFORMATIVE_EXACT_PATTERNS)


def canonicalize_token(token: str) -> str:
    if not token:
        return token
    if token in TOKEN_NORMALIZATION_MAP:
        return TOKEN_NORMALIZATION_MAP[token]
    if any(token.endswith(ending) for ending in INVARIABLE_TOKEN_ENDINGS):
        return token
    return token


def canonicalize_fragment_tokens(text: str) -> str:
    for source, target in PHRASE_NORMALIZATION_MAP.items():
        text = re.sub(rf"\b{re.escape(source)}\b", target, text)
    tokens = [canonicalize_token(token) for token in text.split() if token]
    return " ".join(tokens).strip()


def normalize_fragment_text(fragment: str) -> str:
    cleaned = normalize_free_text(fragment)
    if not cleaned:
        return ""

    if is_non_informative_text(cleaned):
        return ""

    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .,:;-")

    changed = True
    while changed and cleaned:
        previous = cleaned
        for pattern in FRAGMENT_LEADING_PATTERNS:
            cleaned = re.sub(pattern, "", cleaned, flags=re.IGNORECASE)
        for pattern in LEADING_FILLER_PATTERNS:
            cleaned = re.sub(pattern, "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s+", " ", cleaned).strip(" .,:;-")
        changed = cleaned != previous

    for pattern in FRAGMENT_TRAILING_PATTERNS:
        cleaned = re.sub(pattern, "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" .,:;-")

    tokens = [token for token in cleaned.split() if token]
    while tokens and tokens[0] in GENERIC_FRAGMENT_TOKENS:
        tokens.pop(0)
    while tokens and tokens[-1] in GENERIC_FRAGMENT_TOKENS:
        tokens.pop()

    cleaned = canonicalize_fragment_tokens(" ".join(tokens).strip())
    if is_non_informative_text(cleaned):
        return ""
    return cleaned


@lru_cache(maxsize=1)
def get_zero_shot_runtime() -> tuple[object, str, str]:
    try:
        import torch
        from transformers import pipeline
    except ImportError as exc:
        raise RuntimeError(
            "Per la classificazione ML di apprezzamenti e critiche serve il pacchetto 'transformers' "
            "(ed il relativo backend PyTorch)."
        ) from exc

    HF_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    local_only = parse_bool_env("QUAL_SENTIMENT_LOCAL_ONLY", default=False)
    preferred_device = os.environ.get("QUAL_SENTIMENT_DEVICE", "auto").strip().lower()
    use_mps = preferred_device in {"auto", "mps"} and torch.backends.mps.is_built() and torch.backends.mps.is_available()
    if preferred_device == "cpu":
        runtime_device = "cpu"
    elif preferred_device == "mps" and not use_mps:
        runtime_device = "cpu"
    else:
        runtime_device = "mps" if use_mps else "cpu"

    pipeline_kwargs: dict[str, object] = {
        "cache_dir": str(HF_CACHE_DIR),
        "local_files_only": local_only,
        "device": runtime_device,
    }
    if runtime_device == "mps":
        pipeline_kwargs["torch_dtype"] = torch.float16

    last_error: Exception | None = None

    for model_name in deduplicate_preserve_order(ZERO_SHOT_FALLBACK_MODELS):
        try:
            classifier = pipeline(
                "zero-shot-classification",
                model=model_name,
                **pipeline_kwargs,
            )
            return classifier, model_name, runtime_device
        except Exception as exc:  # pragma: no cover - depends on runtime/model availability
            last_error = exc

    raise RuntimeError(
        "Impossibile inizializzare il classificatore Hugging Face per apprezzamenti/critiche. "
        f"Modelli tentati: {', '.join(deduplicate_preserve_order(ZERO_SHOT_FALLBACK_MODELS))}. "
        f"Cache usata: {HF_CACHE_DIR}. Device richiesto: {runtime_device}. Ultimo errore: {last_error}"
    )


def classify_fragments_with_ml(
    fragments: list[str],
    *,
    candidate_labels: list[str],
    threshold: float = DEFAULT_ZERO_SHOT_THRESHOLD,
    max_labels: int = DEFAULT_ZERO_SHOT_MAX_LABELS,
) -> list[list[dict[str, object]]]:
    if not fragments:
        return []

    final_rows: list[list[dict[str, object]] | None] = [None] * len(fragments)
    unresolved_fragments: list[str] = []
    unresolved_positions: list[int] = []

    for idx, fragment in enumerate(fragments):
        explicit_label = classify_fragment_with_explicit_rule(fragment)
        if explicit_label is not None:
            final_rows[idx] = [
                {
                    "frammento": fragment,
                    "tema": explicit_label,
                    "etichetta_modello": explicit_label,
                    "score_modello": 1.0,
                    "rank_modello": 1,
                    "tema_top1_modello": explicit_label,
                    "etichetta_top1_modello": explicit_label,
                    "score_top1_modello": 1.0,
                    "classifica_modello_top3": f"{explicit_label}=1.0000",
                    "metodo_classificazione": "regola_esplicita_o_ml_zero_shot",
                    "modello_hf": "regola_esplicita",
                    "device_inferenza": "n/a",
                }
            ]
        else:
            unresolved_fragments.append(fragment)
            unresolved_positions.append(idx)

    if not unresolved_fragments:
        return [rows if rows is not None else [] for rows in final_rows]

    classifier, model_name, runtime_device = get_zero_shot_runtime()
    raw_result = classifier(
        unresolved_fragments,
        candidate_labels=candidate_labels,
        multi_label=True,
        hypothesis_template=ZERO_SHOT_HYPOTHESIS_TEMPLATE,
        batch_size=DEFAULT_ZERO_SHOT_BATCH_SIZE,
    )
    if isinstance(raw_result, dict):
        raw_results = [raw_result]
    else:
        raw_results = list(raw_result)

    for position, fragment, result in zip(unresolved_positions, unresolved_fragments, raw_results):
        ranked = [
            {
                "tema": SENTIMENT_THEME_LABEL_MAP.get(label, label),
                "etichetta_modello": label,
                "score_modello": round(float(score), 4),
            }
            for label, score in zip(result["labels"], result["scores"])
        ]
        selected = [item for item in ranked if item["score_modello"] >= threshold][:max_labels]
        if not selected and ranked:
            selected = [ranked[0]]

        ranking_preview = " | ".join(f"{item['tema']}={item['score_modello']:.4f}" for item in ranked[:3])
        top1 = ranked[0] if ranked else {"tema": UNDEFINED_LABEL, "etichetta_modello": UNDEFINED_LABEL, "score_modello": 0.0}

        rows: list[dict[str, object]] = []
        for idx, item in enumerate(selected, start=1):
            rows.append(
                {
                    "frammento": fragment,
                    "tema": item["tema"],
                    "etichetta_modello": item["etichetta_modello"],
                    "score_modello": item["score_modello"],
                    "rank_modello": idx,
                    "tema_top1_modello": top1["tema"],
                    "etichetta_top1_modello": top1["etichetta_modello"],
                    "score_top1_modello": top1["score_modello"],
                    "classifica_modello_top3": ranking_preview,
                    "metodo_classificazione": "regola_esplicita_o_ml_zero_shot",
                    "modello_hf": model_name,
                    "device_inferenza": runtime_device,
                }
            )
        final_rows[position] = rows

    return [rows if rows is not None else [] for rows in final_rows]


def classify_text_fragments_ml(
    text: str,
    *,
    candidate_labels: list[str],
    extra_negation_patterns: list[str] | None = None,
    threshold: float = DEFAULT_ZERO_SHOT_THRESHOLD,
    max_labels: int = DEFAULT_ZERO_SHOT_MAX_LABELS,
) -> tuple[list[dict[str, object]], list[str], str]:
    normalized = normalize_free_text(text)
    if is_negative_or_empty_text(normalized, extra_patterns=extra_negation_patterns):
        return [], [], normalized

    # La multi-label e gestita a livello di risposta tramite piu frammenti distinti:
    # ogni frammento riceve una sola categoria prevalente, mentre una risposta come
    # "spiagge e cibo" produce due temi diversi perche viene spezzata in due unita.
    fragments = split_text_fragments(normalized)
    if not fragments:
        fragments = [normalized]

    all_themes: list[str] = []
    valid_fragments = [
        fragment
        for fragment in fragments
        if not is_negative_or_empty_text(fragment, extra_patterns=extra_negation_patterns)
    ]
    if not valid_fragments:
        return [], [], normalized

    batched_matches = classify_fragments_with_ml(
        valid_fragments,
        candidate_labels=candidate_labels,
        threshold=threshold,
        max_labels=max_labels,
    )
    audit_rows: list[dict[str, object]] = []
    for matches in batched_matches:
        audit_rows.extend(matches)
        all_themes.extend(item["tema"] for item in matches)

    return audit_rows, deduplicate_preserve_order(all_themes), normalized


def classify_text_fragments(
    text: str,
    *,
    patterns: list[tuple[str, str]],
    default_label: str,
    extra_negation_patterns: list[str] | None = None,
) -> tuple[list[dict[str, str]], list[str], str]:
    normalized = normalize_free_text(text)
    if is_negative_or_empty_text(normalized, extra_patterns=extra_negation_patterns):
        return [], [], normalized

    fragments = split_text_fragments(normalized)
    if not fragments:
        fragments = [normalized]

    audit_rows: list[dict[str, str]] = []
    all_themes: list[str] = []

    for fragment in fragments:
        if is_negative_or_empty_text(fragment, extra_patterns=extra_negation_patterns):
            continue
        matches: list[tuple[str, str]] = []
        for label, pattern in patterns:
            match = re.search(pattern, fragment, flags=re.IGNORECASE)
            if match:
                matches.append((label, match.group(0)))
        if not matches:
            matches = [(default_label, "fallback")]
        for label, keyword in matches:
            audit_rows.append({"frammento": fragment, "tema": label, "keyword_attivante": keyword})
            all_themes.append(label)

    return audit_rows, deduplicate_preserve_order(all_themes), normalized


def build_open_text_outputs(
    df: pd.DataFrame,
    *,
    text_col: str,
    group_cols: list[str],
    patterns: list[tuple[str, str]],
    default_label: str,
    tema_col_name: str,
    top_n: int,
    audit_sheet_name: str,
    summary_sheet_name: str,
    fragments_sheet_name: str,
    extra_audit_cols: list[str] | None = None,
    extra_negation_patterns: list[str] | None = None,
) -> dict[str, pd.DataFrame]:
    extra_audit_cols = extra_audit_cols or []
    audit_rows: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    fragment_rows: list[dict[str, object]] = []

    for _, row in df.iterrows():
        raw_text = prettify_value(row[text_col])
        audit_fragments, response_themes, normalized = classify_text_fragments(
            str(row[text_col]),
            patterns=patterns,
            default_label=default_label,
            extra_negation_patterns=extra_negation_patterns,
        )
        if not normalized:
            continue

        base_audit = {
            ID_COL: row[ID_COL],
            "testo_originale": raw_text,
            "testo_normalizzato": normalized,
            "tema_primario": response_themes[0] if response_themes else UNDEFINED_LABEL,
            "temi_tutti": " | ".join(response_themes) if response_themes else UNDEFINED_LABEL,
        }
        for col in group_cols + extra_audit_cols:
            base_audit[col] = row[col]

        if not audit_fragments:
            audit_rows.append({**base_audit, "frammento": "", "tema": UNDEFINED_LABEL, "keyword_attivante": ""})
            continue

        for item in audit_fragments:
            audit_rows.append({**base_audit, **item})

        for theme in response_themes:
            base = {col: row[col] for col in group_cols}
            base.update({tema_col_name: theme, ID_COL: row[ID_COL], COMPONENTS_COL: row[COMPONENTS_COL]})
            summary_rows.append(base)

        seen_themes: set[str] = set()
        for item in audit_fragments:
            theme = item["tema"]
            if theme in seen_themes:
                continue
            seen_themes.add(theme)
            base = {col: row[col] for col in group_cols}
            base.update(
                {
                    "frammento": item["frammento"],
                    tema_col_name: theme,
                    ID_COL: row[ID_COL],
                    COMPONENTS_COL: row[COMPONENTS_COL],
                }
            )
            fragment_rows.append(base)

    audit_df = pd.DataFrame(audit_rows)
    if audit_df.empty:
        audit_df = pd.DataFrame(
            columns=[ID_COL, *group_cols, *extra_audit_cols, "testo_originale", "testo_normalizzato", "frammento", "tema", "keyword_attivante", "tema_primario", "temi_tutti"]
        )
    audit_df = compact_rule_based_audit_df(audit_df, group_cols=group_cols, extra_audit_cols=extra_audit_cols)

    summary_df = pd.DataFrame(summary_rows)
    if summary_df.empty:
        summary_out = pd.DataFrame(columns=[*group_cols, tema_col_name, "questionari", "componenti", "pct_su_gruppo"])
    else:
        summary_out = (
            summary_df.groupby([*group_cols, tema_col_name], dropna=False, as_index=False)
            .agg(questionari=(ID_COL, "nunique"), componenti=(COMPONENTS_COL, "sum"))
            .sort_values(
                by=[*group_cols, "questionari", tema_col_name],
                ascending=[True] * len(group_cols) + [False, True],
                kind="mergesort",
            )
        )
        summary_out = add_share_within_group(summary_out, group_cols)

    fragment_df = pd.DataFrame(fragment_rows)
    if fragment_df.empty:
        fragment_out = pd.DataFrame(columns=[*group_cols, "frammento", tema_col_name, "questionari", "componenti", "pct_su_gruppo"])
    else:
        fragment_out = (
            fragment_df.groupby([*group_cols, "frammento", tema_col_name], dropna=False, as_index=False)
            .agg(questionari=(ID_COL, "nunique"), componenti=(COMPONENTS_COL, "sum"))
            .sort_values(
                by=[*group_cols, "questionari", "frammento"],
                ascending=[True] * len(group_cols) + [False, True],
                kind="mergesort",
            )
        )
        fragment_out = fragment_out.groupby(group_cols, dropna=False, group_keys=False).head(top_n).reset_index(drop=True)
        fragment_out = add_share_within_group(fragment_out, group_cols)

    return {
        audit_sheet_name: audit_df,
        summary_sheet_name: summary_out,
        fragments_sheet_name: fragment_out,
    }


def build_ml_open_text_outputs(
    df: pd.DataFrame,
    *,
    text_col: str,
    group_cols: list[str],
    candidate_labels: list[str],
    tema_col_name: str,
    top_n: int,
    audit_sheet_name: str,
    summary_sheet_name: str,
    fragments_sheet_name: str,
    extra_audit_cols: list[str] | None = None,
    extra_negation_patterns: list[str] | None = None,
    progress_label: str | None = None,
) -> dict[str, pd.DataFrame]:
    extra_audit_cols = extra_audit_cols or []
    audit_rows: list[dict[str, object]] = []
    summary_rows: list[dict[str, object]] = []
    fragment_rows: list[dict[str, object]] = []
    row_iterator = df.iterrows()
    if should_show_progress():
        row_iterator = tqdm(
            row_iterator,
            total=len(df),
            desc=progress_label or f"Classificazione ML {text_col}",
            unit="risposta",
            dynamic_ncols=True,
        )

    for _, row in row_iterator:
        raw_text = prettify_value(row[text_col])
        audit_fragments, response_themes, normalized = classify_text_fragments_ml(
            str(row[text_col]),
            candidate_labels=candidate_labels,
            extra_negation_patterns=extra_negation_patterns,
        )
        if not normalized:
            continue

        base_audit = {
            ID_COL: row[ID_COL],
            "testo_originale": raw_text,
            "testo_normalizzato": normalized,
            "tema_primario": response_themes[0] if response_themes else UNDEFINED_LABEL,
            "temi_tutti": " | ".join(response_themes) if response_themes else UNDEFINED_LABEL,
            "metodo_classificazione": ZERO_SHOT_METHOD_LABEL,
        }
        for col in group_cols + extra_audit_cols:
            base_audit[col] = row[col]

        if not audit_fragments:
            audit_rows.append(
                {
                    **base_audit,
                    "frammento": "",
                    "tema": UNDEFINED_LABEL,
                    "score_modello": "",
                    "rank_modello": "",
                    "tema_top1_modello": "",
                    "score_top1_modello": "",
                    "classifica_modello_top3": "",
                    "modello_hf": "",
                }
            )
            continue

        for item in audit_fragments:
            audit_rows.append({**base_audit, **item})

        for theme in response_themes:
            base = {col: row[col] for col in group_cols}
            base.update(
                {
                    tema_col_name: theme,
                    ID_COL: row[ID_COL],
                    COMPONENTS_COL: row[COMPONENTS_COL],
                }
            )
            summary_rows.append(base)

        best_fragment_by_theme: dict[str, tuple[str, float]] = {}
        for item in audit_fragments:
            tema = str(item["tema"])
            frammento = str(item["frammento"])
            score = float(item["score_modello"])
            current = best_fragment_by_theme.get(tema)
            if current is None or score > current[1]:
                best_fragment_by_theme[tema] = (frammento, score)

        for tema, (frammento, _score) in best_fragment_by_theme.items():
            base = {col: row[col] for col in group_cols}
            base.update(
                {
                    "frammento": frammento,
                    tema_col_name: tema,
                    ID_COL: row[ID_COL],
                    COMPONENTS_COL: row[COMPONENTS_COL],
                }
            )
            fragment_rows.append(base)

    audit_df = pd.DataFrame(audit_rows)
    if audit_df.empty:
        audit_df = pd.DataFrame(
            columns=[
                ID_COL,
                *group_cols,
                *extra_audit_cols,
                "testo_originale",
                "testo_normalizzato",
                "frammento",
                "tema",
                "etichetta_modello",
                "score_modello",
                "rank_modello",
                "tema_top1_modello",
                "etichetta_top1_modello",
                "score_top1_modello",
                "classifica_modello_top3",
                "tema_primario",
                "temi_tutti",
                "metodo_classificazione",
                "modello_hf",
                "device_inferenza",
            ]
        )
    audit_df = compact_ml_audit_df(audit_df, group_cols=group_cols, extra_audit_cols=extra_audit_cols)

    summary_df = pd.DataFrame(summary_rows)
    if summary_df.empty:
        summary_out = pd.DataFrame(
            columns=[
                *group_cols,
                tema_col_name,
                "questionari",
                "componenti",
                "pct_su_gruppo",
            ]
        )
    else:
        summary_out = (
            summary_df.groupby([*group_cols, tema_col_name], dropna=False, as_index=False)
            .agg(questionari=(ID_COL, "nunique"), componenti=(COMPONENTS_COL, "sum"))
            .sort_values(
                by=[*group_cols, "questionari", "componenti", tema_col_name],
                ascending=[True] * len(group_cols) + [False, False, True],
                kind="mergesort",
            )
        )
        summary_out = add_share_within_group(summary_out, group_cols)

    fragment_df = pd.DataFrame(fragment_rows)
    if fragment_df.empty:
        fragment_out = pd.DataFrame(
            columns=[
                *group_cols,
                "frammento",
                tema_col_name,
                "questionari",
                "componenti",
                "pct_su_gruppo",
            ]
        )
    else:
        fragment_out = (
            fragment_df.groupby([*group_cols, "frammento", tema_col_name], dropna=False, as_index=False)
            .agg(questionari=(ID_COL, "nunique"), componenti=(COMPONENTS_COL, "sum"))
            .sort_values(
                by=[*group_cols, "questionari", "componenti", "frammento"],
                ascending=[True] * len(group_cols) + [False, False, True],
                kind="mergesort",
            )
        )
        fragment_out = fragment_out.groupby(group_cols, dropna=False, group_keys=False).head(top_n).reset_index(drop=True)
        fragment_out = add_share_within_group(fragment_out, group_cols)

    return {
        audit_sheet_name: audit_df,
        summary_sheet_name: summary_out,
        fragments_sheet_name: fragment_out,
    }


def compact_rule_based_audit_df(
    audit_df: pd.DataFrame,
    *,
    group_cols: list[str],
    extra_audit_cols: list[str],
) -> pd.DataFrame:
    desired_columns = [
        ID_COL,
        *group_cols,
        *extra_audit_cols,
        "testo_originale",
        "frammento",
        "tema",
        "keyword_attivante",
    ]
    available_columns = [col for col in desired_columns if col in audit_df.columns]
    out = audit_df.loc[:, available_columns].copy()
    if ID_COL in out.columns:
        out = out.sort_values(by=[ID_COL], kind="mergesort").reset_index(drop=True)
    return out


def compact_ml_audit_df(
    audit_df: pd.DataFrame,
    *,
    group_cols: list[str],
    extra_audit_cols: list[str],
) -> pd.DataFrame:
    desired_columns = [
        ID_COL,
        *group_cols,
        *extra_audit_cols,
        "testo_originale",
        "frammento",
        "tema",
        "score_modello",
        "classifica_modello_top3",
    ]
    available_columns = [col for col in desired_columns if col in audit_df.columns]
    out = audit_df.loc[:, available_columns].copy()
    if ID_COL in out.columns:
        out = out.sort_values(by=[ID_COL], kind="mergesort").reset_index(drop=True)
    return out


def build_text_theme_table(
    df: pd.DataFrame,
    *,
    text_col: str,
    group_cols: list[str],
    patterns: list[tuple[str, str]],
    default_label: str,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        text = normalize_free_text(row[text_col])
        if not text:
            continue
        themes = classify_text_themes(text, patterns, default_label)
        for theme in themes:
            base = {col: row[col] for col in group_cols}
            base.update({"tema": theme, "ID": row[ID_COL]})
            rows.append(base)

    if not rows:
        return pd.DataFrame(columns=[*group_cols, "tema", "questionari"])

    out = pd.DataFrame(rows)
    return (
        out.groupby([*group_cols, "tema"], dropna=False, as_index=False)
        .agg(questionari=("ID", "nunique"))
        .sort_values(by=[*group_cols, "questionari", "tema"], ascending=[True] * len(group_cols) + [False, True], kind="mergesort")
    )


def build_text_fragment_table(
    df: pd.DataFrame,
    *,
    text_col: str,
    group_cols: list[str],
    top_n: int,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        text = normalize_free_text(row[text_col])
        if not text:
            continue
        fragments = split_text_fragments(text)
        for fragment in fragments:
            base = {col: row[col] for col in group_cols}
            base.update({"frammento": fragment, "ID": row[ID_COL]})
            rows.append(base)

    if not rows:
        return pd.DataFrame(columns=[*group_cols, "frammento", "questionari"])

    out = (
        pd.DataFrame(rows)
        .groupby([*group_cols, "frammento"], dropna=False, as_index=False)
        .agg(questionari=("ID", "nunique"))
    )
    out = out.sort_values(
        by=[*group_cols, "questionari", "frammento"],
        ascending=[True] * len(group_cols) + [False, True],
        kind="mergesort",
    )
    return out.groupby(group_cols, dropna=False, group_keys=False).head(top_n).reset_index(drop=True)


def add_share_within_group(df: pd.DataFrame, group_cols: list[str], value_col: str = "questionari") -> pd.DataFrame:
    if df.empty:
        return df.copy()
    out = df.copy()
    totals = out.groupby(group_cols, dropna=False)[value_col].transform("sum")
    out["pct_su_gruppo"] = (out[value_col] / totals * 100.0).round(2)
    return out


def build_distribution(
    df: pd.DataFrame,
    *,
    group_cols: list[str],
    value_col: str,
    include_components: bool = True,
) -> pd.DataFrame:
    grouped = (
        df.groupby([*group_cols, value_col], dropna=False, as_index=False)
        .agg(
            questionari=(ID_COL, "nunique"),
            componenti=(COMPONENTS_COL, "sum"),
        )
    )
    if not include_components:
        grouped = grouped.drop(columns=["componenti"])
    return grouped.sort_values(
        by=[*group_cols, "questionari", value_col],
        ascending=[True] * len(group_cols) + [False, True],
        kind="mergesort",
    )


def build_top_n_by_group(
    df: pd.DataFrame,
    *,
    group_cols: list[str],
    value_col: str,
    top_n: int,
) -> pd.DataFrame:
    out = build_distribution(df, group_cols=group_cols, value_col=value_col, include_components=True)
    out = out.groupby(group_cols, dropna=False, group_keys=False).head(top_n).reset_index(drop=True)
    return add_share_within_group(out, group_cols)


def build_ranked_theme_priorities(
    df: pd.DataFrame,
    *,
    group_cols: list[str],
    theme_col: str = "tema",
    top_n: int | None = None,
) -> pd.DataFrame:
    if df.empty:
        cols = [*group_cols, theme_col, "questionari", "componenti", "pct_su_gruppo", "rank_nel_gruppo"]
        return pd.DataFrame(columns=cols)

    out = df.sort_values(
        by=[*group_cols, "questionari", "componenti", theme_col],
        ascending=[True] * len(group_cols) + [False, False, True],
        kind="mergesort",
    ).copy()
    out["rank_nel_gruppo"] = out.groupby(group_cols).cumcount() + 1
    if top_n is not None:
        out = out[out["rank_nel_gruppo"] <= top_n].copy()
    return out.reset_index(drop=True)


def compute_text_quality_stats(series: pd.Series) -> dict[str, int]:
    total = int(len(series))
    valid = 0
    discarded_empty_or_null = 0
    discarded_non_informative = 0

    for value in series.tolist():
        normalized = normalize_free_text(value)
        if not normalized:
            discarded_empty_or_null += 1
            continue
        if is_non_informative_text(normalized):
            discarded_non_informative += 1
            continue
        if not split_text_fragments(normalized):
            discarded_non_informative += 1
            continue
        valid += 1

    return {
        "totale_risposte": total,
        "risposte_valide": valid,
        "scartate_vuote_nulle": discarded_empty_or_null,
        "scartate_non_informative": discarded_non_informative,
    }


def prepare_dataframe(input_file: Path, sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(input_file, sheet_name=sheet_name)
    validate_columns(raw)

    df = raw.copy()
    for col in REQUIRED_COLUMNS:
        if col == COMPONENTS_COL:
            continue
        df[col] = normalize_text_series(df[col])

    df[COMPONENTS_COL] = pd.to_numeric(df[COMPONENTS_COL], errors="coerce").fillna(0)
    df[DURATION_COL] = pd.to_numeric(df[DURATION_COL], errors="coerce")
    df["macro_provenienza"] = build_macro_provenienza(df[STATE_COL])
    df["provenienza_associata"] = build_provenienza_label(df)
    df["web_flag"] = normalize_yes_no(df[WEB_COL])
    df["giudizio_norm"] = normalize_judgment(df[JUDGMENT_COL])
    df["giudizio_score"] = df["giudizio_norm"].map(JUDGMENT_SCORE_MAP)

    destination_metrics = build_destination_metrics(df[LOCATIONS_COL])
    df = pd.concat([df, destination_metrics], axis=1)
    df["destinazione_prevalente"] = (
        df["destinazione_prevalente"]
        .astype("string")
        .fillna(UNDEFINED_LABEL)
        .replace({NULL_VALUE: UNDEFINED_LABEL, "": UNDEFINED_LABEL})
    )

    return df


def build_outputs(df: pd.DataFrame, sheet_name: str) -> dict[str, pd.DataFrame]:
    outputs: dict[str, pd.DataFrame] = {}
    _, runtime_model_name, runtime_device = get_zero_shot_runtime()
    strengths_stats = compute_text_quality_stats(df[STRENGTHS_COL])
    weaknesses_stats = compute_text_quality_stats(df[WEAKNESSES_COL])

    meta_rows = [
        {"chiave": "sheet_campione", "valore": sheet_name},
        {"chiave": "questionari_totali", "valore": int(df[ID_COL].nunique())},
        {"chiave": "componenti_totali", "valore": int(df[COMPONENTS_COL].sum())},
        {"chiave": "italiani_questionari", "valore": int(df["macro_provenienza"].eq("ITALIANI").sum())},
        {"chiave": "stranieri_questionari", "valore": int(df["macro_provenienza"].eq("STRANIERI").sum())},
        {"chiave": "web_specifica_metodo", "valore": "regole_testuali_keyword_non_ml"},
        {"chiave": "classificazione_sentiment_metodo", "valore": ZERO_SHOT_METHOD_LABEL},
        {"chiave": "classificazione_sentiment_modello_preferito", "valore": DEFAULT_ZERO_SHOT_MODEL},
        {"chiave": "classificazione_sentiment_modello_effettivo", "valore": runtime_model_name},
        {"chiave": "classificazione_sentiment_cache_dir", "valore": str(HF_CACHE_DIR)},
        {"chiave": "classificazione_sentiment_device", "valore": runtime_device},
        {"chiave": "classificazione_sentiment_batch_size", "valore": DEFAULT_ZERO_SHOT_BATCH_SIZE},
        {"chiave": "classificazione_sentiment_descrizioni_categorie", "valore": "SI"},
        {"chiave": "apprezzamenti_totale_risposte", "valore": strengths_stats["totale_risposte"]},
        {"chiave": "apprezzamenti_risposte_valide", "valore": strengths_stats["risposte_valide"]},
        {"chiave": "apprezzamenti_scartate_vuote_nulle", "valore": strengths_stats["scartate_vuote_nulle"]},
        {"chiave": "apprezzamenti_scartate_non_informative", "valore": strengths_stats["scartate_non_informative"]},
        {"chiave": "miglioramenti_totale_risposte", "valore": weaknesses_stats["totale_risposte"]},
        {"chiave": "miglioramenti_risposte_valide", "valore": weaknesses_stats["risposte_valide"]},
        {"chiave": "miglioramenti_scartate_vuote_nulle", "valore": weaknesses_stats["scartate_vuote_nulle"]},
        {"chiave": "miglioramenti_scartate_non_informative", "valore": weaknesses_stats["scartate_non_informative"]},
    ]
    outputs["meta"] = pd.DataFrame(meta_rows)

    web_usage = build_distribution(df, group_cols=["macro_provenienza"], value_col="web_flag", include_components=True)
    outputs["web_prenotazione"] = add_share_within_group(web_usage, ["macro_provenienza"])

    web_yes = df[df["web_flag"] == "SI"].copy()
    outputs.update(
        build_open_text_outputs(
            web_yes,
            text_col=WEB_DETAIL_COL,
            group_cols=["macro_provenienza"],
            patterns=WEB_THEME_PATTERNS,
            default_label="ALTRO/NON CLASSIFICATO",
            tema_col_name="cosa_prenotata",
            top_n=20,
            audit_sheet_name="audit_web",
            summary_sheet_name="web_cosa_prenotata",
            fragments_sheet_name="web_dettagli_top20",
            extra_negation_patterns=WEB_NEGATION_PATTERNS,
        )
    )

    travel_type = build_distribution(df, group_cols=["macro_provenienza"], value_col="tipo_turismo", include_components=True)
    outputs["stanziale_itinerante"] = add_share_within_group(travel_type, ["macro_provenienza"])

    arrivals = build_distribution(
        df,
        group_cols=["macro_provenienza", "destinazione_prevalente"],
        value_col=ARRIVAL_COL,
        include_components=True,
    )
    outputs["arrivi_x_destinazione"] = add_share_within_group(arrivals, ["macro_provenienza", "destinazione_prevalente"])

    lodging = build_distribution(df, group_cols=["macro_provenienza"], value_col=LODGING_COL, include_components=True)
    outputs["alloggio_prevalente"] = add_share_within_group(lodging, ["macro_provenienza"])

    reason_lodging = build_distribution(
        df,
        group_cols=["macro_provenienza", PRIMARY_REASON_COL],
        value_col=LODGING_COL,
        include_components=True,
    )
    outputs["motivazione_x_alloggio"] = add_share_within_group(reason_lodging, ["macro_provenienza", PRIMARY_REASON_COL])

    outputs["dest_top10_provenienze"] = build_top_n_by_group(
        df,
        group_cols=["macro_provenienza", "destinazione_prevalente"],
        value_col="provenienza_associata",
        top_n=10,
    )
    outputs["dest_top_motivazioni"] = build_top_n_by_group(
        df,
        group_cols=["macro_provenienza", "destinazione_prevalente"],
        value_col=PRIMARY_REASON_COL,
        top_n=10,
    )

    spend_assoc = (
        df.groupby(
            [FUTURE_SPEND_COL, "macro_provenienza", "provenienza_associata", PRIMARY_REASON_COL, "giudizio_norm"],
            dropna=False,
            as_index=False,
        )
        .agg(questionari=(ID_COL, "nunique"), componenti=(COMPONENTS_COL, "sum"))
        .sort_values(
            by=[FUTURE_SPEND_COL, "macro_provenienza", "questionari", "provenienza_associata"],
            ascending=[True, True, False, True],
            kind="mergesort",
        )
    )
    outputs["spesa_futura_assoc"] = spend_assoc

    top20_destinations = (
        df.groupby("destinazione_prevalente", as_index=False)
        .agg(questionari=(ID_COL, "nunique"))
        .sort_values(by=["questionari", "destinazione_prevalente"], ascending=[False, True], kind="mergesort")
        .head(20)["destinazione_prevalente"]
        .tolist()
    )
    top20_df = df[df["destinazione_prevalente"].isin(top20_destinations)].copy()
    top20_reason = build_distribution(
        top20_df,
        group_cols=["destinazione_prevalente"],
        value_col=PRIMARY_REASON_COL,
        include_components=True,
    )
    outputs["top20_dest_x_motivaz"] = add_share_within_group(top20_reason, ["destinazione_prevalente"])

    reason_secondary = build_distribution(
        df,
        group_cols=[PRIMARY_REASON_COL],
        value_col=SECONDARY_REASON_COL,
        include_components=True,
    )
    outputs["motivaz_prim_second"] = add_share_within_group(reason_secondary, [PRIMARY_REASON_COL])

    judgment_dist = build_distribution(
        df,
        group_cols=["macro_provenienza", AGE_COL],
        value_col="giudizio_norm",
        include_components=True,
    )
    outputs["giudizio_distrib"] = add_share_within_group(judgment_dist, ["macro_provenienza", AGE_COL])

    top15_destinations = (
        df.groupby("destinazione_prevalente", as_index=False)
        .agg(questionari=(ID_COL, "nunique"))
        .sort_values(by=["questionari", "destinazione_prevalente"], ascending=[False, True], kind="mergesort")
        .head(15)["destinazione_prevalente"]
        .tolist()
    )
    top15_df = df[df["destinazione_prevalente"].isin(top15_destinations) & df["giudizio_score"].notna()].copy()
    judgment_mean = (
        top15_df.groupby(["destinazione_prevalente", "macro_provenienza"], as_index=False)
        .agg(
            questionari=(ID_COL, "nunique"),
            componenti=(COMPONENTS_COL, "sum"),
            media_giudizio=("giudizio_score", "mean"),
        )
    )
    overall_mean = (
        top15_df.groupby("destinazione_prevalente", as_index=False)
        .agg(
            questionari=(ID_COL, "nunique"),
            componenti=(COMPONENTS_COL, "sum"),
            media_giudizio=("giudizio_score", "mean"),
        )
    )
    overall_mean.insert(1, "macro_provenienza", "TOTALE")
    outputs["giudizio_top15_dest"] = (
        pd.concat([overall_mean, judgment_mean], ignore_index=True)
        .sort_values(
            by=["destinazione_prevalente", "macro_provenienza"],
            ascending=[True, True],
            kind="mergesort",
        )
    )

    outputs.update(
        build_ml_open_text_outputs(
            df,
            text_col=STRENGTHS_COL,
            group_cols=["macro_provenienza"],
            candidate_labels=SENTIMENT_THEME_LABELS,
            tema_col_name="tema",
            top_n=20,
            audit_sheet_name="audit_apprezzamenti",
            summary_sheet_name="punti_forza_temi",
            fragments_sheet_name="punti_forza_top20",
            extra_audit_cols=["destinazione_prevalente", PRIMARY_REASON_COL],
            progress_label="Classificazione ML apprezzamenti",
        )
    )
    outputs["punti_forza_top_temi"] = build_ranked_theme_priorities(
        outputs["punti_forza_temi"],
        group_cols=["macro_provenienza"],
        top_n=10,
    )

    dissent_df = df[df["giudizio_norm"].isin(["SUFFICIENTE", "PESSIMO"])].copy()
    outputs.update(
        build_ml_open_text_outputs(
            dissent_df,
            text_col=WEAKNESSES_COL,
            group_cols=["macro_provenienza", "giudizio_norm"],
            candidate_labels=SENTIMENT_THEME_LABELS,
            tema_col_name="tema",
            top_n=20,
            audit_sheet_name="audit_dissenso",
            summary_sheet_name="dissenso_temi",
            fragments_sheet_name="dissenso_top20",
            extra_audit_cols=["destinazione_prevalente", PRIMARY_REASON_COL],
            progress_label="Classificazione ML dissenso",
        )
    )
    outputs["dissenso_priorita_intervento"] = build_ranked_theme_priorities(
        outputs["dissenso_temi"],
        group_cols=["macro_provenienza", "giudizio_norm"],
        top_n=10,
    )
    outputs["dissenso_dettaglio"] = dissent_df[
        [
            ID_COL,
            "macro_provenienza",
            AGE_COL,
            "provenienza_associata",
            "destinazione_prevalente",
            PRIMARY_REASON_COL,
            "giudizio_norm",
            WEAKNESSES_COL,
        ]
    ].sort_values(by=["macro_provenienza", "giudizio_norm", ID_COL], kind="mergesort")

    return outputs


def prepare_dataframe_1d(input_file: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(input_file, sheet_name=sheet_name)
    required = [STATE_COL, PRIMARY_REASON_COL, STRENGTHS_COL, COMPONENTS_COL]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Colonne mancanti nel foglio {sheet_name}: {missing}")

    out = df.copy()
    out[PRIMARY_REASON_COL] = normalize_text_series(out[PRIMARY_REASON_COL]).str.upper()
    out[STRENGTHS_COL] = normalize_text_series(out[STRENGTHS_COL])
    out["macro_provenienza"] = build_macro_provenienza(out[STATE_COL])
    out[COMPONENTS_COL] = pd.to_numeric(out[COMPONENTS_COL], errors="coerce").fillna(0)
    return out


def build_outputs_1d(df: pd.DataFrame, sheet_name: str) -> dict[str, pd.DataFrame]:
    outputs: dict[str, pd.DataFrame] = {}
    _, runtime_model_name, runtime_device = get_zero_shot_runtime()
    strengths_stats = compute_text_quality_stats(df[STRENGTHS_COL])
    meta_rows = [
        {"chiave": "sheet_campione", "valore": sheet_name},
        {"chiave": "questionari_totali", "valore": int(len(df))},
        {"chiave": "questionari_con_apprezzamenti", "valore": int(((df[PRIMARY_REASON_COL] != "") & (df[STRENGTHS_COL] != "")).sum())},
        {"chiave": "componenti_totali", "valore": int(df[COMPONENTS_COL].sum())},
        {"chiave": "classificazione_sentiment_metodo", "valore": ZERO_SHOT_METHOD_LABEL},
        {"chiave": "classificazione_sentiment_modello_preferito", "valore": DEFAULT_ZERO_SHOT_MODEL},
        {"chiave": "classificazione_sentiment_modello_effettivo", "valore": runtime_model_name},
        {"chiave": "classificazione_sentiment_device", "valore": runtime_device},
        {"chiave": "classificazione_sentiment_batch_size", "valore": DEFAULT_ZERO_SHOT_BATCH_SIZE},
        {"chiave": "classificazione_sentiment_descrizioni_categorie", "valore": "SI"},
        {"chiave": "apprezzamenti_totale_risposte", "valore": strengths_stats["totale_risposte"]},
        {"chiave": "apprezzamenti_risposte_valide", "valore": strengths_stats["risposte_valide"]},
        {"chiave": "apprezzamenti_scartate_vuote_nulle", "valore": strengths_stats["scartate_vuote_nulle"]},
        {"chiave": "apprezzamenti_scartate_non_informative", "valore": strengths_stats["scartate_non_informative"]},
    ]
    outputs["c1d_meta"] = pd.DataFrame(meta_rows)

    working = df[(df[PRIMARY_REASON_COL] != "") & (df[STRENGTHS_COL] != "")].copy()
    ml_outputs = build_ml_open_text_outputs(
        working,
        text_col=STRENGTHS_COL,
        group_cols=["macro_provenienza", PRIMARY_REASON_COL],
        candidate_labels=SENTIMENT_THEME_LABELS,
        tema_col_name="tema",
        top_n=20,
        audit_sheet_name="c1d_audit_apprezzamenti",
        summary_sheet_name="c1d_dettaglio_completo",
        fragments_sheet_name="c1d_frammenti_top20",
        progress_label="Classificazione ML apprezzamenti campione 1d",
    )
    outputs.update(ml_outputs)

    grouped = build_ranked_theme_priorities(
        outputs["c1d_dettaglio_completo"],
        group_cols=["macro_provenienza", PRIMARY_REASON_COL],
        top_n=None,
    )
    outputs["c1d_dettaglio_completo"] = grouped
    outputs["c1d_top_apprezzamenti"] = grouped[grouped["rank_nel_gruppo"] <= 10].copy()
    return outputs


REPORT_LAYOUT = [
    (
        "sintesi",
        "Analisi qualitative - sintesi",
        [
            ("meta", "Metadati del campione"),
            ("web_prenotazione", "Uso del web per la prenotazione"),
            ("stanziale_itinerante", "Turismo stanziale o itinerante"),
            ("arrivi_x_destinazione", "Arrivo in Sardegna per destinazione prevalente"),
            ("alloggio_prevalente", "Tipologia di alloggio prevalente"),
            ("motivazione_x_alloggio", "Tipologia di alloggio per motivazione principale"),
        ],
    ),
    (
        "destinazioni_motivazioni",
        "Analisi qualitative - destinazioni e motivazioni",
        [
            ("dest_top10_provenienze", "Top 10 provenienze associate per destinazione"),
            ("dest_top_motivazioni", "Top motivazioni principali per destinazione"),
            ("top20_dest_x_motivaz", "Top 20 destinazioni per motivazione principale"),
            ("motivaz_prim_second", "Motivazione principale per motivazione secondaria"),
            ("spesa_futura_assoc", "Associazioni con previsione di spesa futura"),
            ("giudizio_distrib", "Distribuzione giudizi per provenienza ed età"),
            ("giudizio_top15_dest", "Giudizio medio top 15 destinazioni"),
        ],
    ),
    (
        "testi_aperti",
        "Analisi qualitative - testi aperti",
        [
            ("web_cosa_prenotata", "Cosa viene prenotato via web"),
            ("web_dettagli_top20", "Top 20 frammenti web"),
            ("punti_forza_temi", "Temi dei punti di forza"),
            ("punti_forza_top_temi", "Top temi dei punti di forza per provenienza"),
            ("punti_forza_top20", "Top 20 frammenti punti di forza"),
            ("dissenso_temi", "Temi del dissenso"),
            ("dissenso_priorita_intervento", "Priorita di intervento per provenienza e giudizio"),
            ("dissenso_top20", "Top 20 frammenti del dissenso"),
            ("dissenso_dettaglio", "Dettaglio questionari con dissenso"),
        ],
    ),
    (
        "audit",
        "Analisi qualitative - audit classificazioni",
        [
            ("audit_web", "Audit classificazione web"),
            ("audit_apprezzamenti", "Audit classificazione apprezzamenti"),
            ("audit_dissenso", "Audit classificazione dissenso"),
        ],
    ),
    (
        "campione_1d",
        "Analisi qualitative - campione 1d",
        [
            ("c1d_meta", "Metadati campione 1d"),
            ("c1d_top_apprezzamenti", "Top 10 temi apprezzamenti per provenienza e motivazione"),
            ("c1d_dettaglio_completo", "Dettaglio completo temi apprezzamenti campione 1d"),
            ("c1d_frammenti_top20", "Top 20 frammenti apprezzamenti campione 1d"),
            ("c1d_audit_apprezzamenti", "Audit classificazione ML apprezzamenti campione 1d"),
        ],
    ),
]


def format_cell_value(value: object) -> object:
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        return round(value, 2)
    return value


def style_worksheet(ws) -> None:
    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    section_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    title_font = Font(bold=True, size=14, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True)

    max_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in (None, ""):
                continue

            if cell.row == 1 and cell.column == 1:
                cell.fill = title_fill
                cell.font = title_font
                cell.alignment = Alignment(horizontal="left", vertical="top")
            elif isinstance(cell.value, str) and cell.value.startswith("Tabella: "):
                cell.font = section_font
                cell.fill = section_fill
                cell.alignment = Alignment(vertical="top")
            elif cell.row > 1 and cell.column == 1 and isinstance(cell.value, str) and cell.value.startswith("Sezione: "):
                cell.font = section_font
                cell.alignment = Alignment(vertical="top")
            elif cell.font and cell.font.bold and cell.row > 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            text_length = len(str(cell.value))
            max_widths[cell.column] = min(max(max_widths.get(cell.column, 0), text_length + 2), 45)

    for col_idx, width in max_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, width)

    ws.freeze_panes = "A2"


def write_table(ws, start_row: int, title: str, df: pd.DataFrame) -> int:
    ws.cell(row=start_row, column=1, value=f"Tabella: {title}")
    start_row += 1

    if df.empty:
        ws.cell(row=start_row, column=1, value="Nessun dato disponibile")
        return start_row + 2

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
    start_row += 1

    for _, data_row in df.iterrows():
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=format_cell_value(data_row[header]))
            if header == "pct_su_gruppo" and cell.value != "":
                cell.number_format = "0.00"
        start_row += 1

    return start_row + 2


def write_structured_excel(
    outputs: dict[str, pd.DataFrame],
    output_file: Path,
    report_layout: list[tuple[str, str, list[tuple[str, str]]]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, sheet_title, table_specs in report_layout:
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.cell(row=1, column=1, value=sheet_title)
        row_idx = 3

        for key, table_title in table_specs:
            if key not in outputs:
                continue
            row_idx = write_table(ws, row_idx, table_title, outputs[key])

        style_worksheet(ws)

    wb.save(output_file)


def write_excel(outputs: dict[str, pd.DataFrame], output_file: Path) -> None:
    write_structured_excel(outputs, output_file, REPORT_LAYOUT)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Costruisce le analisi qualitative del Campione 1.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help=f"File input Excel (default: {DEFAULT_INPUT})")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help=f"Foglio da analizzare (default: {DEFAULT_SHEET})")
    parser.add_argument("--sheet-1d", default=DEFAULT_SHEET_1D, help=f"Foglio campione 1d da includere (default: {DEFAULT_SHEET_1D})")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help=f"File output Excel (default: {DEFAULT_OUTPUT})")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(f"File non trovato: {args.input}")

    df = prepare_dataframe(args.input, args.sheet)
    outputs = build_outputs(df, args.sheet)
    df_1d = prepare_dataframe_1d(args.input, args.sheet_1d)
    outputs.update(build_outputs_1d(df_1d, args.sheet_1d))
    write_excel(outputs, args.output)
    print(f"Analisi qualitative Campione 1 salvate in: {args.output}")


if __name__ == "__main__":
    main()
